#!/usr/bin/env python3
"""Generate both MHL + 7-MTL workbooks from templates and JSON input.

Design goals:
- Keep workbook formulas intact.
- Write only non-formula input cells.
- Support normalized fact inputs (sales by month) plus explicit patches for
  any remaining cells/sheets that need custom logic.

Usage:
    python3 generate_excel_from_template_json.py \
      --input-json template_generation_input.example.json \
      --mhl-template "MHL 2026 Feb.xlsx" \
      --town-template "7-MTL for Township Summary_4.xlsx" \
      --output-dir outputs
"""

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import column_index_from_string


MONTH_NAME_TO_NUM = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}

TOWN_ALIASES = {
    "meiktila": "meiktila",
    "tharzi": "tharzi",
    "thazi": "tharzi",
    "pyawbwe": "pyawbwe",
    "pyawbwee": "pyawbwe",
    "pyawbwe": "pyawbwe",
    "pyawbwe": "pyawbwe",
    "wantwin": "wantwin",
    "wanttwin": "wantwin",
    "wandwin": "wantwin",
    "mahlaing": "mahlaing",
    "mahaling": "mahlaing",
    "yamethin": "yamethin",
    "kyaukpadaung": "kyaukpadaung",
    "kyaukpandaung": "kyaukpadaung",
    "taungthar": "taungthar",
    "myingyan": "myingyan",
    "myingan": "myingyan",
    "bagan": "bagan",
    "pakokku": "pakokku",
    "pakoku": "pakokku",
}


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def as_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = as_text(value).replace(",", "")
    if text in {"", "-", "nan", "None", "#N/A"}:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def normalize_token(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def canonical_town(value: str) -> str:
    token = normalize_token(value)
    return TOWN_ALIASES.get(token, token)


def canonical_product_name(value: str) -> str:
    cleaned = value.lower()
    cleaned = cleaned.replace('"', "")
    cleaned = cleaned.replace("'", "")
    cleaned = re.sub(r"[^a-z0-9]+", "", cleaned)
    return cleaned


def product_key(product_name: str, ml: float) -> tuple[str, int]:
    ml_milli = int(round(as_float(ml) * 1000))
    return canonical_product_name(product_name), ml_milli


def parse_year_month(value: Any) -> tuple[int, int] | None:
    if isinstance(value, datetime):
        return value.year, value.month
    if isinstance(value, date):
        return value.year, value.month
    text = as_text(value)
    if not text:
        return None
    m = re.match(r"^\s*(\d{4})-(\d{1,2})(?:-(\d{1,2}))?\s*$", text)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.match(r"^\s*(\d{4})/(\d{1,2})(?:/(\d{1,2}))?\s*$", text)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.match(r"^\s*(\d{4})(\d{2})\s*$", text)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None


def parse_month_name(value: Any) -> int | None:
    token = as_text(value).lower()
    if not token:
        return None
    token = re.sub(r"[^a-z]", "", token)
    if token in MONTH_NAME_TO_NUM:
        return MONTH_NAME_TO_NUM[token]
    return None


def infer_year_from_sheet(ws: Worksheet, fallback_year: int) -> int:
    for r in range(1, min(ws.max_row, 5) + 1):
        for c in range(1, min(ws.max_column, 10) + 1):
            text = as_text(ws.cell(r, c).value)
            if not text:
                continue
            m = re.search(r"(20\d{2})", text)
            if m:
                return int(m.group(1))
    m = re.search(r"(20\d{2})", ws.title)
    if m:
        return int(m.group(1))
    return fallback_year


def metric_label(value: Any) -> str | None:
    token = normalize_token(as_text(value))
    if token in {"pk", "pkt"}:
        return "pk"
    if token in {"bot", "bottle"}:
        return "bottle"
    if token in {"lit", "liter", "litre"}:
        return "liter"
    return None


def row_has_date_cells(ws: Worksheet, row_idx: int) -> bool:
    if row_idx < 1 or row_idx > ws.max_row:
        return False
    for c in range(1, ws.max_column + 1):
        value = ws.cell(row_idx, c).value
        if isinstance(value, (datetime, date)):
            return True
    return False


def detect_header_row(ws: Worksheet, needle: str, max_scan_rows: int = 8) -> int | None:
    target = normalize_token(needle)
    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        for c in range(1, min(ws.max_column, 30) + 1):
            token = normalize_token(as_text(ws.cell(r, c).value))
            if target in token:
                return r
    return None


def find_col_by_header(ws: Worksheet, header_row: int, candidates: Iterable[str], default_col: int) -> int:
    cand = {normalize_token(v) for v in candidates}
    for c in range(1, ws.max_column + 1):
        token = normalize_token(as_text(ws.cell(header_row, c).value))
        if token in cand:
            return c
    return default_col


def detect_month_metric_columns(
    ws: Worksheet,
    month_row: int,
    metric_row: int,
    fallback_year: int,
) -> dict[tuple[int, int], dict[str, int]]:
    month_cols: dict[tuple[int, int], dict[str, int]] = {}

    # Pattern A: explicit datetime cells in month row.
    for c in range(1, ws.max_column + 1):
        ym = parse_year_month(ws.cell(month_row, c).value)
        if not ym:
            continue
        slots = month_cols.setdefault(ym, {})
        for offset in (0, 1, 2):
            cc = c + offset
            if cc > ws.max_column:
                break
            label = metric_label(ws.cell(metric_row, cc).value)
            if label and label not in slots:
                slots[label] = cc

    if month_cols:
        return month_cols

    # Pattern B: month names (January, Feb, Mar...) with implicit sheet year.
    inferred_year = infer_year_from_sheet(ws, fallback_year)
    for c in range(1, ws.max_column + 1):
        m = parse_month_name(ws.cell(month_row, c).value)
        if not m:
            continue
        ym = (inferred_year, m)
        slots = month_cols.setdefault(ym, {})
        for offset in (0, 1, 2):
            cc = c + offset
            if cc > ws.max_column:
                break
            label = metric_label(ws.cell(metric_row, cc).value)
            if label and label not in slots:
                slots[label] = cc

    return month_cols


def cell_is_formula(cell: Cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")


def write_if_input(ws: Worksheet, row: int, col: int, value: Any) -> bool:
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return False
    if cell_is_formula(cell):
        return False
    if value is None:
        return False
    if isinstance(value, float):
        if abs(value) < 1e-12:
            value = 0.0
    cell.value = value
    return True


@dataclass
class InputModel:
    max_year: int = 2026
    sku_by_id: dict[str, dict[str, Any]] = field(default_factory=dict)
    customer_by_id: dict[str, dict[str, Any]] = field(default_factory=dict)
    township_name_by_code: dict[str, str] = field(default_factory=dict)
    town_product_month: dict[tuple[str, tuple[str, int], tuple[int, int]], dict[str, float]] = field(
        default_factory=lambda: defaultdict(lambda: {"pk": 0.0, "bottle": 0.0, "liter": 0.0})
    )
    town_name_month: dict[tuple[str, str, tuple[int, int]], dict[str, float]] = field(
        default_factory=lambda: defaultdict(lambda: {"pk": 0.0, "bottle": 0.0, "liter": 0.0})
    )
    region_product_month: dict[tuple[tuple[str, int], tuple[int, int]], dict[str, float]] = field(
        default_factory=lambda: defaultdict(lambda: {"pk": 0.0, "bottle": 0.0, "liter": 0.0})
    )
    region_name_month: dict[tuple[str, tuple[int, int]], dict[str, float]] = field(
        default_factory=lambda: defaultdict(lambda: {"pk": 0.0, "bottle": 0.0, "liter": 0.0})
    )
    town_month: dict[tuple[str, tuple[int, int]], dict[str, float]] = field(
        default_factory=lambda: defaultdict(lambda: {"pk": 0.0, "bottle": 0.0, "liter": 0.0})
    )
    customer_month: dict[tuple[str, tuple[int, int]], dict[str, float]] = field(
        default_factory=lambda: defaultdict(lambda: {"pk": 0.0, "bottle": 0.0, "liter": 0.0})
    )


def resolve_town_name(record: dict[str, Any], model: InputModel) -> str:
    code = as_text(record.get("township_code") or record.get("township") or record.get("township_name"))
    if code in model.township_name_by_code:
        return model.township_name_by_code[code]
    return code


def resolve_product_fields(record: dict[str, Any], model: InputModel) -> tuple[str, float]:
    sku_id = as_text(record.get("sku_id") or record.get("stock_id"))
    if sku_id and sku_id in model.sku_by_id:
        src = model.sku_by_id[sku_id]
        return as_text(src.get("product_name")), as_float(src.get("ml"))

    name = as_text(record.get("product_name") or record.get("stock_name") or record.get("particular"))
    ml = as_float(record.get("ml"))
    return name, ml


def add_metrics(bucket: dict[str, float], pk: float, bottle: float, liter: float, has_liter: bool) -> None:
    bucket["pk"] += pk
    bucket["bottle"] += bottle
    if has_liter:
        bucket["liter"] += liter


def build_model(payload: dict[str, Any]) -> InputModel:
    model = InputModel()

    for rec in payload.get("sku_master", []):
        sku_id = as_text(rec.get("sku_id") or rec.get("stock_id"))
        if not sku_id:
            continue
        model.sku_by_id[sku_id] = {
            "product_name": as_text(rec.get("product_name") or rec.get("stock_name")),
            "ml": as_float(rec.get("ml")),
        }

    for rec in payload.get("customer_master", []):
        customer_id = as_text(rec.get("customer_id"))
        if not customer_id:
            continue
        model.customer_by_id[customer_id] = rec

    for rec in payload.get("township_master", []):
        code = as_text(rec.get("township_code"))
        name = as_text(rec.get("township_name"))
        if code and name:
            model.township_name_by_code[code] = name

    sales_sections = ["sales_monthly_sku_township", "future_plan_monthly_sku_township_2026"]
    for section in sales_sections:
        for rec in payload.get(section, []):
            ym = parse_year_month(rec.get("year_month") or rec.get("month") or rec.get("date"))
            if not ym:
                continue
            model.max_year = max(model.max_year, ym[0])

            town = canonical_town(resolve_town_name(rec, model))
            pname, ml = resolve_product_fields(rec, model)
            if not pname:
                continue
            pkey = product_key(pname, ml)
            name_key = canonical_product_name(pname)

            pk = as_float(rec.get("pk_qty") or rec.get("pkt_qty") or rec.get("pk"))
            bottle = as_float(rec.get("bottle_qty") or rec.get("bot_qty") or rec.get("bottle") or rec.get("sales_bot"))
            liter_raw = rec.get("liter")
            liter = as_float(liter_raw)
            has_liter = liter_raw is not None and as_text(liter_raw) != ""
            if not has_liter:
                liter = bottle * as_float(ml)

            add_metrics(model.town_product_month[(town, pkey, ym)], pk, bottle, liter, True)
            add_metrics(model.town_name_month[(town, name_key, ym)], pk, bottle, liter, True)
            add_metrics(model.town_month[(town, ym)], pk, bottle, liter, True)

    for rec in payload.get("sales_monthly_sku_region", []):
        ym = parse_year_month(rec.get("year_month") or rec.get("month") or rec.get("date"))
        if not ym:
            continue
        model.max_year = max(model.max_year, ym[0])

        pname, ml = resolve_product_fields(rec, model)
        if not pname:
            continue
        pkey = product_key(pname, ml)
        name_key = canonical_product_name(pname)

        pk = as_float(rec.get("pk_qty") or rec.get("pkt_qty") or rec.get("pk"))
        bottle = as_float(rec.get("bottle_qty") or rec.get("bot_qty") or rec.get("bottle") or rec.get("sales_bot"))
        liter_raw = rec.get("liter")
        liter = as_float(liter_raw)
        has_liter = liter_raw is not None and as_text(liter_raw) != ""
        if not has_liter:
            liter = bottle * as_float(ml)

        add_metrics(model.region_product_month[(pkey, ym)], pk, bottle, liter, True)
        add_metrics(model.region_name_month[(name_key, ym)], pk, bottle, liter, True)

    # Fallback regional aggregation from township facts if section not provided.
    if not model.region_product_month:
        for (_, pkey, ym), metrics in model.town_product_month.items():
            add_metrics(
                model.region_product_month[(pkey, ym)],
                metrics["pk"],
                metrics["bottle"],
                metrics["liter"],
                True,
            )
        for (_, name_key, ym), metrics in model.town_name_month.items():
            add_metrics(
                model.region_name_month[(name_key, ym)],
                metrics["pk"],
                metrics["bottle"],
                metrics["liter"],
                True,
            )

    for rec in payload.get("sales_monthly_customer", []):
        ym = parse_year_month(rec.get("year_month") or rec.get("month") or rec.get("date"))
        if not ym:
            continue
        model.max_year = max(model.max_year, ym[0])
        cid = as_text(rec.get("customer_id"))
        if not cid:
            continue
        pk = as_float(rec.get("pk_qty") or rec.get("pkt_qty") or rec.get("pk"))
        bottle = as_float(rec.get("bottle_qty") or rec.get("bot_qty") or rec.get("bottle") or rec.get("sales_bot"))
        liter_raw = rec.get("liter")
        liter = as_float(liter_raw)
        has_liter = liter_raw is not None and as_text(liter_raw) != ""
        add_metrics(model.customer_month[(cid, ym)], pk, bottle, liter, has_liter)

    return model


def lookup_product_metrics(
    model: InputModel,
    sheet_town: str | None,
    row_product_name: str,
    row_ml: float,
    ym: tuple[int, int],
) -> dict[str, float] | None:
    pkey = product_key(row_product_name, row_ml)
    name_key = canonical_product_name(row_product_name)

    if sheet_town:
        by_exact = model.town_product_month.get((sheet_town, pkey, ym))
        if by_exact:
            return by_exact
        by_name = model.town_name_month.get((sheet_town, name_key, ym))
        if by_name:
            return by_name
        return None

    by_exact = model.region_product_month.get((pkey, ym))
    if by_exact:
        return by_exact
    by_name = model.region_name_month.get((name_key, ym))
    if by_name:
        return by_name
    return None


def fill_product_monthly_sheet(ws: Worksheet, model: InputModel) -> int:
    header_row = detect_header_row(ws, "product")
    if not header_row:
        return 0

    month_row = header_row
    metric_row = header_row + 1
    if metric_row > ws.max_row:
        return 0

    month_cols = detect_month_metric_columns(ws, month_row, metric_row, model.max_year)
    if not month_cols:
        return 0

    product_col = find_col_by_header(ws, header_row, {"productname", "particular"}, 2)
    ml_col = find_col_by_header(ws, header_row, {"ml"}, 3)
    data_start = metric_row + 1

    sheet_town_key = canonical_town(ws.title)
    has_town_data = any(k[0] == sheet_town_key for k in model.town_month.keys())
    scope_town = sheet_town_key if has_town_data else None

    writes = 0
    for r in range(data_start, ws.max_row + 1):
        pname = as_text(ws.cell(r, product_col).value)
        if not pname:
            continue
        if "total" in pname.lower() and as_float(ws.cell(r, ml_col).value) <= 0:
            continue
        ml = as_float(ws.cell(r, ml_col).value)
        for ym, slots in month_cols.items():
            metrics = lookup_product_metrics(model, scope_town, pname, ml, ym)
            if not metrics:
                continue
            if "pk" in slots:
                writes += int(write_if_input(ws, r, slots["pk"], metrics["pk"]))
            if "bottle" in slots:
                writes += int(write_if_input(ws, r, slots["bottle"], metrics["bottle"]))
            if "liter" in slots:
                liter = metrics["liter"] if metrics["liter"] else (metrics["bottle"] * ml if ml > 0 else 0.0)
                writes += int(write_if_input(ws, r, slots["liter"], liter))
    return writes


def fill_township_monthly_sheet(ws: Worksheet, model: InputModel) -> int:
    header_row = detect_header_row(ws, "township")
    if not header_row:
        return 0

    # Some sheets have month dates on the same header row; others on next row.
    if row_has_date_cells(ws, header_row):
        month_row = header_row
        metric_row = header_row + 1
    else:
        month_row = header_row + 1
        metric_row = header_row + 2

    if metric_row > ws.max_row:
        return 0

    month_cols = detect_month_metric_columns(ws, month_row, metric_row, model.max_year)
    if not month_cols:
        return 0

    township_col = find_col_by_header(ws, header_row, {"township", "townshipname"}, 2)
    data_start = metric_row + 1

    writes = 0
    for r in range(data_start, ws.max_row + 1):
        town_value = as_text(ws.cell(r, township_col).value)
        if not town_value:
            continue
        town_key = canonical_town(town_value)
        if not town_key:
            continue
        for ym, slots in month_cols.items():
            metrics = model.town_month.get((town_key, ym))
            if not metrics:
                continue
            if "pk" in slots:
                writes += int(write_if_input(ws, r, slots["pk"], metrics["pk"]))
            if "bottle" in slots:
                writes += int(write_if_input(ws, r, slots["bottle"], metrics["bottle"]))
            if "liter" in slots:
                writes += int(write_if_input(ws, r, slots["liter"], metrics["liter"]))
    return writes


def fill_customer_monthly_sheet(ws: Worksheet, model: InputModel) -> int:
    header_row = detect_header_row(ws, "customername")
    if not header_row:
        return 0

    month_row = header_row - 1 if row_has_date_cells(ws, header_row - 1) else header_row
    metric_row = header_row
    if metric_row > ws.max_row or month_row < 1:
        return 0

    month_cols = detect_month_metric_columns(ws, month_row, metric_row, model.max_year)
    if not month_cols:
        return 0

    customer_id_col = find_col_by_header(ws, header_row, {"customerid", "no"}, 1)
    data_start = header_row + 1
    writes = 0

    for r in range(data_start, ws.max_row + 1):
        cid = as_text(ws.cell(r, customer_id_col).value)
        if not cid:
            continue
        for ym, slots in month_cols.items():
            metrics = model.customer_month.get((cid, ym))
            if not metrics:
                continue
            if "pk" in slots:
                writes += int(write_if_input(ws, r, slots["pk"], metrics["pk"]))
            if "bottle" in slots:
                writes += int(write_if_input(ws, r, slots["bottle"], metrics["bottle"]))
            if "liter" in slots:
                writes += int(write_if_input(ws, r, slots["liter"], metrics["liter"]))
    return writes


def fill_outlet_list_sheet(ws: Worksheet, outlet_master: list[dict[str, Any]]) -> int:
    if not outlet_master:
        return 0
    if ws.title.strip() != "Outlet List":
        return 0

    start_row = 5
    writes = 0
    for idx, rec in enumerate(outlet_master, start=1):
        r = start_row + idx - 1
        writes += int(write_if_input(ws, r, 1, idx))
        writes += int(write_if_input(ws, r, 2, as_text(rec.get("outlet_name"))))
        writes += int(write_if_input(ws, r, 3, as_text(rec.get("type"))))
        writes += int(write_if_input(ws, r, 4, as_text(rec.get("address"))))
        writes += int(write_if_input(ws, r, 5, as_text(rec.get("township_name") or rec.get("township_code"))))
        writes += int(write_if_input(ws, r, 6, as_text(rec.get("way_code"))))
        writes += int(write_if_input(ws, r, 7, as_text(rec.get("phone"))))
        writes += int(write_if_input(ws, r, 8, as_text(rec.get("owner_customer_id"))))
        writes += int(write_if_input(ws, r, 9, as_float(rec.get("investment_score"))))
    return writes


def fill_way_plan_sheet(ws: Worksheet, way_plan_daily: list[dict[str, Any]]) -> int:
    if not way_plan_daily:
        return 0
    if normalize_token(ws.title) != "wayplan":
        return 0

    rows = sorted(
        way_plan_daily,
        key=lambda r: (
            parse_year_month(r.get("date")) or (9999, 12),
            as_text(r.get("way_code")),
        ),
    )
    writes = 0
    start_row = 2
    for idx, rec in enumerate(rows, start=0):
        r = start_row + idx
        d = rec.get("date")
        if isinstance(d, str):
            # Keep as date string; Excel can parse if opened depending locale.
            d = as_text(d)
        writes += int(write_if_input(ws, r, 1, d))
        writes += int(write_if_input(ws, r, 2, as_text(rec.get("day_name"))))
        writes += int(write_if_input(ws, r, 3, as_text(rec.get("way_code"))))
        writes += int(write_if_input(ws, r, 4, as_text(rec.get("actual_way_name"))))
        writes += int(write_if_input(ws, r, 5, as_float(rec.get("a"))))
        writes += int(write_if_input(ws, r, 6, as_float(rec.get("b"))))
        writes += int(write_if_input(ws, r, 7, as_float(rec.get("c"))))
        writes += int(write_if_input(ws, r, 8, as_float(rec.get("d"))))
        # column I often "S" in template.
        writes += int(write_if_input(ws, r, 9, as_float(rec.get("s"))))
    return writes


def fill_competition_sheet(ws: Worksheet, competition_info: list[dict[str, Any]]) -> int:
    if not competition_info:
        return 0
    if normalize_token(ws.title) != "competitioninformation":
        return 0

    writes = 0
    start_row = 3
    for idx, rec in enumerate(competition_info, start=1):
        r = start_row + idx - 1
        writes += int(write_if_input(ws, r, 1, idx))
        writes += int(write_if_input(ws, r, 2, as_text(rec.get("region"))))
        writes += int(write_if_input(ws, r, 3, as_text(rec.get("town"))))
        writes += int(write_if_input(ws, r, 4, as_text(rec.get("company_name"))))
        writes += int(write_if_input(ws, r, 5, as_text(rec.get("distributor"))))
        writes += int(write_if_input(ws, r, 6, as_text(rec.get("township_code"))))
        writes += int(write_if_input(ws, r, 7, as_text(rec.get("product_name"))))
        writes += int(write_if_input(ws, r, 8, as_float(rec.get("db_landing_price"))))
        writes += int(write_if_input(ws, r, 9, as_float(rec.get("db_selling_price"))))
        writes += int(write_if_input(ws, r, 11, as_text(rec.get("focus_or_budget"))))
        writes += int(write_if_input(ws, r, 12, as_float(rec.get("size_ml"))))
        writes += int(write_if_input(ws, r, 13, as_float(rec.get("packing_size"))))
        writes += int(write_if_input(ws, r, 14, as_float(rec.get("abv"))))
        writes += int(write_if_input(ws, r, 15, as_float(rec.get("buying_price"))))
        writes += int(write_if_input(ws, r, 16, as_float(rec.get("freight_labor"))))
        writes += int(write_if_input(ws, r, 17, as_float(rec.get("trade_promo"))))
        writes += int(write_if_input(ws, r, 19, as_float(rec.get("customer_selling_price"))))
        writes += int(write_if_input(ws, r, 20, as_float(rec.get("estimated_sec"))))
    return writes


def apply_workbook_patches(wb, patches: list[dict[str, Any]]) -> int:
    writes = 0
    for patch in patches:
        sheet_name = as_text(patch.get("sheet"))
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        cell_ref = as_text(patch.get("cell"))
        if not cell_ref:
            continue
        value = patch.get("value")
        cell = ws[cell_ref]
        if cell_is_formula(cell):
            continue
        cell.value = value
        writes += 1
    return writes


def apply_table_patches(wb, table_patches: list[dict[str, Any]]) -> int:
    writes = 0
    for patch in table_patches:
        sheet_name = as_text(patch.get("sheet"))
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        start_row = int(as_float(patch.get("start_row") or 1))
        columns = patch.get("columns") or []
        rows = patch.get("rows") or []
        clear_after = bool(patch.get("clear_after", False))
        if not columns:
            continue
        col_idx = [column_index_from_string(as_text(c).upper()) for c in columns]
        for row_offset, row_values in enumerate(rows):
            r = start_row + row_offset
            for cpos, c in enumerate(col_idx):
                val = row_values[cpos] if cpos < len(row_values) else None
                writes += int(write_if_input(ws, r, c, val))
        if clear_after:
            end_row = start_row + len(rows)
            for r in range(end_row, ws.max_row + 1):
                for c in col_idx:
                    cell = ws.cell(r, c)
                    if cell_is_formula(cell):
                        continue
                    if cell.value not in (None, ""):
                        cell.value = None
                        writes += 1
    return writes


def fill_workbook_from_model(
    workbook_path: Path,
    output_path: Path,
    model: InputModel,
    payload: dict[str, Any],
    workbook_key: str,
) -> dict[str, int]:
    wb = load_workbook(workbook_path)
    stats = defaultdict(int)

    for ws in wb.worksheets:
        stats["product_monthly_writes"] += fill_product_monthly_sheet(ws, model)
        stats["township_monthly_writes"] += fill_township_monthly_sheet(ws, model)
        stats["customer_monthly_writes"] += fill_customer_monthly_sheet(ws, model)
        stats["outlet_writes"] += fill_outlet_list_sheet(ws, payload.get("outlet_master", []))
        stats["way_plan_writes"] += fill_way_plan_sheet(ws, payload.get("way_plan_daily", []))
        stats["competition_writes"] += fill_competition_sheet(ws, payload.get("competition_info", []))

    # Optional explicit patches for anything not covered by normalized fillers.
    raw_patches = payload.get("workbook_patches", [])
    patches = [p for p in raw_patches if as_text(p.get("workbook", "both")).lower() in {workbook_key, "both"}]
    stats["explicit_patch_writes"] += apply_workbook_patches(wb, patches)

    raw_table_patches = payload.get("workbook_table_patches", [])
    table_patches = [
        p for p in raw_table_patches if as_text(p.get("workbook", "both")).lower() in {workbook_key, "both"}
    ]
    stats["explicit_table_patch_writes"] += apply_table_patches(wb, table_patches)

    wb.save(output_path)
    wb.close()
    stats["saved"] = 1
    return dict(stats)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate MHL + 7-MTL workbooks from templates + JSON.")
    parser.add_argument("--input-json", required=True, help="JSON input file path")
    parser.add_argument("--mhl-template", default="MHL 2026 Feb.xlsx", help="Path to MHL template workbook")
    parser.add_argument(
        "--town-template",
        default="7-MTL for Township Summary_4.xlsx",
        help="Path to 7-MTL township template workbook",
    )
    parser.add_argument("--output-dir", default=".", help="Output directory")
    parser.add_argument("--mhl-output", default="MHL_from_json.xlsx", help="Output filename for MHL workbook")
    parser.add_argument(
        "--town-output",
        default="7-MTL_for_Township_Summary_from_json.xlsx",
        help="Output filename for 7-MTL workbook",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    input_json_path = Path(args.input_json).expanduser().resolve()
    mhl_template_path = Path(args.mhl_template).expanduser().resolve()
    town_template_path = Path(args.town_template).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    with input_json_path.open("r", encoding="utf-8") as f:
        payload = json.load(f)

    model = build_model(payload)

    mhl_output_path = output_dir / args.mhl_output
    town_output_path = output_dir / args.town_output

    mhl_stats = fill_workbook_from_model(
        workbook_path=mhl_template_path,
        output_path=mhl_output_path,
        model=model,
        payload=payload,
        workbook_key="mhl",
    )
    town_stats = fill_workbook_from_model(
        workbook_path=town_template_path,
        output_path=town_output_path,
        model=model,
        payload=payload,
        workbook_key="town",
    )

    print(f"Generated: {mhl_output_path}")
    print(f"Generated: {town_output_path}")
    print("MHL stats:", json.dumps(mhl_stats, ensure_ascii=False))
    print("Town stats:", json.dumps(town_stats, ensure_ascii=False))


if __name__ == "__main__":
    main()
