from __future__ import annotations

import calendar
import html
import json
import os
import re
import socket
import time
from collections import Counter, OrderedDict
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path
from typing import Any

from flask import Flask, abort, jsonify, render_template, request
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = ROOT_DIR / "uploads"
WORKBOOK_REGISTRY_PATH = UPLOAD_DIR / "workbook_registry.json"
ACCESS_CONTROL_PATH = UPLOAD_DIR / "access_control.json"
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
ROLE_OWNER = "owner"
ROLE_RSM = "rsm"
ROLE_ASM = "asm"
ROLE_USER = "user"
ALLOWED_USER_ROLES = {ROLE_OWNER, ROLE_RSM, ROLE_ASM, ROLE_USER}
ALLOWED_VIEWER_ROLES = {"owner", "regional_manager", "rsm", "asm", "user"}
ALL_REGION_TOKEN = "ALL"
DEFAULT_REGION_TOKEN = "GLOBAL"

MONTH_LOOKUP = {
    name.lower(): idx
    for idx, name in enumerate(calendar.month_name)
    if idx
}
MONTH_LOOKUP.update(
    {
        name.lower(): idx
        for idx, name in enumerate(calendar.month_abbr)
        if idx
    }
)

REGION_STOP_WORDS = {
    "analysis",
    "detail",
    "for",
    "main",
    "master",
    "month",
    "monthly",
    "overview",
    "reference",
    "sheet",
    "sku",
    "summary",
    "township",
    "workbook",
}
REGION_STOP_WORDS.update({name.lower() for name in calendar.month_name if name})
REGION_STOP_WORDS.update({name.lower() for name in calendar.month_abbr if name})

TOWNSHIP_ALIAS = {
    "meiktila": "meiktila",
    "tharzi": "tharzi",
    "tarzi": "tharzi",
    "pyawbwe": "pyawbwe",
    "pyawbwe": "pyawbwe",
    "wanttwin": "wantwin",
    "wantwin": "wantwin",
    "mahaling": "mahlaing",
    "mahlaing": "mahlaing",
    "yamethin": "yamethin",
    "kyaukpadaung": "kyaukpandaung",
    "kyaukpandaung": "kyaukpandaung",
    "taungthar": "taungthar",
    "taugthar": "taungthar",
    "myingyan": "myingan",
    "myingan": "myingan",
    "bagan": "bagan",
    "pakokku": "pakokku",
    "pakoku": "pakokku",
}


app = Flask(__name__)


def normalize_token(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def canonical_sheet_name(sheet_name: str) -> str:
    token = normalize_token(sheet_name)
    return TOWNSHIP_ALIAS.get(token, token)


def parse_month_from_text(text: str) -> int | None:
    tokens = [t for t in re.split(r"[^a-z]+", text.lower()) if t]
    matched = [MONTH_LOOKUP[token] for token in tokens if token in MONTH_LOOKUP]
    if not matched:
        return None
    if len(set(matched)) > 1:
        return None
    return matched[0]


def extract_year_from_text(text: str) -> int | None:
    match = re.search(r"(20\d{2})", text)
    if not match:
        return None
    return int(match.group(1))


def parse_month_header(value: Any, fallback_year: int | None) -> tuple[int, int] | None:
    if isinstance(value, datetime):
        return value.year, value.month
    if isinstance(value, date):
        return value.year, value.month

    if isinstance(value, str):
        month_num = parse_month_from_text(value)
        if month_num is None:
            return None
        year = extract_year_from_text(value) or fallback_year
        if year is None:
            return None
        return year, month_num

    return None


def parse_month_header_loose(
    value: Any, fallback_year: int | None
) -> tuple[int | None, int] | None:
    if isinstance(value, datetime):
        return value.year, value.month
    if isinstance(value, date):
        return value.year, value.month

    if isinstance(value, str):
        month_num = parse_month_from_text(value)
        if month_num is None:
            return None
        year = extract_year_from_text(value) or fallback_year
        return year, month_num

    return None


def parse_numeric(value: Any) -> int | float | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        try:
            number = float(stripped)
        except ValueError:
            return None
    elif isinstance(value, (int, float)):
        number = float(value)
    else:
        return None

    if number.is_integer():
        return int(number)
    return round(number, 4)


def parse_year_context(worksheet, header_row: int) -> int | None:
    for row_idx in range(1, header_row):
        for col_idx in range(1, min(worksheet.max_column, 30) + 1):
            value = worksheet.cell(row=row_idx, column=col_idx).value
            if isinstance(value, (int, float)) and 2000 <= int(value) <= 2100:
                return int(value)
            if isinstance(value, str):
                year = extract_year_from_text(value)
                if year:
                    return year
    return None


def find_header_layout(worksheet) -> dict[str, int] | None:
    max_scan_col = min(worksheet.max_column, 60)
    for row_idx in range(1, min(worksheet.max_row, 12) + 1):
        columns: dict[str, int] = {}
        for col_idx in range(1, max_scan_col + 1):
            token = normalize_token(worksheet.cell(row=row_idx, column=col_idx).value)
            if not token:
                continue
            if token == "sr":
                columns["sr"] = col_idx
            elif token == "productname":
                columns["product_name"] = col_idx
            elif token == "ml":
                columns["ml"] = col_idx
            elif token == "packing":
                columns["packing"] = col_idx

        required = {"product_name", "ml", "packing"}
        if required.issubset(columns):
            columns["header_row"] = row_idx
            columns["metrics_row"] = row_idx + 1
            return columns
    return None


def parse_month_groups(worksheet, layout: dict[str, int]) -> list[dict[str, Any]]:
    header_row = layout["header_row"]
    metrics_row = layout["metrics_row"]
    fallback_year = parse_year_context(worksheet, header_row)
    start_col = layout["packing"] + 1

    month_groups: list[dict[str, Any]] = []
    seen_keys: set[str] = set()

    for col_idx in range(start_col, worksheet.max_column - 1):
        metric_tokens = [
            normalize_token(worksheet.cell(row=metrics_row, column=col_idx + offset).value)
            for offset in range(3)
        ]
        if metric_tokens[:2] != ["pk", "bottle"]:
            continue
        if metric_tokens[2] not in {"liter", "litre"}:
            continue

        header_value = worksheet.cell(row=header_row, column=col_idx).value
        parsed_month = parse_month_header(header_value, fallback_year)
        if not parsed_month:
            continue

        year, month = parsed_month
        month_key = f"{year:04d}-{month:02d}"
        if month_key in seen_keys:
            continue

        seen_keys.add(month_key)
        month_groups.append(
            {
                "key": month_key,
                "label": f"{calendar.month_abbr[month]} {year}",
                "year": year,
                "month": month,
                "col": col_idx,
            }
        )

    month_groups.sort(key=lambda item: item["key"])
    return month_groups


def row_identity_key(product_name: str, ml: Any, packing: Any) -> str:
    ml_token = "" if ml is None else str(ml).strip()
    packing_token = normalize_token(packing)
    return f"{normalize_token(product_name)}|{ml_token}|{packing_token}"


def parse_rows(worksheet, layout: dict[str, int], month_groups: list[dict[str, Any]]) -> list[dict[str, Any]]:
    sr_col = layout.get("sr")
    product_col = layout["product_name"]
    ml_col = layout["ml"]
    packing_col = layout["packing"]
    data_start = layout["metrics_row"] + 1

    parsed_rows: list[dict[str, Any]] = []
    blank_streak = 0

    for row_idx in range(data_start, worksheet.max_row + 1):
        sr_value = worksheet.cell(row=row_idx, column=sr_col).value if sr_col else None
        product_value = worksheet.cell(row=row_idx, column=product_col).value
        ml_value = worksheet.cell(row=row_idx, column=ml_col).value
        packing_value = worksheet.cell(row=row_idx, column=packing_col).value

        values_by_month: dict[str, dict[str, int | float | None]] = {}
        has_month_data = False

        for group in month_groups:
            col_idx = group["col"]
            pk_val = parse_numeric(worksheet.cell(row=row_idx, column=col_idx).value)
            bottle_val = parse_numeric(worksheet.cell(row=row_idx, column=col_idx + 1).value)
            liter_val = parse_numeric(worksheet.cell(row=row_idx, column=col_idx + 2).value)
            values_by_month[group["key"]] = {
                "pk": pk_val,
                "bottle": bottle_val,
                "liter": liter_val,
            }
            if pk_val is not None or bottle_val is not None or liter_val is not None:
                has_month_data = True

        has_identity = any(
            value not in (None, "") for value in (sr_value, product_value, ml_value, packing_value)
        )

        if not has_identity and not has_month_data:
            blank_streak += 1
            if blank_streak >= 20:
                break
            continue

        blank_streak = 0
        if product_value in (None, ""):
            continue

        product_name = str(product_value).strip()
        parsed_rows.append(
            {
                "sr": sr_value,
                "product_name": product_name,
                "ml": ml_value,
                "packing": packing_value,
                "row_key": row_identity_key(product_name, ml_value, packing_value),
                "values": values_by_month,
            }
        )

    return parsed_rows


def parse_fixed_sheet(worksheet, sheet_name: str) -> dict[str, Any] | None:
    layout = find_header_layout(worksheet)
    if not layout:
        return None

    month_groups = parse_month_groups(worksheet, layout)
    if len(month_groups) < 2:
        return None

    parsed_rows = parse_rows(worksheet, layout, month_groups)
    if not parsed_rows:
        return None

    return {
        "sheet_name": sheet_name,
        "canonical": canonical_sheet_name(sheet_name),
        "months": [
            {
                "key": group["key"],
                "label": group["label"],
                "year": group["year"],
                "month": group["month"],
            }
            for group in month_groups
        ],
        "rows": parsed_rows,
    }


def parse_workbook(path: Path) -> OrderedDict[str, dict[str, Any]]:
    # Use normal mode because this parser relies on random cell access.
    workbook = load_workbook(path, read_only=False, data_only=True, keep_links=False)
    parsed_sheets: OrderedDict[str, dict[str, Any]] = OrderedDict()

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        parsed_sheet = parse_fixed_sheet(worksheet, sheet_name)
        if not parsed_sheet:
            continue

        canonical = parsed_sheet["canonical"]
        if canonical in parsed_sheets:
            # Prefer the first matching sheet in workbook order.
            continue
        parsed_sheets[canonical] = parsed_sheet

    return parsed_sheets


@lru_cache(maxsize=16)
def parse_workbook_cached(path_str: str, mtime_ns: int) -> OrderedDict[str, dict[str, Any]]:
    _ = mtime_ns
    return parse_workbook(Path(path_str))


@lru_cache(maxsize=8)
def load_workbook_cached(path_str: str, mtime_ns: int):
    _ = mtime_ns
    return load_workbook(Path(path_str), read_only=False, data_only=True, keep_links=False)


@lru_cache(maxsize=8)
def workbook_sheetnames_cached(path_str: str, mtime_ns: int) -> tuple[str, ...]:
    _ = mtime_ns
    workbook = load_workbook(
        Path(path_str),
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        visible_names = tuple(
            worksheet.title
            for worksheet in workbook.worksheets
            if getattr(worksheet, "sheet_state", "visible") == "visible"
        )
        if visible_names:
            return visible_names
        return tuple(workbook.sheetnames)
    finally:
        workbook.close()


def normalize_sheet_name_list(sheet_names: tuple[str, ...] | list[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for raw_name in sheet_names:
        if not isinstance(raw_name, str):
            continue
        sheet_name = raw_name
        if not sheet_name or sheet_name in seen:
            continue
        seen.add(sheet_name)
        ordered.append(sheet_name)
    return ordered


def normalize_viewer_role(raw_role: str | None) -> str:
    role = normalize_token(raw_role)
    if role in {"owner"}:
        return ROLE_OWNER
    if role in {"regionalmanager", "regional_manager", "rsm"}:
        return ROLE_RSM
    if role in {"asm"}:
        return ROLE_ASM
    if role in {"user", "staff"}:
        return ROLE_USER
    return ROLE_OWNER


def normalize_user_role(raw_role: str | None, default: str = ROLE_USER) -> str:
    role = normalize_viewer_role(raw_role)
    if role in ALLOWED_USER_ROLES:
        return role
    return default


def normalize_username(raw_username: Any) -> str:
    if raw_username is None:
        return ""
    value = str(raw_username).strip().lower()
    if not value:
        return ""
    cleaned = re.sub(r"[^a-z0-9_.-]+", "_", value).strip("._-")
    return cleaned


def normalize_display_name(raw_display_name: Any, fallback_username: str) -> str:
    if raw_display_name is None:
        return fallback_username
    display_name = str(raw_display_name).strip()
    if not display_name:
        return fallback_username
    return re.sub(r"\s+", " ", display_name)


def normalize_region_scope(raw_region: str | None) -> str:
    token = normalize_token(raw_region).upper()
    if not token or token in {"ALL", "ANY"}:
        return ALL_REGION_TOKEN
    return token


def normalize_region_token(raw_region: Any) -> str:
    token = normalize_token(raw_region).upper()
    if not token:
        return DEFAULT_REGION_TOKEN
    if token in {"ALL", "ANY"}:
        return DEFAULT_REGION_TOKEN
    return token


def infer_region_from_filename(path: Path) -> str:
    for token in re.split(r"[^A-Za-z0-9]+", path.stem):
        if not token:
            continue
        lowered = token.lower()
        if lowered in REGION_STOP_WORDS:
            continue
        if re.fullmatch(r"20\d{2}", token):
            continue
        if token.isdigit():
            continue
        cleaned = re.sub(r"[^A-Za-z0-9]+", "", token)
        if len(cleaned) < 2:
            continue
        return cleaned.upper()
    return DEFAULT_REGION_TOKEN


def load_workbook_registry() -> dict[str, dict[str, str]]:
    if not WORKBOOK_REGISTRY_PATH.exists():
        return {}
    try:
        raw = WORKBOOK_REGISTRY_PATH.read_text(encoding="utf-8")
        payload = json.loads(raw)
    except (OSError, ValueError):
        return {}

    if not isinstance(payload, dict):
        return {}
    files_raw = payload.get("files")
    if not isinstance(files_raw, dict):
        return {}

    normalized: dict[str, dict[str, str]] = {}
    for workbook_name, meta in files_raw.items():
        if not isinstance(workbook_name, str):
            continue
        if not isinstance(meta, dict):
            continue
        region = normalize_region_token(meta.get("region"))
        normalized[workbook_name] = {"region": region}
    return normalized


def save_workbook_registry(registry: dict[str, dict[str, str]]) -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    payload = {
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "files": registry,
    }
    WORKBOOK_REGISTRY_PATH.write_text(
        json.dumps(payload, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def parse_token_list(raw_value: Any) -> list[str]:
    if raw_value is None:
        return []
    if isinstance(raw_value, (list, tuple, set)):
        items = raw_value
    else:
        items = re.split(r"[,\n;]+", str(raw_value))
    tokens: list[str] = []
    for item in items:
        text = str(item).strip()
        if text:
            tokens.append(text)
    return tokens


def normalize_region_list(raw_value: Any) -> list[str]:
    regions: list[str] = []
    seen: set[str] = set()
    for token in parse_token_list(raw_value):
        region = normalize_region_token(token)
        if not region or region in seen:
            continue
        seen.add(region)
        regions.append(region)
    return regions


def normalize_township_list(raw_value: Any) -> list[str]:
    townships: list[str] = []
    seen: set[str] = set()
    for token in parse_token_list(raw_value):
        canonical = canonical_sheet_name(token)
        if not canonical or canonical in seen:
            continue
        seen.add(canonical)
        townships.append(canonical)
    return townships


def default_access_control() -> dict[str, Any]:
    return {
        "users": {
            "owner": {
                "role": ROLE_OWNER,
                "display_name": "Owner",
            }
        },
        "rsm_regions": {},
        "user_to_rsm": {},
        "asm_townships": {},
    }


def load_access_control() -> dict[str, Any]:
    defaults = default_access_control()
    if not ACCESS_CONTROL_PATH.exists():
        return defaults

    try:
        raw = ACCESS_CONTROL_PATH.read_text(encoding="utf-8")
        payload = json.loads(raw)
    except (OSError, ValueError):
        return defaults
    if not isinstance(payload, dict):
        return defaults

    users_raw = payload.get("users")
    users: dict[str, dict[str, str]] = {}
    if isinstance(users_raw, dict):
        for username_raw, meta in users_raw.items():
            username = normalize_username(username_raw)
            if not username:
                continue
            role = ROLE_USER
            display_name = username
            if isinstance(meta, dict):
                role = normalize_user_role(meta.get("role"), default=ROLE_USER)
                display_name = normalize_display_name(meta.get("display_name"), username)
            users[username] = {
                "role": role,
                "display_name": display_name,
            }

    if "owner" not in users:
        users["owner"] = {"role": ROLE_OWNER, "display_name": "Owner"}
    else:
        users["owner"]["role"] = ROLE_OWNER

    rsm_regions: dict[str, list[str]] = {}
    rsm_regions_raw = payload.get("rsm_regions")
    if isinstance(rsm_regions_raw, dict):
        for username_raw, regions_raw in rsm_regions_raw.items():
            username = normalize_username(username_raw)
            if not username:
                continue
            regions = normalize_region_list(regions_raw)
            rsm_regions[username] = regions
            if username not in users:
                users[username] = {"role": ROLE_RSM, "display_name": username}

    user_to_rsm: dict[str, str] = {}
    user_to_rsm_raw = payload.get("user_to_rsm")
    if isinstance(user_to_rsm_raw, dict):
        for username_raw, rsm_raw in user_to_rsm_raw.items():
            username = normalize_username(username_raw)
            manager_rsm = normalize_username(rsm_raw)
            if not username or not manager_rsm:
                continue
            user_to_rsm[username] = manager_rsm
            if username not in users:
                users[username] = {"role": ROLE_USER, "display_name": username}
            if manager_rsm not in users:
                users[manager_rsm] = {"role": ROLE_RSM, "display_name": manager_rsm}

    asm_townships: dict[str, dict[str, list[str]]] = {}
    asm_townships_raw = payload.get("asm_townships")
    if isinstance(asm_townships_raw, dict):
        for asm_raw, region_map_raw in asm_townships_raw.items():
            asm_username = normalize_username(asm_raw)
            if not asm_username or not isinstance(region_map_raw, dict):
                continue
            region_map: dict[str, list[str]] = {}
            for region_raw, towns_raw in region_map_raw.items():
                region = normalize_region_token(region_raw)
                townships = normalize_township_list(towns_raw)
                if townships:
                    region_map[region] = townships
            if region_map:
                asm_townships[asm_username] = region_map
                if asm_username not in users:
                    users[asm_username] = {"role": ROLE_ASM, "display_name": asm_username}

    for username, regions in rsm_regions.items():
        if regions and users[username]["role"] != ROLE_OWNER:
            users[username]["role"] = ROLE_RSM
    for username, regions in asm_townships.items():
        if regions and users[username]["role"] not in {ROLE_OWNER, ROLE_RSM}:
            users[username]["role"] = ROLE_ASM

    return {
        "users": users,
        "rsm_regions": rsm_regions,
        "user_to_rsm": user_to_rsm,
        "asm_townships": asm_townships,
    }


def save_access_control(access: dict[str, Any]) -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    payload = {
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "users": access.get("users", {}),
        "rsm_regions": access.get("rsm_regions", {}),
        "user_to_rsm": access.get("user_to_rsm", {}),
        "asm_townships": access.get("asm_townships", {}),
    }
    ACCESS_CONTROL_PATH.write_text(
        json.dumps(payload, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def ensure_access_user(
    access: dict[str, Any],
    username: str,
    role: str = ROLE_USER,
    display_name: str | None = None,
) -> dict[str, str]:
    users = access.setdefault("users", {})
    record = users.get(username)
    if not isinstance(record, dict):
        record = {}
        users[username] = record
    if "role" not in record:
        record["role"] = normalize_user_role(role, default=ROLE_USER)
    if display_name:
        record["display_name"] = normalize_display_name(display_name, username)
    if "display_name" not in record:
        record["display_name"] = username
    if username == "owner":
        record["role"] = ROLE_OWNER
    return record


def all_known_regions(access: dict[str, Any], entries: list[dict[str, Any]]) -> list[str]:
    regions = {str(entry["region"]) for entry in entries}
    for values in access.get("rsm_regions", {}).values():
        for region in values:
            regions.add(normalize_region_token(region))
    for region_map in access.get("asm_townships", {}).values():
        if not isinstance(region_map, dict):
            continue
        for region in region_map:
            regions.add(normalize_region_token(region))
    return sorted(regions)


def derive_allowed_regions(
    access: dict[str, Any],
    username: str,
    role: str,
    regions_universe: list[str],
) -> list[str]:
    region_set = set(regions_universe)
    if role == ROLE_OWNER:
        return sorted(region_set)
    if role == ROLE_RSM:
        assigned = [
            normalize_region_token(region)
            for region in access.get("rsm_regions", {}).get(username, [])
        ]
        return [region for region in assigned if region in region_set]
    if role == ROLE_ASM:
        asm_regions = []
        region_map = access.get("asm_townships", {}).get(username, {})
        if isinstance(region_map, dict):
            asm_regions = [normalize_region_token(region) for region in region_map.keys()]
        return [region for region in asm_regions if region in region_set]

    manager_rsm = access.get("user_to_rsm", {}).get(username)
    if manager_rsm:
        assigned = [
            normalize_region_token(region)
            for region in access.get("rsm_regions", {}).get(manager_rsm, [])
        ]
        return [region for region in assigned if region in region_set]
    return []


def resolve_selected_region(
    requested_region: str | None,
    allowed_regions: list[str],
    allow_all_regions: bool,
) -> str:
    requested = normalize_region_scope(requested_region)
    if allow_all_regions and requested == ALL_REGION_TOKEN:
        return ALL_REGION_TOKEN
    if requested in allowed_regions:
        return requested
    if allowed_regions:
        return allowed_regions[0]
    return ALL_REGION_TOKEN


def current_request_user_and_region() -> tuple[str, str | None]:
    user_raw = request.args.get("user")
    if not user_raw:
        user_raw = request.form.get("user")
    if not user_raw and request.is_json:
        payload = request.get_json(silent=True) or {}
        if isinstance(payload, dict):
            user_raw = payload.get("user")
    username = normalize_username(user_raw) or "owner"

    region_raw = request.args.get("region")
    if region_raw is None:
        region_raw = request.form.get("region")
    if region_raw is None and request.is_json:
        payload = request.get_json(silent=True) or {}
        if isinstance(payload, dict):
            region_raw = payload.get("region")
    return username, region_raw


def resolve_principal(
    requested_user: str | None = None,
    requested_region: str | None = None,
) -> dict[str, Any]:
    access = load_access_control()
    requested_username = normalize_username(requested_user) or "owner"
    users = access.get("users", {})
    if requested_username not in users:
        requested_username = "owner"
    user_record = users.get(requested_username, {})
    role = normalize_user_role(user_record.get("role"), default=ROLE_USER)

    entries = discover_workbook_entries()
    regions_universe = all_known_regions(access, entries)
    allowed_regions = derive_allowed_regions(access, requested_username, role, regions_universe)
    allow_all_regions = role == ROLE_OWNER
    selected_region = resolve_selected_region(
        requested_region,
        allowed_regions,
        allow_all_regions=allow_all_regions,
    )

    asm_townships = access.get("asm_townships", {}).get(requested_username, {})
    allowed_townships = None
    if role == ROLE_ASM and isinstance(asm_townships, dict):
        region_for_townships = selected_region
        if region_for_townships == ALL_REGION_TOKEN and allowed_regions:
            region_for_townships = allowed_regions[0]
        allowed_townships = set(
            normalize_township_list(asm_townships.get(region_for_townships, []))
        )

    return {
        "username": requested_username,
        "role": role,
        "display_name": normalize_display_name(
            user_record.get("display_name"),
            requested_username,
        ),
        "access": access,
        "entries": entries,
        "regions_universe": regions_universe,
        "allowed_regions": allowed_regions,
        "selected_region": selected_region,
        "allow_all_regions": allow_all_regions,
        "allowed_townships": allowed_townships,
        "rsm_regions": normalize_region_list(
            access.get("rsm_regions", {}).get(requested_username, [])
        ),
    }


def principal_from_request() -> dict[str, Any]:
    requested_user, requested_region = current_request_user_and_region()
    return resolve_principal(
        requested_user=requested_user,
        requested_region=requested_region,
    )


def principal_can_upload_to_region(principal: dict[str, Any], region: str) -> bool:
    role = principal["role"]
    if role == ROLE_OWNER:
        return True
    if role == ROLE_RSM:
        return region in set(principal.get("rsm_regions", []))
    return False


def principal_can_manage_rsm(principal: dict[str, Any]) -> bool:
    return principal["role"] == ROLE_OWNER


def principal_can_manage_asm(principal: dict[str, Any]) -> bool:
    return principal["role"] in {ROLE_OWNER, ROLE_RSM}


def principal_can_manage_region(principal: dict[str, Any], region: str) -> bool:
    if principal["role"] == ROLE_OWNER:
        return True
    if principal["role"] == ROLE_RSM:
        return region in set(principal.get("rsm_regions", []))
    return False


def filter_data_by_allowed_townships(
    data: dict[str, Any],
    allowed_townships: set[str] | None,
) -> dict[str, Any]:
    if allowed_townships is None:
        return data

    filtered_main = {
        canonical: sheet
        for canonical, sheet in data["main"].items()
        if canonical in allowed_townships
    }
    filtered_reference = {
        canonical: sheet
        for canonical, sheet in data["reference"].items()
        if canonical in allowed_townships
    }

    allowed_main_names = {sheet["sheet_name"] for sheet in filtered_main.values()}
    allowed_reference_names = {sheet["sheet_name"] for sheet in filtered_reference.values()}

    filtered_main_tabs = [
        tab
        for tab in data["main_sheet_tabs"]
        if tab.get("canonical") in allowed_townships and tab.get("sheet_name") in allowed_main_names
    ]
    filtered_reference_tabs = [
        tab
        for tab in data["reference_sheet_tabs"]
        if tab.get("canonical") in allowed_townships
        and tab.get("sheet_name") in allowed_reference_names
    ]
    filtered_sheet_index = [
        row
        for row in data["sheet_index"]
        if row.get("canonical") in allowed_townships
    ]

    return {
        **data,
        "main": filtered_main,
        "reference": filtered_reference,
        "sheet_index": filtered_sheet_index,
        "main_sheet_tabs": filtered_main_tabs,
        "reference_sheet_tabs": filtered_reference_tabs,
        "main_sheet_names": normalize_sheet_name_list(
            [tab["sheet_name"] for tab in filtered_main_tabs]
        ),
        "reference_sheet_names": normalize_sheet_name_list(
            [tab["sheet_name"] for tab in filtered_reference_tabs]
        ),
    }

def discover_workbook_entries() -> list[dict[str, Any]]:
    registry = load_workbook_registry()
    entries: list[dict[str, Any]] = []
    seen_names: set[str] = set()
    search_dirs = [ROOT_DIR]
    if UPLOAD_DIR.exists():
        search_dirs.append(UPLOAD_DIR)

    for folder in search_dirs:
        for path in folder.iterdir():
            if not path.is_file():
                continue
            if path.name.startswith("~$"):
                continue
            if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
                continue
            if path.name in seen_names:
                # Keep first entry if duplicate filenames exist in different folders.
                continue

            meta = registry.get(path.name)
            region_token = (
                normalize_region_token(meta.get("region"))
                if meta and isinstance(meta, dict)
                else infer_region_from_filename(path)
            )
            seen_names.add(path.name)
            entries.append({"name": path.name, "path": path, "region": region_token})

    entries.sort(key=lambda item: str(item["name"]).lower())
    return entries


def workbook_name_score(path: Path, role: str) -> int:
    token = normalize_token(path.stem)
    if role == "main":
        keywords = ("overview", "main", "master", "mhl", "summary", "analysis")
    else:
        keywords = ("detail", "reference", "township", "breakdown", "summary", "mtl")

    score = 0
    for idx, keyword in enumerate(keywords):
        if keyword in token:
            score += max(1, 12 - (idx * 2))
    return score


def default_workbook(workbooks: list[Path], role: str, avoid_name: str | None = None) -> Path:
    candidates = [path for path in workbooks if path.name != avoid_name] or workbooks
    ranked = sorted(
        candidates,
        key=lambda path: (
            workbook_name_score(path, role),
            path.stat().st_mtime_ns,
            path.name.lower(),
        ),
        reverse=True,
    )
    return ranked[0]


def resolve_workbook_pair(
    main_name: str | None,
    reference_name: str | None,
    viewer_role: str | None = None,
    region_scope: str | None = None,
    allowed_regions: list[str] | None = None,
    allow_all_regions: bool | None = None,
) -> dict[str, Any]:
    entries = discover_workbook_entries()
    if not entries:
        raise FileNotFoundError("No supported Excel files found in this folder.")

    resolved_role = normalize_viewer_role(viewer_role)
    if allow_all_regions is None:
        allow_all_regions = resolved_role == ROLE_OWNER

    requested_region = normalize_region_scope(region_scope)

    scoped_entries = entries
    if allowed_regions is not None:
        allowed_region_set = {
            normalize_region_token(region)
            for region in allowed_regions
            if region is not None
        }
        scoped_entries = [
            entry for entry in entries if entry["region"] in allowed_region_set
        ]

    available_regions = sorted({str(entry["region"]) for entry in scoped_entries})
    selected_region = resolve_selected_region(
        requested_region,
        available_regions,
        allow_all_regions=allow_all_regions,
    )

    if selected_region == ALL_REGION_TOKEN and allow_all_regions:
        visible_entries = scoped_entries
    else:
        visible_entries = [
            entry for entry in scoped_entries if entry["region"] == selected_region
        ]

    if not visible_entries:
        if selected_region == ALL_REGION_TOKEN:
            raise FileNotFoundError("No supported Excel files found in this folder.")
        raise FileNotFoundError(
            f"No supported Excel files found for region: {selected_region}."
        )

    visible_paths = [entry["path"] for entry in visible_entries]
    by_name = {str(entry["name"]): entry["path"] for entry in visible_entries}
    default_main = default_workbook(visible_paths, "main")
    default_reference = default_workbook(
        visible_paths,
        "reference",
        avoid_name=default_main.name if len(visible_paths) > 1 else None,
    )

    if main_name and main_name not in by_name:
        raise ValueError(f"Unknown main workbook for current scope: {main_name}")
    if reference_name and reference_name not in by_name:
        raise ValueError(f"Unknown reference workbook for current scope: {reference_name}")

    main_path = by_name.get(main_name) if main_name else default_main
    reference_path = by_name.get(reference_name) if reference_name else default_reference

    return {
        "available": [str(entry["name"]) for entry in visible_entries],
        "regions": available_regions,
        "viewer_role": resolved_role,
        "selected_region": selected_region,
        "default_main": default_main.name,
        "default_reference": default_reference.name,
        "main_path": main_path,
        "reference_path": reference_path,
    }


def load_viewer_data(main_path: Path, reference_path: Path) -> dict[str, Any]:
    main_stat = main_path.stat()
    ref_stat = reference_path.stat()
    main_data = parse_workbook_cached(str(main_path), main_stat.st_mtime_ns)
    reference_data = parse_workbook_cached(str(reference_path), ref_stat.st_mtime_ns)
    version = f"{main_stat.st_mtime_ns}:{ref_stat.st_mtime_ns}"
    try:
        main_sheet_names = normalize_sheet_name_list(
            workbook_sheetnames_cached(str(main_path), main_stat.st_mtime_ns)
        )
    except Exception:
        main_sheet_names = normalize_sheet_name_list(
            [sheet["sheet_name"] for sheet in main_data.values()]
        )
    try:
        reference_sheet_names = normalize_sheet_name_list(
            workbook_sheetnames_cached(str(reference_path), ref_stat.st_mtime_ns)
        )
    except Exception:
        reference_sheet_names = normalize_sheet_name_list(
            [sheet["sheet_name"] for sheet in reference_data.values()]
        )

    sheet_index = []
    fixed_by_sheet_name: dict[str, dict[str, Any]] = {}
    for canonical, main_sheet in main_data.items():
        reference_sheet = reference_data.get(canonical)
        fixed_by_sheet_name[main_sheet["sheet_name"]] = {
            "canonical": canonical,
            "has_reference": bool(reference_sheet),
        }
        sheet_index.append(
            {
                "canonical": canonical,
                "main_sheet_name": main_sheet["sheet_name"],
                "reference_sheet_name": (
                    reference_sheet["sheet_name"] if reference_sheet else None
                ),
                "main_month_count": len(main_sheet["months"]),
                "reference_month_count": (
                    len(reference_sheet["months"]) if reference_sheet else 0
                ),
                "main_row_count": len(main_sheet["rows"]),
            }
        )

    main_sheet_tabs = []
    for sheet_name in main_sheet_names:
        fixed_info = fixed_by_sheet_name.get(sheet_name)
        main_sheet_tabs.append(
            {
                "sheet_name": sheet_name,
                "canonical": fixed_info["canonical"] if fixed_info else None,
                "filterable": bool(fixed_info),
                "has_reference": fixed_info["has_reference"] if fixed_info else False,
            }
        )

    fixed_reference_by_sheet_name: dict[str, dict[str, Any]] = {}
    for canonical, reference_sheet in reference_data.items():
        fixed_reference_by_sheet_name[reference_sheet["sheet_name"]] = {
            "canonical": canonical,
            "has_main": canonical in main_data,
        }

    reference_sheet_tabs = []
    for sheet_name in reference_sheet_names:
        fixed_info = fixed_reference_by_sheet_name.get(sheet_name)
        canonical = fixed_info["canonical"] if fixed_info else None
        reference_sheet_tabs.append(
            {
                "sheet_name": sheet_name,
                "canonical": canonical,
                "filterable": bool(fixed_info),
                "has_main": fixed_info["has_main"] if fixed_info else False,
            }
        )

    return {
        "main_workbook": main_path.name,
        "reference_workbook": reference_path.name,
        "version": version,
        "main_sheet_names": main_sheet_names,
        "reference_sheet_names": reference_sheet_names,
        "main": main_data,
        "reference": reference_data,
        "sheet_index": sheet_index,
        "main_sheet_tabs": main_sheet_tabs,
        "reference_sheet_tabs": reference_sheet_tabs,
    }


def color_to_css_hex(color) -> str | None:
    if not color:
        return None
    try:
        rgb = getattr(color, "rgb", None)
    except Exception:
        return None
    if not rgb:
        return None
    try:
        rgb = str(rgb)
    except Exception:
        return None
    if len(rgb) == 8:
        rgb = rgb[2:]
    if len(rgb) != 6:
        return None
    if not re.fullmatch(r"[0-9a-fA-F]{6}", rgb):
        return None
    return f"#{rgb.lower()}"


DEFAULT_TABLE_TEXT_COLOR = "#e7f4fb"
FALLBACK_READABLE_DARK_TEXT = "#102733"
FALLBACK_READABLE_LIGHT_TEXT = "#f5fbff"
MIN_TEXT_CONTRAST_RATIO = 4.5


def hex_to_rgb_triplet(css_hex: str) -> tuple[int, int, int] | None:
    if not css_hex or not re.fullmatch(r"#[0-9a-fA-F]{6}", css_hex):
        return None
    return (
        int(css_hex[1:3], 16),
        int(css_hex[3:5], 16),
        int(css_hex[5:7], 16),
    )


def relative_luminance(rgb: tuple[int, int, int]) -> float:
    def to_linear(channel: int) -> float:
        normalized = channel / 255
        if normalized <= 0.03928:
            return normalized / 12.92
        return ((normalized + 0.055) / 1.055) ** 2.4

    red = to_linear(rgb[0])
    green = to_linear(rgb[1])
    blue = to_linear(rgb[2])
    return 0.2126 * red + 0.7152 * green + 0.0722 * blue


def contrast_ratio(foreground_hex: str, background_hex: str) -> float:
    foreground_rgb = hex_to_rgb_triplet(foreground_hex)
    background_rgb = hex_to_rgb_triplet(background_hex)
    if foreground_rgb is None or background_rgb is None:
        return 1.0
    foreground_l = relative_luminance(foreground_rgb)
    background_l = relative_luminance(background_rgb)
    lighter = max(foreground_l, background_l)
    darker = min(foreground_l, background_l)
    return (lighter + 0.05) / (darker + 0.05)


def pick_accessible_text_color(background_hex: str, preferred_text_hex: str | None) -> str | None:
    current_text = preferred_text_hex or DEFAULT_TABLE_TEXT_COLOR
    if contrast_ratio(current_text, background_hex) >= MIN_TEXT_CONTRAST_RATIO:
        return preferred_text_hex

    dark_contrast = contrast_ratio(FALLBACK_READABLE_DARK_TEXT, background_hex)
    light_contrast = contrast_ratio(FALLBACK_READABLE_LIGHT_TEXT, background_hex)
    if dark_contrast >= light_contrast:
        return FALLBACK_READABLE_DARK_TEXT
    return FALLBACK_READABLE_LIGHT_TEXT


def border_side_css(side) -> str | None:
    if side is None:
        return None
    try:
        side_style = side.style
    except Exception:
        return None
    if side_style is None:
        return None
    width_map = {
        "hair": "1px",
        "thin": "1px",
        "medium": "2px",
        "thick": "3px",
        "dotted": "1px",
        "dashDot": "1px",
        "dashDotDot": "1px",
        "dashed": "1px",
        "double": "3px",
    }
    line_map = {
        "dotted": "dotted",
        "dashDot": "dashed",
        "dashDotDot": "dashed",
        "dashed": "dashed",
        "double": "double",
    }
    width = width_map.get(side_style, "1px")
    line_style = line_map.get(side_style, "solid")
    try:
        side_color = side.color
    except Exception:
        side_color = None
    color = color_to_css_hex(side_color) or "#6f8797"
    return f"{width} {line_style} {color}"


def cell_css(cell) -> str:
    rules: list[str] = []
    font_color: str | None = None
    fill_color: str | None = None

    try:
        alignment = cell.alignment
    except Exception:
        alignment = None
    if alignment:
        if alignment.horizontal:
            rules.append(f"text-align:{alignment.horizontal}")
        if alignment.vertical:
            rules.append(f"vertical-align:{alignment.vertical}")
        if alignment.wrap_text:
            rules.append("white-space:pre-wrap")

    try:
        font = cell.font
    except Exception:
        font = None
    if font:
        if font.bold:
            rules.append("font-weight:700")
        if font.italic:
            rules.append("font-style:italic")
        if font.underline and font.underline != "none":
            rules.append("text-decoration:underline")
        if font.size:
            rules.append(f"font-size:{font.size:.1f}pt")
        if font.name:
            rules.append(f"font-family:'{font.name}', sans-serif")
        try:
            font_color = color_to_css_hex(font.color)
        except Exception:
            font_color = None

    try:
        fill = cell.fill
    except Exception:
        fill = None
    if fill and fill.fill_type == "solid":
        try:
            fill_color = color_to_css_hex(fill.fgColor) or color_to_css_hex(fill.start_color)
        except Exception:
            fill_color = None
        if fill_color:
            rules.append(f"background-color:{fill_color}")

    try:
        border = cell.border
    except Exception:
        border = None
    if border:
        left = border_side_css(border.left)
        right = border_side_css(border.right)
        top = border_side_css(border.top)
        bottom = border_side_css(border.bottom)
        if left:
            rules.append(f"border-left:{left}")
        if right:
            rules.append(f"border-right:{right}")
        if top:
            rules.append(f"border-top:{top}")
        if bottom:
            rules.append(f"border-bottom:{bottom}")

    resolved_font_color = font_color
    if fill_color:
        resolved_font_color = pick_accessible_text_color(fill_color, font_color)
    if resolved_font_color:
        rules.append(f"color:{resolved_font_color}")

    return ";".join(rules)


def display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.4f}".rstrip("0").rstrip(".")
    return str(value)


def join_unique(items: list[int]) -> list[int]:
    seen: set[int] = set()
    ordered: list[int] = []
    for item in items:
        if item in seen:
            continue
        seen.add(item)
        ordered.append(item)
    return ordered


def parse_clamped_int(raw_value: str | None, default: int, min_value: int, max_value: int) -> int:
    try:
        parsed = int(raw_value) if raw_value is not None else default
    except (TypeError, ValueError):
        parsed = default
    return max(min_value, min(max_value, parsed))


def detect_month_column_groups(worksheet) -> list[dict[str, Any]]:
    header_scan_rows = min(4, worksheet.max_row)
    if header_scan_rows < 1:
        return []

    fallback_year = parse_year_context(worksheet, header_scan_rows + 1)
    anchors: list[dict[str, int | None]] = []

    for col_idx in range(1, worksheet.max_column + 1):
        parsed_month: tuple[int | None, int] | None = None
        for row_idx in range(1, header_scan_rows + 1):
            parsed_month = parse_month_header_loose(
                worksheet.cell(row=row_idx, column=col_idx).value,
                fallback_year,
            )
            if parsed_month:
                break
        if not parsed_month:
            continue
        year, month = parsed_month
        anchors.append({"col": col_idx, "year": year, "month": month})

    if not anchors:
        return []

    known_years = [item["year"] for item in anchors if item["year"] is not None]
    if known_years:
        first_year = int(known_years[0])
        for item in anchors:
            if item["year"] is None:
                item["year"] = first_year
    else:
        synthetic_year = 2000
        previous_month = None
        for item in anchors:
            current_month = int(item["month"])
            if previous_month is not None and current_month < previous_month:
                synthetic_year += 1
            item["year"] = synthetic_year
            previous_month = current_month

    diffs = [
        int(anchors[idx + 1]["col"]) - int(anchors[idx]["col"])
        for idx in range(len(anchors) - 1)
        if int(anchors[idx + 1]["col"]) > int(anchors[idx]["col"])
    ]
    default_width = Counter(diffs).most_common(1)[0][0] if diffs else 1
    if default_width < 1:
        default_width = 1

    groups_by_key: dict[str, dict[str, Any]] = {}
    for idx, anchor in enumerate(anchors):
        start_col = int(anchor["col"])
        if idx + 1 < len(anchors):
            end_col = int(anchors[idx + 1]["col"]) - 1
        else:
            end_col = min(worksheet.max_column, start_col + default_width - 1)
        if end_col < start_col:
            end_col = start_col

        year = int(anchor["year"])
        month = int(anchor["month"])
        key = f"{year:04d}-{month:02d}"
        label = f"{calendar.month_abbr[month]} {year}"

        group = groups_by_key.get(key)
        if group is None:
            group = {
                "key": key,
                "label": label,
                "year": year,
                "month": month,
                "cols": [],
                "start_col": start_col,
            }
            groups_by_key[key] = group

        group["cols"].extend(range(start_col, end_col + 1))
        group["start_col"] = min(group["start_col"], start_col)

    groups = sorted(groups_by_key.values(), key=lambda item: item["key"])
    for group in groups:
        group["cols"] = join_unique(
            [col for col in group["cols"] if 1 <= col <= worksheet.max_column]
        )
    return groups


def pick_month_keys_for_mode(
    month_groups: list[dict[str, Any]],
    mode: str,
    n_value: int,
    month_value: int | None,
) -> list[str]:
    if not month_groups:
        return []

    sorted_groups = sorted(month_groups, key=lambda item: item["key"])
    n_value = max(1, min(60, n_value))
    selected = sorted_groups

    if mode == "same_month_years":
        if month_value is None or month_value < 1 or month_value > 12:
            month_value = int(sorted_groups[-1]["month"])
        selected = [group for group in sorted_groups if int(group["month"]) == month_value]

    if not selected:
        return []
    return [item["key"] for item in selected[-n_value:]]


@lru_cache(maxsize=128)
def build_main_sheet_html_cached(
    path_str: str,
    mtime_ns: int,
    sheet_name: str,
    month_keys_csv: str,
    mode: str,
    n_value: int,
    month_value: int,
) -> dict[str, Any]:
    workbook = load_workbook_cached(path_str, mtime_ns)
    worksheet = workbook[sheet_name]

    layout = find_header_layout(worksheet)
    detected_month_groups = detect_month_column_groups(worksheet)
    selected_month_labels: list[str] = []
    available_months = sorted({int(group["month"]) for group in detected_month_groups})
    frozen_count = 0

    if detected_month_groups:
        month_by_key = {item["key"]: item for item in detected_month_groups}
        selected_month_keys = [key for key in month_keys_csv.split(",") if key in month_by_key]

        if not selected_month_keys:
            selected_month_keys = pick_month_keys_for_mode(
                detected_month_groups,
                mode,
                n_value,
                month_value if month_value > 0 else None,
            )

        if not selected_month_keys:
            selected_month_keys = [detected_month_groups[-1]["key"]]

        first_month_col = min(
            int(group["start_col"])
            for group in detected_month_groups
            if group.get("cols")
        )
        frozen_count = max(0, first_month_col - 1)
        selected_columns: list[int] = list(range(1, first_month_col))

        for key in selected_month_keys:
            group = month_by_key.get(key)
            if not group:
                continue
            selected_columns.extend(group["cols"])
            selected_month_labels.append(group["label"])

        selected_columns = join_unique(
            [col for col in selected_columns if 1 <= col <= worksheet.max_column]
        )
        min_required_row = 1
    elif layout:
        # Sheet matches fixed format layout but no month groups were detected.
        selected_columns = list(range(1, layout["packing"] + 1))
        frozen_count = len(selected_columns)
        min_required_row = layout["metrics_row"] + 1
    else:
        used_columns: list[int] = []
        for col_idx in range(1, worksheet.max_column + 1):
            has_value = any(
                worksheet.cell(row=row_idx, column=col_idx).value not in (None, "")
                for row_idx in range(1, worksheet.max_row + 1)
            )
            if has_value:
                used_columns.append(col_idx)

        if not used_columns:
            return {
                "sheet_name": sheet_name,
                "row_count": 0,
                "col_count": 0,
                "frozen_count": 0,
                "selected_month_labels": [],
                "available_months": [],
                "html": '<div class="empty">No data in this sheet.</div>',
            }

        if len(used_columns) > 70:
            first_cols = used_columns[:20]
            last_cols = used_columns[-20:]
            header_cols = [
                col_idx
                for col_idx in used_columns
                if any(
                    worksheet.cell(row=row_idx, column=col_idx).value not in (None, "")
                    for row_idx in range(1, min(worksheet.max_row, 4) + 1)
                )
            ]
            selected_columns = join_unique(first_cols + header_cols + last_cols)
        else:
            selected_columns = used_columns

        min_required_row = 1

    selected_col_set = set(selected_columns)

    last_row = 1
    for row_idx in range(1, worksheet.max_row + 1):
        row_has_value = False
        for col_idx in selected_columns:
            value = worksheet.cell(row=row_idx, column=col_idx).value
            if value not in (None, ""):
                row_has_value = True
                break
        if row_has_value:
            last_row = row_idx
    last_row = max(last_row, min_required_row)

    merge_top_left: dict[tuple[int, int], tuple[int, int]] = {}
    merge_skip: set[tuple[int, int]] = set()
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.max_row > last_row:
            continue
        fully_selected = all(
            col in selected_col_set
            for col in range(merged_range.min_col, merged_range.max_col + 1)
        )
        if not fully_selected:
            continue
        top_left = (merged_range.min_row, merged_range.min_col)
        merge_top_left[top_left] = (
            merged_range.max_row - merged_range.min_row + 1,
            merged_range.max_col - merged_range.min_col + 1,
        )
        for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
            for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                if (row_idx, col_idx) != top_left:
                    merge_skip.add((row_idx, col_idx))

    style_to_class: dict[int, str] = {}
    style_rules: list[str] = []

    def style_class_for(cell) -> str:
        try:
            style_id = int(cell.style_id)
        except Exception:
            style_id = -1
        if style_id in style_to_class:
            return style_to_class[style_id]
        class_name = f"sx{len(style_to_class)}"
        style_to_class[style_id] = class_name
        css = cell_css(cell)
        style_rules.append(f".main-source-table td.{class_name}{{{css}}}")
        return class_name

    colgroup_html: list[str] = []
    for col_idx in selected_columns:
        col_dim = worksheet.column_dimensions.get(get_column_letter(col_idx))
        width = getattr(col_dim, "width", None)
        if width:
            pixel_width = max(44, int(width * 7))
            colgroup_html.append(f'<col style="width:{pixel_width}px">')
        else:
            colgroup_html.append("<col>")

    rows_html: list[str] = []
    for row_idx in range(1, last_row + 1):
        height = worksheet.row_dimensions[row_idx].height
        row_style = f' style="height:{int(height)}px"' if height else ""
        cell_html_parts: list[str] = []
        for col_idx in selected_columns:
            if (row_idx, col_idx) in merge_skip:
                continue
            cell = worksheet.cell(row=row_idx, column=col_idx)
            class_name = style_class_for(cell)
            attrs = [f'class="{class_name}"']
            if (row_idx, col_idx) in merge_top_left:
                rowspan, colspan = merge_top_left[(row_idx, col_idx)]
                if rowspan > 1:
                    attrs.append(f'rowspan="{rowspan}"')
                if colspan > 1:
                    attrs.append(f'colspan="{colspan}"')

            text = html.escape(display_value(cell.value)).replace("\n", "<br>")
            cell_html_parts.append(f"<td {' '.join(attrs)}>{text}</td>")

        rows_html.append(f"<tr{row_style}>{''.join(cell_html_parts)}</tr>")

    inline_styles = "".join(style_rules)
    base_style = (
        "<style>"
        ".main-source-table{border-collapse:collapse;width:max-content;min-width:100%;}"
        ".main-source-table td{padding:6px 8px;white-space:nowrap;border:1px solid rgba(133,184,208,0.22);}"
        f"{inline_styles}"
        "</style>"
    )
    table_html = (
        f"{base_style}<table class=\"main-source-table\">"
        f"<colgroup>{''.join(colgroup_html)}</colgroup>"
        f"<tbody>{''.join(rows_html)}</tbody>"
        "</table>"
    )

    return {
        "sheet_name": sheet_name,
        "row_count": last_row,
        "col_count": len(selected_columns),
        "frozen_count": min(frozen_count, len(selected_columns)),
        "selected_month_labels": selected_month_labels,
        "available_months": available_months,
        "html": table_html,
    }


def clear_workbook_caches() -> None:
    parse_workbook_cached.cache_clear()
    load_workbook_cached.cache_clear()
    workbook_sheetnames_cached.cache_clear()
    build_main_sheet_html_cached.cache_clear()


def safe_uploaded_filename(original_name: str) -> str:
    path = Path(original_name)
    stem = re.sub(r"[^A-Za-z0-9._ -]+", "_", path.stem).strip(" ._")
    if not stem:
        stem = "workbook"
    suffix = path.suffix.lower()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    return f"{stem}__{timestamp}{suffix}"


def upload_region_for_file(upload_region: str | None, original_name: str) -> str:
    scoped_region = normalize_region_scope(upload_region)
    if scoped_region != ALL_REGION_TOKEN:
        return scoped_region
    return infer_region_from_filename(Path(original_name))


@app.route("/")
def index() -> str:
    return render_template("index.html", static_version=int(time.time()))


@app.route("/api/workbooks")
def api_workbooks():
    try:
        selection = resolve_workbook_pair(
            None,
            None,
            viewer_role=request.args.get("role"),
            region_scope=request.args.get("region"),
        )
    except FileNotFoundError as exc:
        return jsonify({"error": str(exc), "workbooks": []}), 404

    return jsonify(
        {
            "workbooks": selection["available"],
            "regions": selection["regions"],
            "viewer_role": selection["viewer_role"],
            "selected_region": selection["selected_region"],
            "default_main": selection["default_main"],
            "default_reference": selection["default_reference"],
        }
    )


@app.route("/api/upload-workbooks", methods=["POST"])
def api_upload_workbooks():
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "No Excel files were uploaded."}), 400

    viewer_role = request.form.get("role")
    viewer_region = request.form.get("region")
    upload_region = request.form.get("upload_region")

    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    saved_files: list[str] = []
    uploaded_regions: list[dict[str, str]] = []
    skipped_files: list[str] = []
    registry = load_workbook_registry()

    for file_storage in files:
        original_name = (file_storage.filename or "").strip()
        if not original_name:
            continue

        suffix = Path(original_name).suffix.lower()
        if suffix not in SUPPORTED_EXTENSIONS:
            skipped_files.append(original_name)
            continue

        output_name = safe_uploaded_filename(original_name)
        file_storage.save(UPLOAD_DIR / output_name)
        saved_files.append(output_name)
        region = upload_region_for_file(upload_region, original_name)
        uploaded_regions.append({"file": output_name, "region": region})
        registry[output_name] = {"region": region}

    if not saved_files:
        return (
            jsonify(
                {
                    "error": "No supported Excel files were uploaded.",
                    "skipped_files": skipped_files,
                }
            ),
            400,
        )

    save_workbook_registry(registry)
    clear_workbook_caches()
    selection = resolve_workbook_pair(
        None,
        None,
        viewer_role=viewer_role,
        region_scope=viewer_region,
    )
    return jsonify(
        {
            "uploaded_files": saved_files,
            "uploaded_regions": uploaded_regions,
            "skipped_files": skipped_files,
            "workbooks": selection["available"],
            "regions": selection["regions"],
            "viewer_role": selection["viewer_role"],
            "selected_region": selection["selected_region"],
            "default_main": selection["default_main"],
            "default_reference": selection["default_reference"],
        }
    )


@app.route("/api/sheets")
def api_sheets():
    try:
        selection = resolve_workbook_pair(
            request.args.get("main"),
            request.args.get("reference"),
            viewer_role=request.args.get("role"),
            region_scope=request.args.get("region"),
        )
    except FileNotFoundError as exc:
        return jsonify({"error": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    data = load_viewer_data(selection["main_path"], selection["reference_path"])
    return jsonify(
        {
            "main_workbook": data["main_workbook"],
            "reference_workbook": data["reference_workbook"],
            "version": data["version"],
            "available_workbooks": selection["available"],
            "main_sheet_names": data["main_sheet_names"],
            "reference_sheet_names": data["reference_sheet_names"],
            "regions": selection["regions"],
            "viewer_role": selection["viewer_role"],
            "selected_region": selection["selected_region"],
            "sheets": data["sheet_index"],
            "main_sheet_tabs": data["main_sheet_tabs"],
            "reference_sheet_tabs": data["reference_sheet_tabs"],
        }
    )


@app.route("/api/sheet/<canonical>")
def api_sheet(canonical: str):
    try:
        selection = resolve_workbook_pair(
            request.args.get("main"),
            request.args.get("reference"),
            viewer_role=request.args.get("role"),
            region_scope=request.args.get("region"),
        )
    except FileNotFoundError as exc:
        return jsonify({"error": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    data = load_viewer_data(selection["main_path"], selection["reference_path"])
    if canonical not in data["main"] and canonical not in data["reference"]:
        abort(404, description=f"Unknown sheet: {canonical}")

    return jsonify(
        {
            "canonical": canonical,
            "version": data["version"],
            "main": data["main"].get(canonical),
            "reference": data["reference"].get(canonical),
        }
    )


@app.route("/api/main-styled/<canonical>")
def api_main_styled(canonical: str):
    try:
        selection = resolve_workbook_pair(
            request.args.get("main"),
            request.args.get("reference"),
            viewer_role=request.args.get("role"),
            region_scope=request.args.get("region"),
        )
    except FileNotFoundError as exc:
        return jsonify({"error": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    data = load_viewer_data(selection["main_path"], selection["reference_path"])
    if canonical not in data["main"]:
        abort(404, description=f"Unknown sheet: {canonical}")

    month_keys_csv = request.args.get("month_keys", "")
    mode = request.args.get("mode", "past_months")
    if mode not in {"past_months", "same_month_years"}:
        mode = "past_months"
    n_value = parse_clamped_int(request.args.get("n"), default=6, min_value=1, max_value=60)
    month_value = parse_clamped_int(request.args.get("month"), default=0, min_value=0, max_value=12)
    main_stat = selection["main_path"].stat()
    sheet_name = data["main"][canonical]["sheet_name"]
    html_payload = build_main_sheet_html_cached(
        str(selection["main_path"]),
        main_stat.st_mtime_ns,
        sheet_name,
        month_keys_csv,
        mode,
        n_value,
        month_value,
    )
    return jsonify(
        {
            "version": data["version"],
            "canonical": canonical,
            "filterable": True,
            "sheet_name": html_payload["sheet_name"],
            "row_count": html_payload["row_count"],
            "col_count": html_payload["col_count"],
            "frozen_count": html_payload["frozen_count"],
            "selected_month_labels": html_payload["selected_month_labels"],
            "available_months": html_payload["available_months"],
            "html": html_payload["html"],
        }
    )


@app.route("/api/main-styled-sheet")
def api_main_styled_sheet():
    try:
        selection = resolve_workbook_pair(
            request.args.get("main"),
            request.args.get("reference"),
            viewer_role=request.args.get("role"),
            region_scope=request.args.get("region"),
        )
    except FileNotFoundError as exc:
        return jsonify({"error": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    data = load_viewer_data(selection["main_path"], selection["reference_path"])
    sheet_name = request.args.get("sheet")
    if not sheet_name:
        return jsonify({"error": "Missing required query parameter: sheet"}), 400

    tab_match = next(
        (tab for tab in data["main_sheet_tabs"] if tab["sheet_name"] == sheet_name),
        None,
    )
    if not tab_match:
        abort(404, description=f"Unknown main sheet: {sheet_name}")

    month_keys_csv = request.args.get("month_keys", "")
    mode = request.args.get("mode", "past_months")
    if mode not in {"past_months", "same_month_years"}:
        mode = "past_months"
    n_value = parse_clamped_int(request.args.get("n"), default=6, min_value=1, max_value=60)
    month_value = parse_clamped_int(request.args.get("month"), default=0, min_value=0, max_value=12)
    main_stat = selection["main_path"].stat()
    html_payload = build_main_sheet_html_cached(
        str(selection["main_path"]),
        main_stat.st_mtime_ns,
        sheet_name,
        month_keys_csv,
        mode,
        n_value,
        month_value,
    )
    return jsonify(
        {
            "version": data["version"],
            "canonical": tab_match["canonical"],
            "filterable": tab_match["filterable"],
            "sheet_name": html_payload["sheet_name"],
            "row_count": html_payload["row_count"],
            "col_count": html_payload["col_count"],
            "frozen_count": html_payload["frozen_count"],
            "selected_month_labels": html_payload["selected_month_labels"],
            "available_months": html_payload["available_months"],
            "html": html_payload["html"],
        }
    )


def pick_available_port(candidates: list[int]) -> int:
    for port in candidates:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            try:
                sock.bind(("127.0.0.1", port))
                return port
            except OSError:
                continue
    raise RuntimeError("No available port found in candidate list.")


def resolve_runtime_host_port() -> tuple[str, int]:
    render_port = os.getenv("PORT")
    if render_port:
        try:
            return "0.0.0.0", int(render_port)
        except ValueError as exc:
            raise RuntimeError("PORT environment variable must be an integer.") from exc

    return "127.0.0.1", pick_available_port([5055, 8000, 8080, 5000])


if __name__ == "__main__":
    host, port = resolve_runtime_host_port()
    print(f"Starting viewer at http://{host}:{port}")
    app.run(host=host, port=port, debug=False)
