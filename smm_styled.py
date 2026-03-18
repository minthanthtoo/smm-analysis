from __future__ import annotations

import calendar
import html
import re
from collections import Counter
from datetime import date, datetime
from functools import lru_cache
from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

from smm_services import (
    find_header_layout,
    load_workbook_cached,
    parse_month_header_loose,
    parse_year_context,
)

def color_to_css_hex(color) -> str | None:
    if not color:
        return None
    try:
        rgb = getattr(color, "rgb", None)
    except Exception:
        return None
    if not rgb:
        return None
    try:
        rgb = str(rgb)
    except Exception:
        return None
    if len(rgb) == 8:
        rgb = rgb[2:]
    if len(rgb) != 6:
        return None
    if not re.fullmatch(r"[0-9a-fA-F]{6}", rgb):
        return None
    return f"#{rgb.lower()}"


DEFAULT_TABLE_TEXT_COLOR = "#e7f4fb"
FALLBACK_READABLE_DARK_TEXT = "#102733"
FALLBACK_READABLE_LIGHT_TEXT = "#f5fbff"
MIN_TEXT_CONTRAST_RATIO = 4.5


def hex_to_rgb_triplet(css_hex: str) -> tuple[int, int, int] | None:
    if not css_hex or not re.fullmatch(r"#[0-9a-fA-F]{6}", css_hex):
        return None
    return (
        int(css_hex[1:3], 16),
        int(css_hex[3:5], 16),
        int(css_hex[5:7], 16),
    )


def relative_luminance(rgb: tuple[int, int, int]) -> float:
    def to_linear(channel: int) -> float:
        normalized = channel / 255
        if normalized <= 0.03928:
            return normalized / 12.92
        return ((normalized + 0.055) / 1.055) ** 2.4

    red = to_linear(rgb[0])
    green = to_linear(rgb[1])
    blue = to_linear(rgb[2])
    return 0.2126 * red + 0.7152 * green + 0.0722 * blue


def contrast_ratio(foreground_hex: str, background_hex: str) -> float:
    foreground_rgb = hex_to_rgb_triplet(foreground_hex)
    background_rgb = hex_to_rgb_triplet(background_hex)
    if foreground_rgb is None or background_rgb is None:
        return 1.0
    foreground_l = relative_luminance(foreground_rgb)
    background_l = relative_luminance(background_rgb)
    lighter = max(foreground_l, background_l)
    darker = min(foreground_l, background_l)
    return (lighter + 0.05) / (darker + 0.05)


def pick_accessible_text_color(background_hex: str, preferred_text_hex: str | None) -> str | None:
    current_text = preferred_text_hex or DEFAULT_TABLE_TEXT_COLOR
    if contrast_ratio(current_text, background_hex) >= MIN_TEXT_CONTRAST_RATIO:
        return preferred_text_hex

    dark_contrast = contrast_ratio(FALLBACK_READABLE_DARK_TEXT, background_hex)
    light_contrast = contrast_ratio(FALLBACK_READABLE_LIGHT_TEXT, background_hex)
    if dark_contrast >= light_contrast:
        return FALLBACK_READABLE_DARK_TEXT
    return FALLBACK_READABLE_LIGHT_TEXT


def border_side_css(side) -> str | None:
    if side is None:
        return None
    try:
        side_style = side.style
    except Exception:
        return None
    if side_style is None:
        return None
    width_map = {
        "hair": "1px",
        "thin": "1px",
        "medium": "2px",
        "thick": "3px",
        "dotted": "1px",
        "dashDot": "1px",
        "dashDotDot": "1px",
        "dashed": "1px",
        "double": "3px",
    }
    line_map = {
        "dotted": "dotted",
        "dashDot": "dashed",
        "dashDotDot": "dashed",
        "dashed": "dashed",
        "double": "double",
    }
    width = width_map.get(side_style, "1px")
    line_style = line_map.get(side_style, "solid")
    try:
        side_color = side.color
    except Exception:
        side_color = None
    color = color_to_css_hex(side_color) or "#6f8797"
    return f"{width} {line_style} {color}"


def cell_css(cell) -> str:
    rules: list[str] = []
    font_color: str | None = None
    fill_color: str | None = None

    try:
        alignment = cell.alignment
    except Exception:
        alignment = None
    if alignment:
        if alignment.horizontal:
            rules.append(f"text-align:{alignment.horizontal}")
        if alignment.vertical:
            rules.append(f"vertical-align:{alignment.vertical}")
        if alignment.wrap_text:
            rules.append("white-space:pre-wrap")

    try:
        font = cell.font
    except Exception:
        font = None
    if font:
        if font.bold:
            rules.append("font-weight:700")
        if font.italic:
            rules.append("font-style:italic")
        if font.underline and font.underline != "none":
            rules.append("text-decoration:underline")
        if font.size:
            rules.append(f"font-size:{font.size:.1f}pt")
        if font.name:
            rules.append(f"font-family:'{font.name}', sans-serif")
        try:
            font_color = color_to_css_hex(font.color)
        except Exception:
            font_color = None

    try:
        fill = cell.fill
    except Exception:
        fill = None
    if fill and fill.fill_type == "solid":
        try:
            fill_color = color_to_css_hex(fill.fgColor) or color_to_css_hex(fill.start_color)
        except Exception:
            fill_color = None
        if fill_color:
            rules.append(f"background-color:{fill_color}")

    try:
        border = cell.border
    except Exception:
        border = None
    if border:
        left = border_side_css(border.left)
        right = border_side_css(border.right)
        top = border_side_css(border.top)
        bottom = border_side_css(border.bottom)
        if left:
            rules.append(f"border-left:{left}")
        if right:
            rules.append(f"border-right:{right}")
        if top:
            rules.append(f"border-top:{top}")
        if bottom:
            rules.append(f"border-bottom:{bottom}")

    resolved_font_color = font_color
    if fill_color:
        resolved_font_color = pick_accessible_text_color(fill_color, font_color)
    if resolved_font_color:
        rules.append(f"color:{resolved_font_color}")

    return ";".join(rules)


def display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.4f}".rstrip("0").rstrip(".")
    return str(value)


def join_unique(items: list[int]) -> list[int]:
    seen: set[int] = set()
    ordered: list[int] = []
    for item in items:
        if item in seen:
            continue
        seen.add(item)
        ordered.append(item)
    return ordered


def parse_clamped_int(raw_value: str | None, default: int, min_value: int, max_value: int) -> int:
    try:
        parsed = int(raw_value) if raw_value is not None else default
    except (TypeError, ValueError):
        parsed = default
    return max(min_value, min(max_value, parsed))


def parse_month_values_csv(month_values_csv: str | None) -> tuple[int, ...]:
    if not month_values_csv:
        return ()
    month_values: set[int] = set()
    for token in str(month_values_csv).split(","):
        try:
            month_value = int(token)
        except (TypeError, ValueError):
            continue
        if 1 <= month_value <= 12:
            month_values.add(month_value)
    if not month_values:
        return ()
    return tuple(sorted(month_values))


def points_to_css_px(points_value: Any, fallback_px: int | None = None) -> int | None:
    try:
        points = float(points_value)
    except (TypeError, ValueError):
        points = None
    if points is None or points <= 0:
        return fallback_px
    pixels = int(round(points * 96.0 / 72.0))
    return max(1, pixels)


def freeze_counts_from_worksheet(worksheet) -> tuple[int, int]:
    freeze_marker = getattr(worksheet, "freeze_panes", None)
    if not freeze_marker:
        return 0, 0

    coordinate = getattr(freeze_marker, "coordinate", None) or str(freeze_marker)
    if not coordinate:
        return 0, 0

    coordinate = str(coordinate).upper()
    if coordinate == "A1":
        return 0, 0

    try:
        column_letters, row_idx = coordinate_from_string(coordinate)
        col_idx = column_index_from_string(column_letters)
        row_idx = int(row_idx)
    except Exception:
        return 0, 0

    return max(0, row_idx - 1), max(0, col_idx - 1)


def is_column_hidden(worksheet, col_idx: int) -> bool:
    col_dim = worksheet.column_dimensions.get(get_column_letter(col_idx))
    return bool(getattr(col_dim, "hidden", False))


def is_row_hidden(worksheet, row_idx: int) -> bool:
    row_dim = worksheet.row_dimensions.get(row_idx)
    return bool(getattr(row_dim, "hidden", False))


def detect_month_column_groups(worksheet) -> list[dict[str, Any]]:
    header_scan_rows = min(4, worksheet.max_row)
    if header_scan_rows < 1:
        return []

    fallback_year = parse_year_context(worksheet, header_scan_rows + 1)
    anchors: list[dict[str, int | None]] = []

    for col_idx in range(1, worksheet.max_column + 1):
        parsed_month: tuple[int | None, int] | None = None
        for row_idx in range(1, header_scan_rows + 1):
            parsed_month = parse_month_header_loose(
                worksheet.cell(row=row_idx, column=col_idx).value,
                fallback_year,
            )
            if parsed_month:
                break
        if not parsed_month:
            continue
        year, month = parsed_month
        anchors.append({"col": col_idx, "year": year, "month": month})

    if not anchors:
        return []

    known_years = [item["year"] for item in anchors if item["year"] is not None]
    if known_years:
        first_year = int(known_years[0])
        for item in anchors:
            if item["year"] is None:
                item["year"] = first_year
    else:
        synthetic_year = 2000
        previous_month = None
        for item in anchors:
            current_month = int(item["month"])
            if previous_month is not None and current_month < previous_month:
                synthetic_year += 1
            item["year"] = synthetic_year
            previous_month = current_month

    diffs = [
        int(anchors[idx + 1]["col"]) - int(anchors[idx]["col"])
        for idx in range(len(anchors) - 1)
        if int(anchors[idx + 1]["col"]) > int(anchors[idx]["col"])
    ]
    default_width = Counter(diffs).most_common(1)[0][0] if diffs else 1
    if default_width < 1:
        default_width = 1

    groups_by_key: dict[str, dict[str, Any]] = {}
    for idx, anchor in enumerate(anchors):
        start_col = int(anchor["col"])
        if idx + 1 < len(anchors):
            end_col = int(anchors[idx + 1]["col"]) - 1
        else:
            end_col = min(worksheet.max_column, start_col + default_width - 1)
        if end_col < start_col:
            end_col = start_col

        year = int(anchor["year"])
        month = int(anchor["month"])
        key = f"{year:04d}-{month:02d}"
        label = f"{calendar.month_abbr[month]} {year}"

        group = groups_by_key.get(key)
        if group is None:
            group = {
                "key": key,
                "label": label,
                "year": year,
                "month": month,
                "cols": [],
                "start_col": start_col,
            }
            groups_by_key[key] = group

        group["cols"].extend(range(start_col, end_col + 1))
        group["start_col"] = min(group["start_col"], start_col)

    groups = sorted(groups_by_key.values(), key=lambda item: item["key"])
    for group in groups:
        group["cols"] = join_unique(
            [col for col in group["cols"] if 1 <= col <= worksheet.max_column]
        )
    return groups


def pick_month_keys_for_mode(
    month_groups: list[dict[str, Any]],
    mode: str,
    n_value: int,
    month_value: int | None,
    month_values: tuple[int, ...] | None = None,
) -> list[str]:
    if not month_groups:
        return []

    sorted_groups = sorted(month_groups, key=lambda item: item["key"])
    n_value = max(1, min(60, n_value))
    selected = sorted_groups

    if mode == "same_month_years":
        if month_value is None or month_value < 1 or month_value > 12:
            month_value = int(sorted_groups[-1]["month"])
        selected = [group for group in sorted_groups if int(group["month"]) == month_value]
    elif mode == "multi_month_years":
        valid_months = [
            int(value)
            for value in (month_values or ())
            if isinstance(value, int) and 1 <= int(value) <= 12
        ]
        if not valid_months:
            valid_months = [int(sorted_groups[-1]["month"])]

        picked: list[dict[str, Any]] = []
        for month in valid_months:
            month_groups_for_value = [
                group for group in sorted_groups if int(group["month"]) == month
            ]
            picked.extend(month_groups_for_value[-n_value:])
        selected = sorted(picked, key=lambda item: item["key"])

    if not selected:
        return []
    if mode == "multi_month_years":
        return [item["key"] for item in selected]
    return [item["key"] for item in selected[-n_value:]]


@lru_cache(maxsize=128)
def build_main_sheet_html_cached(
    path_str: str,
    mtime_ns: int,
    sheet_name: str,
    month_keys_csv: str,
    mode: str,
    n_value: int,
    month_value: int,
    month_values_csv: str = "",
) -> dict[str, Any]:
    workbook = load_workbook_cached(path_str, mtime_ns)
    worksheet = workbook[sheet_name]

    layout = find_header_layout(worksheet)
    detected_month_groups = detect_month_column_groups(worksheet)
    selected_month_labels: list[str] = []
    selected_month_keys: list[str] = []
    available_months = sorted({int(group["month"]) for group in detected_month_groups})
    worksheet_frozen_rows, worksheet_frozen_cols = freeze_counts_from_worksheet(worksheet)
    has_explicit_freeze = worksheet_frozen_rows > 0 or worksheet_frozen_cols > 0
    frozen_row_boundary = worksheet_frozen_rows
    frozen_col_boundary = worksheet_frozen_cols
    is_generic_fallback_sheet = False

    if detected_month_groups:
        month_by_key = {item["key"]: item for item in detected_month_groups}
        selected_month_keys = [key for key in month_keys_csv.split(",") if key in month_by_key]

        if not selected_month_keys:
            selected_month_keys = pick_month_keys_for_mode(
                detected_month_groups,
                mode,
                n_value,
                month_value if month_value > 0 else None,
                parse_month_values_csv(month_values_csv),
            )

        if not selected_month_keys:
            selected_month_keys = [detected_month_groups[-1]["key"]]

        first_month_col = min(
            int(group["start_col"])
            for group in detected_month_groups
            if group.get("cols")
        )
        if not has_explicit_freeze:
            frozen_col_boundary = max(0, first_month_col - 1)
        selected_columns: list[int] = list(range(1, first_month_col))

        for key in selected_month_keys:
            group = month_by_key.get(key)
            if not group:
                continue
            selected_columns.extend(group["cols"])

        selected_columns = join_unique(
            [col for col in selected_columns if 1 <= col <= worksheet.max_column]
        )
        min_required_row = 1
    elif layout:
        # Sheet matches fixed format layout but no month groups were detected.
        selected_columns = list(range(1, layout["packing"] + 1))
        if not has_explicit_freeze:
            frozen_col_boundary = len(selected_columns)
        min_required_row = layout["metrics_row"] + 1
    else:
        is_generic_fallback_sheet = True
        used_columns: list[int] = []
        for col_idx in range(1, worksheet.max_column + 1):
            has_value = any(
                worksheet.cell(row=row_idx, column=col_idx).value not in (None, "")
                for row_idx in range(1, worksheet.max_row + 1)
            )
            if has_value:
                used_columns.append(col_idx)

        if not used_columns:
            return {
                "sheet_name": sheet_name,
                "row_count": 0,
                "col_count": 0,
                "frozen_count": 0,
                "selected_month_labels": [],
                "available_months": [],
                "html": '<div class="empty">No data in this sheet.</div>',
            }

        if len(used_columns) > 70:
            first_cols = used_columns[:20]
            last_cols = used_columns[-20:]
            header_cols = [
                col_idx
                for col_idx in used_columns
                if any(
                    worksheet.cell(row=row_idx, column=col_idx).value not in (None, "")
                    for row_idx in range(1, min(worksheet.max_row, 4) + 1)
                )
            ]
            selected_columns = join_unique(first_cols + header_cols + last_cols)
        else:
            selected_columns = used_columns

        min_required_row = 1

    hidden_columns_skipped = sum(
        1
        for col_idx in selected_columns
        if 1 <= col_idx <= worksheet.max_column and is_column_hidden(worksheet, col_idx)
    )
    selected_columns = join_unique(
        [
            col_idx
            for col_idx in selected_columns
            if 1 <= col_idx <= worksheet.max_column and not is_column_hidden(worksheet, col_idx)
        ]
    )
    if is_generic_fallback_sheet and selected_columns:
        # Avoid collapsing sparse sheets into a 1x1 view; keep a minimum visible grid footprint.
        min_visible_columns = 12
        if len(selected_columns) < min_visible_columns:
            max_selected_col = max(selected_columns)
            padded_target_col = max(min_visible_columns, max_selected_col)
            selected_columns = [
                col_idx
                for col_idx in range(1, padded_target_col + 1)
                if not is_column_hidden(worksheet, col_idx)
            ]
    if detected_month_groups and selected_month_keys:
        visible_column_set = set(selected_columns)
        month_by_key = {item["key"]: item for item in detected_month_groups}
        selected_month_labels = [
            month_by_key[key]["label"]
            for key in selected_month_keys
            if key in month_by_key and any(col_idx in visible_column_set for col_idx in month_by_key[key]["cols"])
        ]
    if not selected_columns:
        return {
            "sheet_name": sheet_name,
            "row_count": 0,
            "col_count": 0,
            "frozen_count": 0,
            "frozen_rows": 0,
            "frozen_columns": 0,
            "selected_month_labels": selected_month_labels,
            "available_months": available_months,
            "hidden_rows_skipped": 0,
            "hidden_columns_skipped": hidden_columns_skipped,
            "html": '<div class="empty">All selected columns are hidden in this sheet.</div>',
        }

    visible_row_candidates = [
        row_idx for row_idx in range(1, worksheet.max_row + 1) if not is_row_hidden(worksheet, row_idx)
    ]
    if not visible_row_candidates:
        return {
            "sheet_name": sheet_name,
            "row_count": 0,
            "col_count": len(selected_columns),
            "frozen_count": 0,
            "frozen_rows": 0,
            "frozen_columns": 0,
            "selected_month_labels": selected_month_labels,
            "available_months": available_months,
            "hidden_rows_skipped": worksheet.max_row,
            "hidden_columns_skipped": hidden_columns_skipped,
            "html": '<div class="empty">All rows in this sheet are hidden.</div>',
        }

    last_row = 0
    for row_idx in visible_row_candidates:
        row_has_value = False
        for col_idx in selected_columns:
            value = worksheet.cell(row=row_idx, column=col_idx).value
            if value not in (None, ""):
                row_has_value = True
                break
        if row_has_value:
            last_row = row_idx

    if last_row < min_required_row:
        last_row = next((row_idx for row_idx in visible_row_candidates if row_idx >= min_required_row), visible_row_candidates[-1])

    if is_generic_fallback_sheet and last_row < 24:
        last_row = 24

    visible_rows = [row_idx for row_idx in range(1, last_row + 1) if not is_row_hidden(worksheet, row_idx)]
    visible_row_set = set(visible_rows)
    hidden_rows_skipped = max(0, last_row - len(visible_rows))

    sparse_sheet = len(visible_rows) <= 2 and len(selected_columns) <= 8
    if sparse_sheet:
        min_sparse_rows = 24
        min_sparse_columns = 10
        if last_row < min_sparse_rows:
            last_row = min_sparse_rows
        if selected_columns:
            max_selected_col = max(selected_columns)
            if len(selected_columns) < min_sparse_columns:
                padded_target_col = max(min_sparse_columns, max_selected_col)
                selected_columns = [
                    col_idx
                    for col_idx in range(1, padded_target_col + 1)
                    if not is_column_hidden(worksheet, col_idx)
                ]
        visible_rows = [row_idx for row_idx in range(1, last_row + 1) if not is_row_hidden(worksheet, row_idx)]
        visible_row_set = set(visible_rows)
        hidden_rows_skipped = max(0, last_row - len(visible_rows))

    selected_col_set = set(selected_columns)

    merge_top_left: dict[tuple[int, int], tuple[int, int]] = {}
    merge_skip: set[tuple[int, int]] = set()
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.max_row > last_row:
            continue
        fully_visible = all(
            row_idx in visible_row_set
            for row_idx in range(merged_range.min_row, merged_range.max_row + 1)
        )
        if not fully_visible:
            continue
        fully_selected = all(
            col in selected_col_set
            for col in range(merged_range.min_col, merged_range.max_col + 1)
        )
        if not fully_selected:
            continue
        top_left = (merged_range.min_row, merged_range.min_col)
        merge_top_left[top_left] = (
            merged_range.max_row - merged_range.min_row + 1,
            merged_range.max_col - merged_range.min_col + 1,
        )
        for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
            for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                if (row_idx, col_idx) != top_left:
                    merge_skip.add((row_idx, col_idx))

    style_to_class: dict[int, str] = {}
    style_rules: list[str] = []

    def style_class_for(cell) -> str:
        try:
            style_id = int(cell.style_id)
        except Exception:
            style_id = -1
        if style_id in style_to_class:
            return style_to_class[style_id]
        class_name = f"sx{len(style_to_class)}"
        style_to_class[style_id] = class_name
        css = cell_css(cell)
        style_rules.append(f".main-source-table td.{class_name}{{{css}}}")
        return class_name

    colgroup_html: list[str] = []
    for col_idx in selected_columns:
        col_dim = worksheet.column_dimensions.get(get_column_letter(col_idx))
        width = getattr(col_dim, "width", None)
        if width:
            pixel_width = max(44, int(width * 7))
            colgroup_html.append(f'<col style="width:{pixel_width}px">')
        else:
            colgroup_html.append("<col>")

    sheet_format = getattr(worksheet, "sheet_format", None)
    default_row_height_points = getattr(sheet_format, "defaultRowHeight", None)
    default_row_height_px = points_to_css_px(default_row_height_points, fallback_px=20) or 20

    rows_html: list[str] = []
    for row_idx in visible_rows:
        row_dim = worksheet.row_dimensions.get(row_idx)
        custom_height_points = getattr(row_dim, "height", None)
        custom_height_enabled = bool(getattr(row_dim, "customHeight", False))
        if (
            not custom_height_enabled
            and custom_height_points is not None
            and default_row_height_points is not None
        ):
            try:
                custom_height_enabled = abs(float(custom_height_points) - float(default_row_height_points)) > 1e-6
            except (TypeError, ValueError):
                custom_height_enabled = False

        row_height_px = (
            points_to_css_px(custom_height_points, fallback_px=default_row_height_px)
            if custom_height_enabled
            else default_row_height_px
        )
        row_height_px = row_height_px or default_row_height_px
        row_height_source = "custom" if custom_height_enabled else "default"
        row_attrs = f' data-excel-row-height="{row_height_px}" data-excel-row-height-source="{row_height_source}"'
        row_style = f' style="height:{row_height_px}px;min-height:{row_height_px}px"'
        cell_html_parts: list[str] = []
        for col_idx in selected_columns:
            if (row_idx, col_idx) in merge_skip:
                continue
            cell = worksheet.cell(row=row_idx, column=col_idx)
            class_name = style_class_for(cell)
            attrs = [f'class="{class_name}"']
            if (row_idx, col_idx) in merge_top_left:
                rowspan, colspan = merge_top_left[(row_idx, col_idx)]
                if rowspan > 1:
                    attrs.append(f'rowspan="{rowspan}"')
                if colspan > 1:
                    attrs.append(f'colspan="{colspan}"')

            text = html.escape(display_value(cell.value)).replace("\n", "<br>")
            cell_html_parts.append(f"<td {' '.join(attrs)}>{text}</td>")

        rows_html.append(f"<tr{row_attrs}{row_style}>{''.join(cell_html_parts)}</tr>")

    inline_styles = "".join(style_rules)
    base_style = (
        "<style>"
        ".main-source-table{border-collapse:collapse;width:max-content;min-width:100%;}"
        ".main-source-table td{padding:6px 8px;white-space:nowrap;border:1px solid rgba(133,184,208,0.22);}"
        f"{inline_styles}"
        "</style>"
    )
    table_html = (
        f"{base_style}<table class=\"main-source-table\">"
        f"<colgroup>{''.join(colgroup_html)}</colgroup>"
        f"<tbody>{''.join(rows_html)}</tbody>"
        "</table>"
    )

    frozen_columns = min(
        len(selected_columns),
        sum(1 for col_idx in selected_columns if col_idx <= frozen_col_boundary),
    )
    frozen_rows = min(
        len(visible_rows),
        sum(1 for row_idx in visible_rows if row_idx <= frozen_row_boundary),
    )

    return {
        "sheet_name": sheet_name,
        "row_count": len(visible_rows),
        "col_count": len(selected_columns),
        "frozen_count": frozen_columns,
        "frozen_rows": frozen_rows,
        "frozen_columns": frozen_columns,
        "selected_month_labels": selected_month_labels,
        "available_months": available_months,
        "hidden_rows_skipped": hidden_rows_skipped,
        "hidden_columns_skipped": hidden_columns_skipped,
        "html": table_html,
    }




def clear_styled_cache() -> None:
    build_main_sheet_html_cached.cache_clear()
