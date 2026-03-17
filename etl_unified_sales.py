#!/usr/bin/env python3
"""ETL both raw workbooks into unified DB-ready CSV + SQL load script.

Usage:
    python3 etl_unified_sales.py \
      --mlm-file "9-MLM Table and DailySales-2026_Feb.xlsx" \
      --mtl-file "7-MTL for Table and DailySales(2026 Jan to Mar).xlsx" \
      --output-dir outputs/unified
"""

from __future__ import annotations

import argparse
import csv
import json
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def as_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    text = str(value).strip()
    return text if text else None


def as_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text or text in {"-", "#N/A", "None", "nan"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def as_smallint(value: Any) -> int | None:
    number = as_number(value)
    if number is None:
        return None
    return int(number)


def as_timestamp(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day).strftime("%Y-%m-%d %H:%M:%S")
    text = as_text(value)
    if not text:
        return None
    # Keep raw text if it looks like datetime text in source.
    return text


def normalize_output_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float):
        # Keep integers clean for CSV readability.
        if value.is_integer():
            return int(value)
    return value


@dataclass(frozen=True)
class SourceConfig:
    label: str
    path: Path
    table_map: dict[str, int]
    daily_map: dict[str, int]


TABLE_COLUMNS = [
    "source_file",
    "source_sheet",
    "source_row",
    "stock_id",
    "stock_particular",
    "ml",
    "qty",
    "product_group",
    "sales_price",
    "customer_code",
    "customer_name",
    "customer_address",
    "township",
    "sales_channel",
    "sale_particular",
    "car_no",
    "pg_name",
    "secondary_stock_id",
    "secondary_particular",
    "secondary_ml",
    "secondary_qty",
    "secondary_product_group",
    "secondary_sales_price",
    "team_or_van",
    "extra_unmapped",
]


DAILY_COLUMNS = [
    "source_file",
    "source_sheet",
    "source_row",
    "year",
    "month",
    "sale_date",
    "snapshot_date",
    "period_text",
    "product_group",
    "voucher_no",
    "car_no",
    "customer_id",
    "customer_name",
    "township",
    "customer_address",
    "sales_channel",
    "sale_particular",
    "stock_id",
    "stock_name",
    "ml",
    "pack_size",
    "bottle",
    "parking",
    "sales_pk",
    "sales_bot",
    "sales_liter",
    "unit_price",
    "line_amount",
    "opening_balance",
    "discount_old_price",
    "commission",
    "discount_extra",
    "transport_discount",
    "transport_plus",
    "payment_date_1",
    "payment_amount_1",
    "payment_date_2",
    "payment_amount_2",
    "payment_date_3",
    "payment_amount_3",
    "payment_date_4",
    "payment_amount_4",
    "outstanding_balance",
    "report_date",
    "report_month",
    "report_year",
    "report_customer_name",
    "report_township",
    "report_team",
    "report_particular",
    "report_stock_name",
    "report_sales_ctns",
    "report_sales_bot",
    "report_sales_liter",
    "extra_unmapped",
]


def extract_unmapped(row: tuple[Any, ...], mapped_cols: set[int]) -> dict[str, Any]:
    extras: dict[str, Any] = {}
    for idx, value in enumerate(row, start=1):
        if idx in mapped_cols:
            continue
        text = as_text(value)
        if text is not None:
            extras[f"col_{idx}"] = text
    return extras


def build_table_record(source: SourceConfig, row_idx: int, row: tuple[Any, ...]) -> dict[str, Any]:
    m = source.table_map
    mapped_cols = set(m.values())
    extras = extract_unmapped(row, mapped_cols)

    def g(col_name: str) -> Any:
        col = m.get(col_name)
        if col is None or col > len(row):
            return None
        return row[col - 1]

    rec = {
        "source_file": source.path.name,
        "source_sheet": "Table",
        "source_row": row_idx,
        "stock_id": as_text(g("stock_id")),
        "stock_particular": as_text(g("stock_particular")),
        "ml": as_number(g("ml")),
        "qty": as_number(g("qty")),
        "product_group": as_text(g("product_group")),
        "sales_price": as_number(g("sales_price")),
        "customer_code": as_text(g("customer_code")),
        "customer_name": as_text(g("customer_name")),
        "customer_address": as_text(g("customer_address")),
        "township": as_text(g("township")),
        "sales_channel": as_text(g("sales_channel")),
        "sale_particular": as_text(g("sale_particular")),
        "car_no": as_text(g("car_no")),
        "pg_name": as_text(g("pg_name")),
        "secondary_stock_id": as_text(g("secondary_stock_id")),
        "secondary_particular": as_text(g("secondary_particular")),
        "secondary_ml": as_number(g("secondary_ml")),
        "secondary_qty": as_number(g("secondary_qty")),
        "secondary_product_group": as_text(g("secondary_product_group")),
        "secondary_sales_price": as_number(g("secondary_sales_price")),
        "team_or_van": as_text(g("team_or_van")),
        "extra_unmapped": json.dumps(extras, ensure_ascii=False),
    }
    return rec


def has_table_business_data(rec: dict[str, Any]) -> bool:
    keys = [
        "stock_id",
        "stock_particular",
        "customer_code",
        "customer_name",
        "sales_channel",
        "sale_particular",
        "secondary_stock_id",
    ]
    return any(rec.get(k) not in (None, "") for k in keys)


def build_daily_record(source: SourceConfig, row_idx: int, row: tuple[Any, ...]) -> dict[str, Any]:
    m = source.daily_map
    mapped_cols = set(m.values())
    extras = extract_unmapped(row, mapped_cols)

    def g(col_name: str) -> Any:
        col = m.get(col_name)
        if col is None or col > len(row):
            return None
        return row[col - 1]

    rec = {
        "source_file": source.path.name,
        "source_sheet": "DailySales",
        "source_row": row_idx,
        "year": as_smallint(g("year")),
        "month": as_smallint(g("month")),
        "sale_date": as_timestamp(g("sale_date")),
        "snapshot_date": as_timestamp(g("snapshot_date")),
        "period_text": as_text(g("period_text")),
        "product_group": as_text(g("product_group")),
        "voucher_no": as_text(g("voucher_no")),
        "car_no": as_text(g("car_no")),
        "customer_id": as_text(g("customer_id")),
        "customer_name": as_text(g("customer_name")),
        "township": as_text(g("township")),
        "customer_address": as_text(g("customer_address")),
        "sales_channel": as_text(g("sales_channel")),
        "sale_particular": as_text(g("sale_particular")),
        "stock_id": as_text(g("stock_id")),
        "stock_name": as_text(g("stock_name")),
        "ml": as_number(g("ml")),
        "pack_size": as_number(g("pack_size")),
        "bottle": as_number(g("bottle")),
        "parking": as_number(g("parking")),
        "sales_pk": as_number(g("sales_pk")),
        "sales_bot": as_number(g("sales_bot")),
        "sales_liter": as_number(g("sales_liter")),
        "unit_price": as_number(g("unit_price")),
        "line_amount": as_number(g("line_amount")),
        "opening_balance": as_number(g("opening_balance")),
        "discount_old_price": as_number(g("discount_old_price")),
        "commission": as_number(g("commission")),
        "discount_extra": as_number(g("discount_extra")),
        "transport_discount": as_number(g("transport_discount")),
        "transport_plus": as_number(g("transport_plus")),
        "payment_date_1": as_timestamp(g("payment_date_1")),
        "payment_amount_1": as_number(g("payment_amount_1")),
        "payment_date_2": as_timestamp(g("payment_date_2")),
        "payment_amount_2": as_number(g("payment_amount_2")),
        "payment_date_3": as_timestamp(g("payment_date_3")),
        "payment_amount_3": as_number(g("payment_amount_3")),
        "payment_date_4": as_timestamp(g("payment_date_4")),
        "payment_amount_4": as_number(g("payment_amount_4")),
        "outstanding_balance": as_number(g("outstanding_balance")),
        "report_date": as_timestamp(g("report_date")),
        "report_month": as_text(g("report_month")),
        "report_year": as_smallint(g("report_year")),
        "report_customer_name": as_text(g("report_customer_name")),
        "report_township": as_text(g("report_township")),
        "report_team": as_text(g("report_team")),
        "report_particular": as_text(g("report_particular")),
        "report_stock_name": as_text(g("report_stock_name")),
        "report_sales_ctns": as_number(g("report_sales_ctns")),
        "report_sales_bot": as_number(g("report_sales_bot")),
        "report_sales_liter": as_number(g("report_sales_liter")),
        "extra_unmapped": json.dumps(extras, ensure_ascii=False),
    }
    return rec


def has_daily_business_data(rec: dict[str, Any]) -> bool:
    keys = [
        "sale_date",
        "voucher_no",
        "customer_id",
        "customer_name",
        "sales_channel",
        "stock_id",
        "stock_name",
        "line_amount",
        "opening_balance",
    ]
    return any(rec.get(k) not in (None, "") for k in keys)


def write_csv(path: Path, columns: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            normalized = {k: normalize_output_value(row.get(k)) for k in columns}
            writer.writerow(normalized)


def build_sql_template(table_csv: Path, daily_csv: Path) -> str:
    table_cols = ", ".join(TABLE_COLUMNS)
    daily_cols = ", ".join(DAILY_COLUMNS)

    return f"""-- Optional: run DDL from docs/unified_sales_schema.md first.

-- Load unified Table rows
COPY stg_sales_table_unified ({table_cols})
FROM '{table_csv.as_posix()}'
WITH (FORMAT csv, HEADER true, ENCODING 'UTF8');

-- Load unified DailySales rows
COPY stg_daily_sales_unified ({daily_cols})
FROM '{daily_csv.as_posix()}'
WITH (FORMAT csv, HEADER true, ENCODING 'UTF8');
"""


def extract_source(source: SourceConfig) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_workbook(source.path, read_only=True, data_only=True)
    table_ws = wb["Table"]
    daily_ws = wb["DailySales"]

    table_rows: list[dict[str, Any]] = []
    for row_idx, row in enumerate(
        table_ws.iter_rows(min_row=2, max_row=table_ws.max_row, min_col=1, max_col=table_ws.max_column, values_only=True),
        start=2,
    ):
        rec = build_table_record(source, row_idx, row)
        if has_table_business_data(rec):
            table_rows.append(rec)

    daily_rows: list[dict[str, Any]] = []
    for row_idx, row in enumerate(
        daily_ws.iter_rows(min_row=2, max_row=daily_ws.max_row, min_col=1, max_col=daily_ws.max_column, values_only=True),
        start=2,
    ):
        rec = build_daily_record(source, row_idx, row)
        if has_daily_business_data(rec):
            daily_rows.append(rec)

    wb.close()
    return table_rows, daily_rows


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export both raw workbooks into unified DB-ready CSVs.")
    parser.add_argument(
        "--mlm-file",
        type=Path,
        default=Path("9-MLM Table and DailySales-2026_Feb.xlsx"),
        help="Path to 9-MLM workbook",
    )
    parser.add_argument(
        "--mtl-file",
        type=Path,
        default=Path("7-MTL for Table and DailySales(2026 Jan to Mar).xlsx"),
        help="Path to 7-MTL workbook",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("outputs/unified"),
        help="Directory where CSV/SQL output files are written",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    mlm_file = args.mlm_file.resolve()
    mtl_file = args.mtl_file.resolve()
    output_dir = args.output_dir.resolve()

    for required in (mlm_file, mtl_file):
        if not required.exists():
            raise FileNotFoundError(f"Missing source workbook: {required}")

    mlm = SourceConfig(
        label="9-MLM",
        path=mlm_file,
        table_map={
            "stock_id": 1,
            "stock_particular": 2,
            "ml": 3,
            "qty": 4,
            "product_group": 5,
            "sales_price": 6,
            "customer_code": 9,
            "customer_name": 10,
            "customer_address": 11,
            "township": 12,
            "sales_channel": 13,
            "sale_particular": 15,
            "car_no": 17,
            "team_or_van": 19,
        },
        daily_map={
            "year": 1,
            "month": 2,
            "sale_date": 3,
            "snapshot_date": 4,
            "period_text": 5,
            "product_group": 6,
            "voucher_no": 7,
            "car_no": 8,
            "customer_id": 9,
            "customer_name": 10,
            "township": 11,
            "sales_channel": 12,
            "sale_particular": 13,
            "stock_id": 14,
            "stock_name": 15,
            "ml": 16,
            "pack_size": 17,
            "bottle": 18,
            "parking": 19,
            "sales_pk": 20,
            "sales_bot": 21,
            "sales_liter": 22,
            "unit_price": 23,
            "line_amount": 24,
            "opening_balance": 25,
            "discount_old_price": 26,
            "commission": 27,
            "payment_date_1": 28,
            "payment_amount_1": 29,
            "payment_date_2": 30,
            "payment_amount_2": 31,
            "outstanding_balance": 32,
            "report_date": 34,
            "report_month": 35,
            "report_year": 36,
            "report_customer_name": 37,
            "report_township": 38,
            "report_team": 39,
            "report_particular": 40,
            "report_stock_name": 41,
            "report_sales_ctns": 42,
            "report_sales_bot": 43,
            "report_sales_liter": 44,
        },
    )

    mtl = SourceConfig(
        label="7-MTL",
        path=mtl_file,
        table_map={
            "stock_id": 1,
            "stock_particular": 2,
            "ml": 3,
            "qty": 4,
            "product_group": 5,
            "sales_price": 6,
            "customer_code": 8,
            "customer_name": 9,
            "customer_address": 10,
            "township": 11,
            "sales_channel": 12,
            "sale_particular": 14,
            "car_no": 16,
            "pg_name": 18,
            "secondary_stock_id": 20,
            "secondary_particular": 21,
            "secondary_ml": 22,
            "secondary_qty": 23,
            "secondary_product_group": 24,
            "secondary_sales_price": 25,
        },
        daily_map={
            "year": 1,
            "month": 2,
            "sale_date": 3,
            "snapshot_date": 4,
            "period_text": 5,
            "product_group": 6,
            "voucher_no": 7,
            "car_no": 8,
            "customer_id": 9,
            "customer_name": 10,
            "township": 11,
            "customer_address": 12,
            "sales_channel": 13,
            "sale_particular": 14,
            "stock_id": 15,
            "stock_name": 16,
            "ml": 17,
            "pack_size": 18,
            "bottle": 19,
            "parking": 20,
            "sales_pk": 21,
            "sales_bot": 22,
            "sales_liter": 23,
            "unit_price": 24,
            "line_amount": 25,
            "opening_balance": 26,
            "discount_old_price": 27,
            "commission": 28,
            "discount_extra": 29,
            "transport_discount": 30,
            "transport_plus": 31,
            "payment_date_1": 32,
            "payment_amount_1": 33,
            "payment_date_2": 34,
            "payment_amount_2": 35,
            "payment_date_3": 36,
            "payment_amount_3": 37,
            "payment_date_4": 38,
            "payment_amount_4": 39,
            "outstanding_balance": 40,
        },
    )

    table_rows: list[dict[str, Any]] = []
    daily_rows: list[dict[str, Any]] = []

    for source in (mlm, mtl):
        source_table, source_daily = extract_source(source)
        table_rows.extend(source_table)
        daily_rows.extend(source_daily)

    table_csv = output_dir / "stg_sales_table_unified.csv"
    daily_csv = output_dir / "stg_daily_sales_unified.csv"
    sql_path = output_dir / "load_unified_sales.sql"
    summary_path = output_dir / "unified_etl_summary.json"

    write_csv(table_csv, TABLE_COLUMNS, table_rows)
    write_csv(daily_csv, DAILY_COLUMNS, daily_rows)
    sql_path.write_text(build_sql_template(table_csv, daily_csv), encoding="utf-8")

    summary = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "sources": [mlm.path.as_posix(), mtl.path.as_posix()],
        "outputs": {
            "table_csv": table_csv.as_posix(),
            "daily_csv": daily_csv.as_posix(),
            "sql_load_file": sql_path.as_posix(),
        },
        "row_counts": {
            "stg_sales_table_unified": len(table_rows),
            "stg_daily_sales_unified": len(daily_rows),
        },
    }
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Wrote: {table_csv}")
    print(f"Wrote: {daily_csv}")
    print(f"Wrote: {sql_path}")
    print(f"Wrote: {summary_path}")
    print(f"Rows: table={len(table_rows)} daily={len(daily_rows)}")


if __name__ == "__main__":
    main()
