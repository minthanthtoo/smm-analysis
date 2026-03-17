#!/usr/bin/env python3
"""Generate SKU and township analysis workbooks from a single DailySales source.

Source workbook expected structure:
- Sheet `Table`: stock and customer masters.
- Sheet `DailySales`: transactional rows with formula-resolved values.

Outputs:
- <prefix>_SKU_Analysis.xlsx
- <prefix>_Township_Analysis.xlsx
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


@dataclass
class SalesRow:
    tx_date: date
    year: int
    month: int
    month_name: str
    voucher_no: str
    car_no: str
    customer_id: str
    customer_name: str
    township: str
    team: str
    particular: str
    stock_id: str
    stock_name: str
    ml: float
    packing: int
    bottle: float
    sales_pk: float
    sales_bot: float
    liter: float
    price: float
    amount: float


class Col:
    DATE = 3
    VOUCHER = 7
    CAR_NO = 8
    CUSTOMER_ID = 9
    CUSTOMER_NAME = 10
    TOWNSHIP = 11
    TEAM = 13
    PARTICULAR = 14
    STOCK_ID = 15
    STOCK_NAME = 16
    ML = 17
    PACKING = 18
    BOTTLE = 19
    SALES_PK = 21
    SALES_BOT = 22
    LITER = 23
    PRICE = 24
    AMOUNT = 25


def as_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text in {"None", "nan", "#N/A", "00:00:00"}:
        return ""
    return text


def as_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text or text in {"#N/A", "-"}:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def as_int(value: Any) -> int:
    number = as_float(value)
    return int(number) if number else 0


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def month_name(month: int) -> str:
    return date(2000, month, 1).strftime("%b")


def load_masters(source_path: Path) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    wb = load_workbook(source_path, data_only=True, read_only=False)
    ws = wb["Table"]

    stock_map: dict[str, dict[str, Any]] = {}
    for row in range(2, ws.max_row + 1):
        stock_id = as_text(ws.cell(row, 1).value)
        if not stock_id:
            continue
        stock_map[stock_id] = {
            "stock_name": as_text(ws.cell(row, 2).value),
            "ml": as_float(ws.cell(row, 3).value),
            "packing": as_int(ws.cell(row, 4).value),
            "team": as_text(ws.cell(row, 5).value),
            "price": as_float(ws.cell(row, 6).value),
        }

    customer_map: dict[str, dict[str, Any]] = {}
    for row in range(2, ws.max_row + 1):
        customer_id = as_text(ws.cell(row, 8).value)
        if not customer_id:
            continue
        customer_map[customer_id] = {
            "customer_name": as_text(ws.cell(row, 9).value),
            "address": as_text(ws.cell(row, 10).value),
            "township": as_text(ws.cell(row, 11).value),
            "team": as_text(ws.cell(row, 12).value),
        }

    wb.close()
    return stock_map, customer_map


def extract_sales_rows(
    source_path: Path,
    stock_map: dict[str, dict[str, Any]],
    customer_map: dict[str, dict[str, Any]],
) -> tuple[list[SalesRow], dict[str, int]]:
    wb = load_workbook(source_path, data_only=True, read_only=False)
    ws = wb["DailySales"]

    rows: list[SalesRow] = []
    skipped = defaultdict(int)

    for row in range(2, ws.max_row + 1):
        tx_date = parse_date(ws.cell(row, Col.DATE).value)
        if not tx_date:
            skipped["missing_date"] += 1
            continue

        customer_id = as_text(ws.cell(row, Col.CUSTOMER_ID).value)
        stock_id = as_text(ws.cell(row, Col.STOCK_ID).value)
        particular = as_text(ws.cell(row, Col.PARTICULAR).value)

        customer_ref = customer_map.get(customer_id, {})
        stock_ref = stock_map.get(stock_id, {})

        customer_name = as_text(ws.cell(row, Col.CUSTOMER_NAME).value) or customer_ref.get("customer_name", "")
        township = as_text(ws.cell(row, Col.TOWNSHIP).value) or customer_ref.get("township", "")
        team = as_text(ws.cell(row, Col.TEAM).value) or customer_ref.get("team", "")

        stock_name = as_text(ws.cell(row, Col.STOCK_NAME).value) or stock_ref.get("stock_name", "")
        ml = as_float(ws.cell(row, Col.ML).value) or float(stock_ref.get("ml", 0.0))
        packing = as_int(ws.cell(row, Col.PACKING).value) or int(stock_ref.get("packing", 0))

        bottle = as_float(ws.cell(row, Col.BOTTLE).value)
        sales_pk = as_float(ws.cell(row, Col.SALES_PK).value)
        sales_bot = as_float(ws.cell(row, Col.SALES_BOT).value)
        liter = as_float(ws.cell(row, Col.LITER).value)
        price = as_float(ws.cell(row, Col.PRICE).value) or float(stock_ref.get("price", 0.0))
        amount = as_float(ws.cell(row, Col.AMOUNT).value)

        # Filter to true sales lines for SKU/township analysis.
        if sales_bot <= 0 and liter <= 0:
            skipped["no_sales_qty"] += 1
            continue
        if ml <= 0 or packing <= 0:
            skipped["invalid_sku_dimension"] += 1
            continue
        if not stock_id:
            skipped["missing_stock_id"] += 1
            continue
        if not township:
            skipped["missing_township"] += 1
            continue

        if amount <= 0 and price > 0:
            amount = sales_bot * price

        rows.append(
            SalesRow(
                tx_date=tx_date,
                year=tx_date.year,
                month=tx_date.month,
                month_name=month_name(tx_date.month),
                voucher_no=as_text(ws.cell(row, Col.VOUCHER).value),
                car_no=as_text(ws.cell(row, Col.CAR_NO).value),
                customer_id=customer_id,
                customer_name=customer_name,
                township=township,
                team=team,
                particular=particular,
                stock_id=stock_id,
                stock_name=stock_name,
                ml=ml,
                packing=packing,
                bottle=bottle,
                sales_pk=sales_pk,
                sales_bot=sales_bot,
                liter=liter,
                price=price,
                amount=amount,
            )
        )

    wb.close()
    skipped["kept_rows"] = len(rows)
    return rows, dict(skipped)


def sorted_month_keys(rows: Iterable[SalesRow]) -> list[tuple[int, int]]:
    keys = {(r.year, r.month) for r in rows}
    return sorted(keys)


def autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 38)


def write_table_sheet(
    wb: Workbook,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    number_formats: dict[int, str] | None = None,
) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    for row in rows:
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    if number_formats:
        for col_idx, fmt in number_formats.items():
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col_idx).number_format = fmt

    for r in range(2, ws.max_row + 1):
        ws.cell(r, 1).alignment = LEFT

    autosize(ws)


def write_header_row(ws, row_num: int, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[row_num]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def safe_sheet_name(name: str, fallback: str) -> str:
    invalid = set("[]:*?/\\")
    cleaned = "".join(ch for ch in name if ch not in invalid).strip()
    if not cleaned:
        cleaned = fallback
    return cleaned[:31]


def build_sku_workbook(
    output_path: Path,
    source_path: Path,
    rows: list[SalesRow],
    profile: dict[str, int],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # README
    readme = wb.create_sheet("README")
    readme.append(["Source File", str(source_path.name)])
    readme.append(["Generated At", datetime.now().isoformat(timespec="seconds")])
    readme.append(["Rows Used", profile.get("kept_rows", 0)])
    for key in sorted(k for k in profile.keys() if k != "kept_rows"):
        readme.append([f"Skipped: {key}", profile[key]])
    autosize(readme)

    raw_headers = [
        "Date",
        "Year",
        "Month",
        "MonthName",
        "VoucherNo",
        "CarNo",
        "CustomerID",
        "CustomerName",
        "Township",
        "Team",
        "Particular",
        "StockID",
        "StockName",
        "ML",
        "Packing",
        "Bottle",
        "SalesPK",
        "SalesBot",
        "Liter",
        "Price",
        "Amount",
    ]
    raw_rows = [
        [
            r.tx_date,
            r.year,
            r.month,
            r.month_name,
            r.voucher_no,
            r.car_no,
            r.customer_id,
            r.customer_name,
            r.township,
            r.team,
            r.particular,
            r.stock_id,
            r.stock_name,
            r.ml,
            r.packing,
            r.bottle,
            r.sales_pk,
            r.sales_bot,
            r.liter,
            r.price,
            r.amount,
        ]
        for r in sorted(rows, key=lambda x: (x.tx_date, x.stock_name, x.customer_id))
    ]
    write_table_sheet(
        wb,
        "Raw_Normalized",
        raw_headers,
        raw_rows,
        number_formats={1: "yyyy-mm-dd", 14: "0.00", 16: "0.00", 17: "0.00", 18: "0.00", 19: "0.00", 20: "#,##0.00", 21: "#,##0.00"},
    )

    month_totals_liter: dict[tuple[int, int], float] = defaultdict(float)
    month_totals_amt: dict[tuple[int, int], float] = defaultdict(float)
    sku_month = defaultdict(lambda: {"rows": 0, "sales_pk": 0.0, "bottle": 0.0, "sales_bot": 0.0, "liter": 0.0, "amount": 0.0, "customers": set(), "townships": set()})

    for r in rows:
        mkey = (r.year, r.month)
        month_totals_liter[mkey] += r.liter
        month_totals_amt[mkey] += r.amount
        key = (r.year, r.month, r.stock_id, r.stock_name, r.ml, r.packing, r.team)
        bucket = sku_month[key]
        bucket["rows"] += 1
        bucket["sales_pk"] += r.sales_pk
        bucket["bottle"] += r.bottle
        bucket["sales_bot"] += r.sales_bot
        bucket["liter"] += r.liter
        bucket["amount"] += r.amount
        bucket["customers"].add(r.customer_id)
        bucket["townships"].add(r.township)

    sku_month_rows: list[list[Any]] = []
    for key in sorted(sku_month.keys(), key=lambda k: (k[0], k[1], -sku_month[k]["liter"], k[3])):
        year, month, stock_id, stock_name, ml, packing, team = key
        b = sku_month[key]
        denom_liter = month_totals_liter[(year, month)]
        denom_amt = month_totals_amt[(year, month)]
        sku_month_rows.append(
            [
                year,
                month,
                month_name(month),
                stock_id,
                stock_name,
                ml,
                packing,
                team,
                b["rows"],
                len(b["customers"]),
                len(b["townships"]),
                b["sales_pk"],
                b["bottle"],
                b["sales_bot"],
                b["liter"],
                b["amount"],
                (b["liter"] / denom_liter) if denom_liter else 0.0,
                (b["amount"] / denom_amt) if denom_amt else 0.0,
            ]
        )

    write_table_sheet(
        wb,
        "SKU_Monthly",
        [
            "Year",
            "Month",
            "MonthName",
            "StockID",
            "StockName",
            "ML",
            "Packing",
            "Team",
            "TxnRows",
            "Outlets",
            "Townships",
            "SalesPK",
            "Bottle",
            "SalesBot",
            "Liter",
            "Amount",
            "LiterContribution",
            "AmountContribution",
        ],
        sku_month_rows,
        number_formats={6: "0.00", 12: "0.00", 13: "0.00", 14: "0.00", 15: "0.00", 16: "#,##0.00", 17: "0.00%", 18: "0.00%"},
    )

    sku_total = defaultdict(lambda: {"rows": 0, "sales_bot": 0.0, "liter": 0.0, "amount": 0.0, "customers": set(), "townships": set(), "months": set()})
    grand_liter = sum(r.liter for r in rows)
    grand_amt = sum(r.amount for r in rows)

    for r in rows:
        key = (r.stock_id, r.stock_name, r.ml, r.packing, r.team)
        b = sku_total[key]
        b["rows"] += 1
        b["sales_bot"] += r.sales_bot
        b["liter"] += r.liter
        b["amount"] += r.amount
        b["customers"].add(r.customer_id)
        b["townships"].add(r.township)
        b["months"].add((r.year, r.month))

    sku_total_rows: list[list[Any]] = []
    for key in sorted(sku_total.keys(), key=lambda k: (-sku_total[k]["liter"], k[1])):
        stock_id, stock_name, ml, packing, team = key
        b = sku_total[key]
        sku_total_rows.append(
            [
                stock_id,
                stock_name,
                ml,
                packing,
                team,
                b["rows"],
                len(b["months"]),
                len(b["customers"]),
                len(b["townships"]),
                b["sales_bot"],
                b["liter"],
                b["amount"],
                (b["liter"] / grand_liter) if grand_liter else 0.0,
                (b["amount"] / grand_amt) if grand_amt else 0.0,
            ]
        )

    write_table_sheet(
        wb,
        "SKU_Total",
        [
            "StockID",
            "StockName",
            "ML",
            "Packing",
            "Team",
            "TxnRows",
            "ActiveMonths",
            "Outlets",
            "Townships",
            "SalesBot",
            "Liter",
            "Amount",
            "LiterContribution",
            "AmountContribution",
        ],
        sku_total_rows,
        number_formats={3: "0.00", 10: "0.00", 11: "0.00", 12: "#,##0.00", 13: "0.00%", 14: "0.00%"},
    )

    sku_town_rows = defaultdict(lambda: {"sales_bot": 0.0, "liter": 0.0, "amount": 0.0, "rows": 0})
    for r in rows:
        key = (r.year, r.month, r.township, r.stock_id, r.stock_name, r.ml, r.packing)
        b = sku_town_rows[key]
        b["sales_bot"] += r.sales_bot
        b["liter"] += r.liter
        b["amount"] += r.amount
        b["rows"] += 1

    sku_town_out = []
    for key in sorted(sku_town_rows.keys(), key=lambda k: (k[0], k[1], k[2], -sku_town_rows[k]["liter"], k[4])):
        year, month, township, stock_id, stock_name, ml, packing = key
        b = sku_town_rows[key]
        sku_town_out.append([
            year,
            month,
            month_name(month),
            township,
            stock_id,
            stock_name,
            ml,
            packing,
            b["rows"],
            b["sales_bot"],
            b["liter"],
            b["amount"],
        ])

    write_table_sheet(
        wb,
        "SKU_by_Township",
        [
            "Year",
            "Month",
            "MonthName",
            "Township",
            "StockID",
            "StockName",
            "ML",
            "Packing",
            "TxnRows",
            "SalesBot",
            "Liter",
            "Amount",
        ],
        sku_town_out,
        number_formats={7: "0.00", 10: "0.00", 11: "0.00", 12: "#,##0.00"},
    )

    wb.save(output_path)


def build_township_workbook(
    output_path: Path,
    source_path: Path,
    rows: list[SalesRow],
    profile: dict[str, int],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    readme = wb.create_sheet("README")
    readme.append(["Source File", str(source_path.name)])
    readme.append(["Generated At", datetime.now().isoformat(timespec="seconds")])
    readme.append(["Rows Used", profile.get("kept_rows", 0)])
    for key in sorted(k for k in profile.keys() if k != "kept_rows"):
        readme.append([f"Skipped: {key}", profile[key]])
    autosize(readme)

    month_totals_liter: dict[tuple[int, int], float] = defaultdict(float)
    month_totals_amt: dict[tuple[int, int], float] = defaultdict(float)
    town_month = defaultdict(lambda: {"rows": 0, "sales_bot": 0.0, "liter": 0.0, "amount": 0.0, "customers": set(), "skus": set()})

    for r in rows:
        mkey = (r.year, r.month)
        month_totals_liter[mkey] += r.liter
        month_totals_amt[mkey] += r.amount

        key = (r.year, r.month, r.township)
        b = town_month[key]
        b["rows"] += 1
        b["sales_bot"] += r.sales_bot
        b["liter"] += r.liter
        b["amount"] += r.amount
        b["customers"].add(r.customer_id)
        b["skus"].add((r.stock_id, r.stock_name, r.ml, r.packing))

    town_month_rows = []
    grouped_by_month: dict[tuple[int, int], list[tuple[tuple[int, int, str], dict[str, Any]]]] = defaultdict(list)
    for key, val in town_month.items():
        grouped_by_month[(key[0], key[1])].append((key, val))

    for mkey in sorted(grouped_by_month.keys()):
        chunk = sorted(grouped_by_month[mkey], key=lambda item: (-item[1]["liter"], item[0][2]))
        rank = 1
        for (year, month, township), b in chunk:
            denom_liter = month_totals_liter[(year, month)]
            denom_amt = month_totals_amt[(year, month)]
            town_month_rows.append(
                [
                    year,
                    month,
                    month_name(month),
                    township,
                    rank,
                    b["rows"],
                    len(b["customers"]),
                    len(b["skus"]),
                    b["sales_bot"],
                    b["liter"],
                    b["amount"],
                    (b["liter"] / denom_liter) if denom_liter else 0.0,
                    (b["amount"] / denom_amt) if denom_amt else 0.0,
                ]
            )
            rank += 1

    write_table_sheet(
        wb,
        "Township_Monthly",
        [
            "Year",
            "Month",
            "MonthName",
            "Township",
            "RankInMonth",
            "TxnRows",
            "Outlets",
            "ActiveSKUs",
            "SalesBot",
            "Liter",
            "Amount",
            "LiterContribution",
            "AmountContribution",
        ],
        town_month_rows,
        number_formats={9: "0.00", 10: "0.00", 11: "#,##0.00", 12: "0.00%", 13: "0.00%"},
    )

    town_total = defaultdict(lambda: {"rows": 0, "sales_bot": 0.0, "liter": 0.0, "amount": 0.0, "customers": set(), "months": set(), "skus": set()})
    grand_liter = sum(r.liter for r in rows)
    grand_amt = sum(r.amount for r in rows)

    for r in rows:
        b = town_total[r.township]
        b["rows"] += 1
        b["sales_bot"] += r.sales_bot
        b["liter"] += r.liter
        b["amount"] += r.amount
        b["customers"].add(r.customer_id)
        b["months"].add((r.year, r.month))
        b["skus"].add((r.stock_id, r.stock_name, r.ml, r.packing))

    town_total_rows = []
    rank = 1
    for township, b in sorted(town_total.items(), key=lambda item: (-item[1]["liter"], item[0])):
        town_total_rows.append(
            [
                rank,
                township,
                b["rows"],
                len(b["months"]),
                len(b["customers"]),
                len(b["skus"]),
                b["sales_bot"],
                b["liter"],
                b["amount"],
                (b["liter"] / grand_liter) if grand_liter else 0.0,
                (b["amount"] / grand_amt) if grand_amt else 0.0,
            ]
        )
        rank += 1

    write_table_sheet(
        wb,
        "Township_Total",
        [
            "Rank",
            "Township",
            "TxnRows",
            "ActiveMonths",
            "Outlets",
            "ActiveSKUs",
            "SalesBot",
            "Liter",
            "Amount",
            "LiterContribution",
            "AmountContribution",
        ],
        town_total_rows,
        number_formats={7: "0.00", 8: "0.00", 9: "#,##0.00", 10: "0.00%", 11: "0.00%"},
    )

    town_sku_month = defaultdict(lambda: {"rows": 0, "sales_bot": 0.0, "liter": 0.0, "amount": 0.0, "customers": set()})
    for r in rows:
        key = (r.year, r.month, r.township, r.stock_id, r.stock_name, r.ml, r.packing)
        b = town_sku_month[key]
        b["rows"] += 1
        b["sales_bot"] += r.sales_bot
        b["liter"] += r.liter
        b["amount"] += r.amount
        b["customers"].add(r.customer_id)

    town_sku_rows = []
    for key in sorted(town_sku_month.keys(), key=lambda k: (k[0], k[1], k[2], -town_sku_month[k]["liter"], k[4])):
        year, month, township, stock_id, stock_name, ml, packing = key
        b = town_sku_month[key]
        town_sku_rows.append(
            [
                year,
                month,
                month_name(month),
                township,
                stock_id,
                stock_name,
                ml,
                packing,
                b["rows"],
                len(b["customers"]),
                b["sales_bot"],
                b["liter"],
                b["amount"],
            ]
        )

    write_table_sheet(
        wb,
        "Township_SKU_Monthly",
        [
            "Year",
            "Month",
            "MonthName",
            "Township",
            "StockID",
            "StockName",
            "ML",
            "Packing",
            "TxnRows",
            "Outlets",
            "SalesBot",
            "Liter",
            "Amount",
        ],
        town_sku_rows,
        number_formats={7: "0.00", 11: "0.00", 12: "0.00", 13: "#,##0.00"},
    )

    town_channel_month = defaultdict(lambda: {"rows": 0, "sales_bot": 0.0, "liter": 0.0, "amount": 0.0, "customers": set()})
    for r in rows:
        key = (r.year, r.month, r.township, r.team or "(blank)")
        b = town_channel_month[key]
        b["rows"] += 1
        b["sales_bot"] += r.sales_bot
        b["liter"] += r.liter
        b["amount"] += r.amount
        b["customers"].add(r.customer_id)

    town_channel_rows = []
    for key in sorted(town_channel_month.keys(), key=lambda k: (k[0], k[1], k[2], -town_channel_month[k]["liter"], k[3])):
        year, month, township, team = key
        b = town_channel_month[key]
        town_channel_rows.append(
            [
                year,
                month,
                month_name(month),
                township,
                team,
                b["rows"],
                len(b["customers"]),
                b["sales_bot"],
                b["liter"],
                b["amount"],
            ]
        )

    write_table_sheet(
        wb,
        "Township_Channel_Monthly",
        [
            "Year",
            "Month",
            "MonthName",
            "Township",
            "Team",
            "TxnRows",
            "Outlets",
            "SalesBot",
            "Liter",
            "Amount",
        ],
        town_channel_rows,
        number_formats={8: "0.00", 9: "0.00", 10: "#,##0.00"},
    )

    daily_town = defaultdict(lambda: {"rows": 0, "sales_bot": 0.0, "liter": 0.0, "amount": 0.0, "customers": set()})
    for r in rows:
        key = (r.tx_date, r.township)
        b = daily_town[key]
        b["rows"] += 1
        b["sales_bot"] += r.sales_bot
        b["liter"] += r.liter
        b["amount"] += r.amount
        b["customers"].add(r.customer_id)

    daily_town_rows = []
    for key in sorted(daily_town.keys()):
        tx_date, township = key
        b = daily_town[key]
        daily_town_rows.append(
            [
                tx_date,
                tx_date.year,
                tx_date.month,
                month_name(tx_date.month),
                township,
                b["rows"],
                len(b["customers"]),
                b["sales_bot"],
                b["liter"],
                b["amount"],
            ]
        )

    write_table_sheet(
        wb,
        "Daily_Township",
        [
            "Date",
            "Year",
            "Month",
            "MonthName",
            "Township",
            "TxnRows",
            "Outlets",
            "SalesBot",
            "Liter",
            "Amount",
        ],
        daily_town_rows,
        number_formats={1: "yyyy-mm-dd", 8: "0.00", 9: "0.00", 10: "#,##0.00"},
    )

    wb.save(output_path)


def build_region_sku_template_workbook(
    output_path: Path,
    source_path: Path,
    rows: list[SalesRow],
    profile: dict[str, int],
    region_code: str,
    customer_map: dict[str, dict[str, Any]],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    months = sorted_month_keys(rows)
    month_labels = [f"{y}-{m:02d}" for y, m in months]

    readme = wb.create_sheet("README")
    readme.append(["Source File", str(source_path.name)])
    readme.append(["Generated At", datetime.now().isoformat(timespec="seconds")])
    readme.append(["Region", region_code])
    readme.append(["Rows Used", profile.get("kept_rows", 0)])
    for key in sorted(k for k in profile.keys() if k != "kept_rows"):
        readme.append([f"Skipped: {key}", profile[key]])
    autosize(readme)

    sku_sheet_name = safe_sheet_name(f"7-{region_code}", "7-Region")
    ws = wb.create_sheet(sku_sheet_name)
    headers = ["Sr", "StockID", "Product Name", "ML", "Packing"]
    for label in month_labels:
        headers.extend([f"{label} Bot", f"{label} Lit"])
    headers.extend(["Total Bot", "Total Lit", "Total Amount"])
    write_header_row(ws, 1, headers)

    sku_totals = defaultdict(lambda: {"bot": 0.0, "lit": 0.0, "amt": 0.0})
    sku_month = defaultdict(lambda: {"bot": 0.0, "lit": 0.0})
    for r in rows:
        sku_key = (r.stock_id, r.stock_name, r.ml, r.packing)
        month_key = (r.year, r.month)
        sku_totals[sku_key]["bot"] += r.sales_bot
        sku_totals[sku_key]["lit"] += r.liter
        sku_totals[sku_key]["amt"] += r.amount
        sku_month[(sku_key, month_key)]["bot"] += r.sales_bot
        sku_month[(sku_key, month_key)]["lit"] += r.liter

    sorted_skus = sorted(sku_totals.keys(), key=lambda k: (-sku_totals[k]["lit"], k[1], k[0]))
    grand_bot = grand_lit = grand_amt = 0.0
    for idx, sku_key in enumerate(sorted_skus, start=1):
        stock_id, stock_name, ml, packing = sku_key
        row_data: list[Any] = [idx, stock_id, stock_name, ml, packing]
        for mkey in months:
            cell = sku_month.get((sku_key, mkey), {"bot": 0.0, "lit": 0.0})
            row_data.extend([cell["bot"], cell["lit"]])
        row_data.extend([sku_totals[sku_key]["bot"], sku_totals[sku_key]["lit"], sku_totals[sku_key]["amt"]])
        grand_bot += sku_totals[sku_key]["bot"]
        grand_lit += sku_totals[sku_key]["lit"]
        grand_amt += sku_totals[sku_key]["amt"]
        ws.append(row_data)

    total_row = ["", "", "All Total", "", ""]
    for mkey in months:
        month_bot = sum(sku_month[(sku_key, mkey)]["bot"] for sku_key in sorted_skus if (sku_key, mkey) in sku_month)
        month_lit = sum(sku_month[(sku_key, mkey)]["lit"] for sku_key in sorted_skus if (sku_key, mkey) in sku_month)
        total_row.extend([month_bot, month_lit])
    total_row.extend([grand_bot, grand_lit, grand_amt])
    ws.append(total_row)

    for cell in ws[ws.max_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for r in range(2, ws.max_row + 1):
        ws.cell(r, 1).alignment = CENTER
    autosize(ws)

    business_rows = []
    by_month = defaultdict(lambda: {"rows": 0, "customers": set(), "skus": set(), "townships": set(), "bot": 0.0, "lit": 0.0, "amt": 0.0})
    for r in rows:
        key = (r.year, r.month)
        b = by_month[key]
        b["rows"] += 1
        b["customers"].add(r.customer_id)
        b["skus"].add((r.stock_id, r.stock_name, r.ml, r.packing))
        b["townships"].add(r.township)
        b["bot"] += r.sales_bot
        b["lit"] += r.liter
        b["amt"] += r.amount

    for key in sorted(by_month.keys()):
        y, m = key
        b = by_month[key]
        business_rows.append(
            [
                y,
                m,
                month_name(m),
                b["rows"],
                len(b["customers"]),
                len(b["skus"]),
                len(b["townships"]),
                b["bot"],
                b["lit"],
                b["amt"],
            ]
        )

    write_table_sheet(
        wb,
        "Business Summary",
        ["Year", "Month", "MonthName", "TxnRows", "Outlets", "ActiveSKUs", "Townships", "SalesBot", "Liter", "Amount"],
        business_rows,
        number_formats={8: "0.00", 9: "0.00", 10: "#,##0.00"},
    )

    town_sheet = wb.create_sheet("Township wise Analysis")
    town_headers = ["No", "Township"]
    for label in month_labels:
        town_headers.extend([f"{label} Bot", f"{label} Lit"])
    town_headers.extend(["Total Bot", "Total Lit", "Total Amount"])
    write_header_row(town_sheet, 1, town_headers)

    town_totals = defaultdict(lambda: {"bot": 0.0, "lit": 0.0, "amt": 0.0})
    town_month = defaultdict(lambda: {"bot": 0.0, "lit": 0.0})
    for r in rows:
        tkey = r.township
        mkey = (r.year, r.month)
        town_totals[tkey]["bot"] += r.sales_bot
        town_totals[tkey]["lit"] += r.liter
        town_totals[tkey]["amt"] += r.amount
        town_month[(tkey, mkey)]["bot"] += r.sales_bot
        town_month[(tkey, mkey)]["lit"] += r.liter

    sorted_towns = sorted(town_totals.keys(), key=lambda t: (-town_totals[t]["lit"], t))
    for idx, tkey in enumerate(sorted_towns, start=1):
        row_data = [idx, tkey]
        for mkey in months:
            cell = town_month.get((tkey, mkey), {"bot": 0.0, "lit": 0.0})
            row_data.extend([cell["bot"], cell["lit"]])
        row_data.extend([town_totals[tkey]["bot"], town_totals[tkey]["lit"], town_totals[tkey]["amt"]])
        town_sheet.append(row_data)
    autosize(town_sheet)

    town_sku = defaultdict(lambda: {"rows": 0, "bot": 0.0, "lit": 0.0, "amt": 0.0, "customers": set(), "months": set()})
    for r in rows:
        key = (r.township, r.stock_id, r.stock_name, r.ml, r.packing)
        b = town_sku[key]
        b["rows"] += 1
        b["bot"] += r.sales_bot
        b["lit"] += r.liter
        b["amt"] += r.amount
        b["customers"].add(r.customer_id)
        b["months"].add((r.year, r.month))

    final_rows = []
    for key in sorted(town_sku.keys(), key=lambda k: (k[0], -town_sku[k]["lit"], k[2])):
        township, stock_id, stock_name, ml, packing = key
        b = town_sku[key]
        denom = town_totals[township]["lit"] if township in town_totals else 0.0
        final_rows.append(
            [
                township,
                stock_id,
                stock_name,
                ml,
                packing,
                b["rows"],
                len(b["months"]),
                len(b["customers"]),
                b["bot"],
                b["lit"],
                b["amt"],
                (b["lit"] / denom) if denom else 0.0,
            ]
        )

    write_table_sheet(
        wb,
        "Final Town SKU wise Analysis",
        [
            "Township",
            "StockID",
            "StockName",
            "ML",
            "Packing",
            "TxnRows",
            "ActiveMonths",
            "Outlets",
            "SalesBot",
            "Liter",
            "Amount",
            "TownshipLiterContribution",
        ],
        final_rows,
        number_formats={4: "0.00", 9: "0.00", 10: "0.00", 11: "#,##0.00", 12: "0.00%"},
    )

    outlet_totals = defaultdict(lambda: {"rows": 0, "lit": 0.0, "amt": 0.0})
    for r in rows:
        outlet_totals[r.customer_id]["rows"] += 1
        outlet_totals[r.customer_id]["lit"] += r.liter
        outlet_totals[r.customer_id]["amt"] += r.amount

    outlet_rows = []
    for cid, values in sorted(outlet_totals.items(), key=lambda item: (-item[1]["lit"], item[0])):
        ref = customer_map.get(cid, {})
        outlet_rows.append(
            [
                cid,
                ref.get("customer_name", ""),
                ref.get("township", ""),
                ref.get("team", ""),
                values["rows"],
                values["lit"],
                values["amt"],
            ]
        )

    write_table_sheet(
        wb,
        "Outlet Summary",
        ["CustomerID", "CustomerName", "Township", "Team", "TxnRows", "Liter", "Amount"],
        outlet_rows,
        number_formats={6: "0.00", 7: "#,##0.00"},
    )

    outlet_list_rows = []
    for cid, ref in sorted(customer_map.items(), key=lambda item: (item[1].get("township", ""), item[1].get("customer_name", ""), item[0])):
        is_active = "Y" if cid in outlet_totals else "N"
        outlet_list_rows.append([cid, ref.get("customer_name", ""), ref.get("address", ""), ref.get("township", ""), ref.get("team", ""), is_active])

    write_table_sheet(
        wb,
        "Outlet List",
        ["CustomerID", "CustomerName", "Address", "Township", "Team", "ActiveInPeriod"],
        outlet_list_rows,
    )

    wb.save(output_path)


def build_region_township_template_workbook(
    output_path: Path,
    source_path: Path,
    rows: list[SalesRow],
    profile: dict[str, int],
    region_code: str,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    months = sorted_month_keys(rows)
    month_labels = [f"{y}-{m:02d}" for y, m in months]

    readme = wb.create_sheet("README")
    readme.append(["Source File", str(source_path.name)])
    readme.append(["Generated At", datetime.now().isoformat(timespec="seconds")])
    readme.append(["Region", region_code])
    readme.append(["Rows Used", profile.get("kept_rows", 0)])
    for key in sorted(k for k in profile.keys() if k != "kept_rows"):
        readme.append([f"Skipped: {key}", profile[key]])
    autosize(readme)

    summary_sheet_name = safe_sheet_name(f"7-{region_code}", "7-Region")
    ws_sum = wb.create_sheet(summary_sheet_name)
    headers = ["No", "Township"]
    for label in month_labels:
        headers.extend([f"{label} Bot", f"{label} Lit"])
    headers.extend(["Total Bot", "Total Lit", "Total Amount"])
    write_header_row(ws_sum, 1, headers)

    town_totals = defaultdict(lambda: {"bot": 0.0, "lit": 0.0, "amt": 0.0})
    town_month = defaultdict(lambda: {"bot": 0.0, "lit": 0.0})
    town_sku_month = defaultdict(lambda: {"bot": 0.0, "lit": 0.0, "amt": 0.0})
    for r in rows:
        tkey = r.township
        mkey = (r.year, r.month)
        town_totals[tkey]["bot"] += r.sales_bot
        town_totals[tkey]["lit"] += r.liter
        town_totals[tkey]["amt"] += r.amount
        town_month[(tkey, mkey)]["bot"] += r.sales_bot
        town_month[(tkey, mkey)]["lit"] += r.liter
        sku_key = (tkey, r.stock_id, r.stock_name, r.ml, r.packing, mkey)
        town_sku_month[sku_key]["bot"] += r.sales_bot
        town_sku_month[sku_key]["lit"] += r.liter
        town_sku_month[sku_key]["amt"] += r.amount

    sorted_towns = sorted(town_totals.keys(), key=lambda t: (-town_totals[t]["lit"], t))
    for idx, township in enumerate(sorted_towns, start=1):
        row_data = [idx, township]
        for mkey in months:
            cell = town_month.get((township, mkey), {"bot": 0.0, "lit": 0.0})
            row_data.extend([cell["bot"], cell["lit"]])
        row_data.extend([town_totals[township]["bot"], town_totals[township]["lit"], town_totals[township]["amt"]])
        ws_sum.append(row_data)
    autosize(ws_sum)

    for township in sorted_towns:
        ws_name = safe_sheet_name(township, "Township")
        ws = wb.create_sheet(ws_name)
        sub_headers = ["Sr", "StockID", "StockName", "ML", "Packing"]
        for label in month_labels:
            sub_headers.extend([f"{label} Bot", f"{label} Lit"])
        sub_headers.extend(["Total Bot", "Total Lit", "Total Amount"])
        write_header_row(ws, 1, sub_headers)

        sku_totals = defaultdict(lambda: {"bot": 0.0, "lit": 0.0, "amt": 0.0})
        for key, values in town_sku_month.items():
            tkey, stock_id, stock_name, ml, packing, _ = key
            if tkey != township:
                continue
            sku_key = (stock_id, stock_name, ml, packing)
            sku_totals[sku_key]["bot"] += values["bot"]
            sku_totals[sku_key]["lit"] += values["lit"]
            sku_totals[sku_key]["amt"] += values["amt"]

        sorted_skus = sorted(sku_totals.keys(), key=lambda k: (-sku_totals[k]["lit"], k[1], k[0]))
        for idx, sku_key in enumerate(sorted_skus, start=1):
            stock_id, stock_name, ml, packing = sku_key
            row_data: list[Any] = [idx, stock_id, stock_name, ml, packing]
            for mkey in months:
                values = town_sku_month.get((township, stock_id, stock_name, ml, packing, mkey), {"bot": 0.0, "lit": 0.0, "amt": 0.0})
                row_data.extend([values["bot"], values["lit"]])
            row_data.extend([sku_totals[sku_key]["bot"], sku_totals[sku_key]["lit"], sku_totals[sku_key]["amt"]])
            ws.append(row_data)
        autosize(ws)

    wb.save(output_path)


def detect_base_sources(source_path: Path) -> dict[str, Any]:
    wb = load_workbook(source_path, data_only=False, read_only=False)
    has_table = "Table" in wb.sheetnames
    has_daily = "DailySales" in wb.sheetnames
    table_refs = []
    for ws in wb.worksheets:
        for tname in ws.tables:
            t = ws.tables[tname]
            table_refs.append((ws.title, t.name, t.ref))

    wb.close()
    return {
        "has_table_sheet": has_table,
        "has_dailysales_sheet": has_daily,
        "excel_tables": table_refs,
    }


def run(
    source_path: Path,
    output_dir: Path,
    prefix: str,
) -> tuple[Path, Path, dict[str, int], dict[str, Any], list[SalesRow], dict[str, dict[str, Any]]]:
    source_info = detect_base_sources(source_path)
    stock_map, customer_map = load_masters(source_path)
    rows, profile = extract_sales_rows(source_path, stock_map, customer_map)

    output_dir.mkdir(parents=True, exist_ok=True)
    sku_path = output_dir / f"{prefix}_SKU_Analysis.xlsx"
    township_path = output_dir / f"{prefix}_Township_Analysis.xlsx"

    build_sku_workbook(sku_path, source_path, rows, profile)
    build_township_workbook(township_path, source_path, rows, profile)
    return sku_path, township_path, profile, source_info, rows, customer_map


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate SKU and township reports from DailySales source.")
    parser.add_argument("--source", required=True, help="Path to source workbook (e.g., 9-MLM Table and DailySales-2026_Feb.xlsx)")
    parser.add_argument("--output-dir", default=".", help="Directory for output files")
    parser.add_argument("--prefix", default="generated", help="Output filename prefix")
    parser.add_argument("--emit-region-templates", action="store_true", help="Also emit region-named template-style workbooks.")
    parser.add_argument("--region-code", default="", help="Region code label for template sheets/files (e.g., MLM).")
    parser.add_argument("--template-sku-name", default="", help="Custom filename for region SKU template workbook.")
    parser.add_argument("--template-township-name", default="", help="Custom filename for region township template workbook.")
    args = parser.parse_args()

    source_path = Path(args.source).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()

    if not source_path.exists():
        raise SystemExit(f"Source file not found: {source_path}")

    sku_path, township_path, profile, source_info, rows, customer_map = run(source_path, output_dir, args.prefix)

    template_sku_path = None
    template_township_path = None
    if args.emit_region_templates:
        region_code = (args.region_code or args.prefix.split("_")[0] or "REGION").strip()
        sku_filename = args.template_sku_name.strip() or f"{region_code} Template SKU.xlsx"
        township_filename = args.template_township_name.strip() or f"7-{region_code} for Township Summary.xlsx"
        template_sku_path = output_dir / sku_filename
        template_township_path = output_dir / township_filename

        build_region_sku_template_workbook(
            template_sku_path,
            source_path,
            rows,
            profile,
            region_code,
            customer_map,
        )
        build_region_township_template_workbook(
            template_township_path,
            source_path,
            rows,
            profile,
            region_code,
        )

    print("Base source detection:")
    print(f" - Table sheet present: {source_info['has_table_sheet']}")
    print(f" - DailySales sheet present: {source_info['has_dailysales_sheet']}")
    print(" - Excel tables:")
    for ws_name, table_name, table_ref in source_info["excel_tables"]:
        print(f"   * {ws_name}.{table_name} -> {table_ref}")

    print("\nSales row profile:")
    for key in sorted(profile.keys()):
        print(f" - {key}: {profile[key]}")

    print("\nGenerated files:")
    print(f" - {sku_path}")
    print(f" - {township_path}")
    if template_sku_path and template_township_path:
        print(f" - {template_sku_path}")
        print(f" - {template_township_path}")


if __name__ == "__main__":
    main()
