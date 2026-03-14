from __future__ import annotations

import calendar
import html
import json
import os
import re
import socket
import time
from collections import Counter, OrderedDict
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path
from typing import Any

from flask import request
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = ROOT_DIR / "uploads"
WORKBOOK_REGISTRY_PATH = UPLOAD_DIR / "workbook_registry.json"
ACCESS_CONTROL_PATH = UPLOAD_DIR / "access_control.json"
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
MACRO_ENABLED_EXTENSIONS = {".xlsm", ".xltm"}
EXCEL_SHEET_TITLE_MAX_LENGTH = 31
EXCEL_SHEET_TITLE_INVALID_CHARS = set("\\/*?:[]")
ROLE_OWNER = "owner"
ROLE_RSM = "rsm"
ROLE_ASM = "asm"
ROLE_USER = "user"
ALLOWED_USER_ROLES = {ROLE_OWNER, ROLE_RSM, ROLE_ASM, ROLE_USER}
ALLOWED_VIEWER_ROLES = {"owner", "regional_manager", "rsm", "asm", "user"}
ALL_REGION_TOKEN = "ALL"
DEFAULT_REGION_TOKEN = "GLOBAL"

MONTH_LOOKUP = {
    name.lower(): idx
    for idx, name in enumerate(calendar.month_name)
    if idx
}
MONTH_LOOKUP.update(
    {
        name.lower(): idx
        for idx, name in enumerate(calendar.month_abbr)
        if idx
    }
)

REGION_STOP_WORDS = {
    "analysis",
    "detail",
    "for",
    "main",
    "master",
    "month",
    "monthly",
    "overview",
    "reference",
    "sheet",
    "sku",
    "summary",
    "township",
    "workbook",
}
REGION_STOP_WORDS.update({name.lower() for name in calendar.month_name if name})
REGION_STOP_WORDS.update({name.lower() for name in calendar.month_abbr if name})

TOWNSHIP_ALIAS = {
    "meiktila": "meiktila",
    "tharzi": "tharzi",
    "tarzi": "tharzi",
    "pyawbwe": "pyawbwe",
    "pyawbwe": "pyawbwe",
    "wanttwin": "wantwin",
    "wantwin": "wantwin",
    "mahaling": "mahlaing",
    "mahlaing": "mahlaing",
    "yamethin": "yamethin",
    "kyaukpadaung": "kyaukpandaung",
    "kyaukpandaung": "kyaukpandaung",
    "taungthar": "taungthar",
    "taugthar": "taungthar",
    "myingyan": "myingan",
    "myingan": "myingan",
    "bagan": "bagan",
    "pakokku": "pakokku",
    "pakoku": "pakokku",
}


def normalize_token(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def canonical_sheet_name(sheet_name: str) -> str:
    token = normalize_token(sheet_name)
    return TOWNSHIP_ALIAS.get(token, token)


def parse_month_from_text(text: str) -> int | None:
    tokens = [t for t in re.split(r"[^a-z]+", text.lower()) if t]
    matched = [MONTH_LOOKUP[token] for token in tokens if token in MONTH_LOOKUP]
    if not matched:
        return None
    if len(set(matched)) > 1:
        return None
    return matched[0]


def extract_year_from_text(text: str) -> int | None:
    match = re.search(r"(20\d{2})", text)
    if not match:
        return None
    return int(match.group(1))


def parse_month_header(value: Any, fallback_year: int | None) -> tuple[int, int] | None:
    if isinstance(value, datetime):
        return value.year, value.month
    if isinstance(value, date):
        return value.year, value.month

    if isinstance(value, str):
        month_num = parse_month_from_text(value)
        if month_num is None:
            return None
        year = extract_year_from_text(value) or fallback_year
        if year is None:
            return None
        return year, month_num

    return None


def parse_month_header_loose(
    value: Any, fallback_year: int | None
) -> tuple[int | None, int] | None:
    if isinstance(value, datetime):
        return value.year, value.month
    if isinstance(value, date):
        return value.year, value.month

    if isinstance(value, str):
        month_num = parse_month_from_text(value)
        if month_num is None:
            return None
        year = extract_year_from_text(value) or fallback_year
        return year, month_num

    return None


def parse_numeric(value: Any) -> int | float | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        try:
            number = float(stripped)
        except ValueError:
            return None
    elif isinstance(value, (int, float)):
        number = float(value)
    else:
        return None

    if number.is_integer():
        return int(number)
    return round(number, 4)


def parse_year_context(worksheet, header_row: int) -> int | None:
    for row_idx in range(1, header_row):
        for col_idx in range(1, min(worksheet.max_column, 30) + 1):
            value = worksheet.cell(row=row_idx, column=col_idx).value
            if isinstance(value, (int, float)) and 2000 <= int(value) <= 2100:
                return int(value)
            if isinstance(value, str):
                year = extract_year_from_text(value)
                if year:
                    return year
    return None


def find_header_layout(worksheet) -> dict[str, int] | None:
    max_scan_col = min(worksheet.max_column, 60)
    for row_idx in range(1, min(worksheet.max_row, 12) + 1):
        columns: dict[str, int] = {}
        for col_idx in range(1, max_scan_col + 1):
            token = normalize_token(worksheet.cell(row=row_idx, column=col_idx).value)
            if not token:
                continue
            if token == "sr":
                columns["sr"] = col_idx
            elif token == "productname":
                columns["product_name"] = col_idx
            elif token == "ml":
                columns["ml"] = col_idx
            elif token == "packing":
                columns["packing"] = col_idx

        required = {"product_name", "ml", "packing"}
        if required.issubset(columns):
            columns["header_row"] = row_idx
            columns["metrics_row"] = row_idx + 1
            return columns
    return None


def parse_month_groups(worksheet, layout: dict[str, int]) -> list[dict[str, Any]]:
    header_row = layout["header_row"]
    metrics_row = layout["metrics_row"]
    fallback_year = parse_year_context(worksheet, header_row)
    start_col = layout["packing"] + 1

    month_groups: list[dict[str, Any]] = []
    seen_keys: set[str] = set()

    for col_idx in range(start_col, worksheet.max_column - 1):
        metric_tokens = [
            normalize_token(worksheet.cell(row=metrics_row, column=col_idx + offset).value)
            for offset in range(3)
        ]
        if metric_tokens[:2] != ["pk", "bottle"]:
            continue
        if metric_tokens[2] not in {"liter", "litre"}:
            continue

        header_value = worksheet.cell(row=header_row, column=col_idx).value
        parsed_month = parse_month_header(header_value, fallback_year)
        if not parsed_month:
            continue

        year, month = parsed_month
        month_key = f"{year:04d}-{month:02d}"
        if month_key in seen_keys:
            continue

        seen_keys.add(month_key)
        month_groups.append(
            {
                "key": month_key,
                "label": f"{calendar.month_abbr[month]} {year}",
                "year": year,
                "month": month,
                "col": col_idx,
            }
        )

    month_groups.sort(key=lambda item: item["key"])
    return month_groups


def row_identity_key(product_name: str, ml: Any, packing: Any) -> str:
    ml_token = "" if ml is None else str(ml).strip()
    packing_token = normalize_token(packing)
    return f"{normalize_token(product_name)}|{ml_token}|{packing_token}"


def parse_rows(worksheet, layout: dict[str, int], month_groups: list[dict[str, Any]]) -> list[dict[str, Any]]:
    sr_col = layout.get("sr")
    product_col = layout["product_name"]
    ml_col = layout["ml"]
    packing_col = layout["packing"]
    data_start = layout["metrics_row"] + 1

    parsed_rows: list[dict[str, Any]] = []
    blank_streak = 0

    for row_idx in range(data_start, worksheet.max_row + 1):
        sr_value = worksheet.cell(row=row_idx, column=sr_col).value if sr_col else None
        product_value = worksheet.cell(row=row_idx, column=product_col).value
        ml_value = worksheet.cell(row=row_idx, column=ml_col).value
        packing_value = worksheet.cell(row=row_idx, column=packing_col).value

        values_by_month: dict[str, dict[str, int | float | None]] = {}
        has_month_data = False

        for group in month_groups:
            col_idx = group["col"]
            pk_val = parse_numeric(worksheet.cell(row=row_idx, column=col_idx).value)
            bottle_val = parse_numeric(worksheet.cell(row=row_idx, column=col_idx + 1).value)
            liter_val = parse_numeric(worksheet.cell(row=row_idx, column=col_idx + 2).value)
            values_by_month[group["key"]] = {
                "pk": pk_val,
                "bottle": bottle_val,
                "liter": liter_val,
            }
            if pk_val is not None or bottle_val is not None or liter_val is not None:
                has_month_data = True

        has_identity = any(
            value not in (None, "") for value in (sr_value, product_value, ml_value, packing_value)
        )

        if not has_identity and not has_month_data:
            blank_streak += 1
            if blank_streak >= 20:
                break
            continue

        blank_streak = 0
        if product_value in (None, ""):
            continue

        product_name = str(product_value).strip()
        parsed_rows.append(
            {
                "sr": sr_value,
                "product_name": product_name,
                "ml": ml_value,
                "packing": packing_value,
                "row_key": row_identity_key(product_name, ml_value, packing_value),
                "values": values_by_month,
            }
        )

    return parsed_rows


def parse_fixed_sheet(worksheet, sheet_name: str) -> dict[str, Any] | None:
    layout = find_header_layout(worksheet)
    if not layout:
        return None

    month_groups = parse_month_groups(worksheet, layout)
    if len(month_groups) < 2:
        return None

    parsed_rows = parse_rows(worksheet, layout, month_groups)
    if not parsed_rows:
        return None

    return {
        "sheet_name": sheet_name,
        "canonical": canonical_sheet_name(sheet_name),
        "months": [
            {
                "key": group["key"],
                "label": group["label"],
                "year": group["year"],
                "month": group["month"],
            }
            for group in month_groups
        ],
        "rows": parsed_rows,
    }


def parse_workbook(path: Path) -> OrderedDict[str, dict[str, Any]]:
    # Use normal mode because this parser relies on random cell access.
    workbook = load_workbook(path, read_only=False, data_only=True, keep_links=False)
    parsed_sheets: OrderedDict[str, dict[str, Any]] = OrderedDict()

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        parsed_sheet = parse_fixed_sheet(worksheet, sheet_name)
        if not parsed_sheet:
            continue

        canonical = parsed_sheet["canonical"]
        if canonical in parsed_sheets:
            # Prefer the first matching sheet in workbook order.
            continue
        parsed_sheets[canonical] = parsed_sheet

    return parsed_sheets


@lru_cache(maxsize=16)
def parse_workbook_cached(path_str: str, mtime_ns: int) -> OrderedDict[str, dict[str, Any]]:
    _ = mtime_ns
    return parse_workbook(Path(path_str))


@lru_cache(maxsize=8)
def load_workbook_cached(path_str: str, mtime_ns: int):
    _ = mtime_ns
    return load_workbook(Path(path_str), read_only=False, data_only=True, keep_links=False)


@lru_cache(maxsize=8)
def workbook_sheetnames_cached(path_str: str, mtime_ns: int) -> tuple[str, ...]:
    _ = mtime_ns
    workbook = load_workbook(
        Path(path_str),
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        visible_names = tuple(
            worksheet.title
            for worksheet in workbook.worksheets
            if getattr(worksheet, "sheet_state", "visible") == "visible"
        )
        if visible_names:
            return visible_names
        return tuple(workbook.sheetnames)
    finally:
        workbook.close()


def normalize_sheet_name_list(sheet_names: tuple[str, ...] | list[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for raw_name in sheet_names:
        if not isinstance(raw_name, str):
            continue
        sheet_name = raw_name
        if not sheet_name or sheet_name in seen:
            continue
        seen.add(sheet_name)
        ordered.append(sheet_name)
    return ordered


def normalize_viewer_role(raw_role: str | None) -> str:
    role = normalize_token(raw_role)
    if role in {"owner"}:
        return ROLE_OWNER
    if role in {"regionalmanager", "regional_manager", "rsm"}:
        return ROLE_RSM
    if role in {"asm"}:
        return ROLE_ASM
    if role in {"user", "staff"}:
        return ROLE_USER
    return ROLE_OWNER


def normalize_user_role(raw_role: str | None, default: str = ROLE_USER) -> str:
    role = normalize_viewer_role(raw_role)
    if role in ALLOWED_USER_ROLES:
        return role
    return default


def normalize_username(raw_username: Any) -> str:
    if raw_username is None:
        return ""
    value = str(raw_username).strip().lower()
    if not value:
        return ""
    cleaned = re.sub(r"[^a-z0-9_.-]+", "_", value).strip("._-")
    return cleaned


def normalize_display_name(raw_display_name: Any, fallback_username: str) -> str:
    if raw_display_name is None:
        return fallback_username
    display_name = str(raw_display_name).strip()
    if not display_name:
        return fallback_username
    return re.sub(r"\s+", " ", display_name)


def normalize_region_scope(raw_region: str | None) -> str:
    token = normalize_token(raw_region).upper()
    if not token or token in {"ALL", "ANY"}:
        return ALL_REGION_TOKEN
    return token


def normalize_region_token(raw_region: Any) -> str:
    token = normalize_token(raw_region).upper()
    if not token:
        return DEFAULT_REGION_TOKEN
    if token in {"ALL", "ANY"}:
        return DEFAULT_REGION_TOKEN
    return token


def infer_region_from_filename(path: Path) -> str:
    for token in re.split(r"[^A-Za-z0-9]+", path.stem):
        if not token:
            continue
        lowered = token.lower()
        if lowered in REGION_STOP_WORDS:
            continue
        if re.fullmatch(r"20\d{2}", token):
            continue
        if token.isdigit():
            continue
        cleaned = re.sub(r"[^A-Za-z0-9]+", "", token)
        if len(cleaned) < 2:
            continue
        return cleaned.upper()
    return DEFAULT_REGION_TOKEN


def load_workbook_registry() -> dict[str, dict[str, str]]:
    if not WORKBOOK_REGISTRY_PATH.exists():
        return {}
    try:
        raw = WORKBOOK_REGISTRY_PATH.read_text(encoding="utf-8")
        payload = json.loads(raw)
    except (OSError, ValueError):
        return {}

    if not isinstance(payload, dict):
        return {}
    files_raw = payload.get("files")
    if not isinstance(files_raw, dict):
        return {}

    normalized: dict[str, dict[str, str]] = {}
    for workbook_name, meta in files_raw.items():
        if not isinstance(workbook_name, str):
            continue
        if not isinstance(meta, dict):
            continue
        region = normalize_region_token(meta.get("region"))
        normalized[workbook_name] = {"region": region}
    return normalized


def save_workbook_registry(registry: dict[str, dict[str, str]]) -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    payload = {
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "files": registry,
    }
    WORKBOOK_REGISTRY_PATH.write_text(
        json.dumps(payload, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def parse_token_list(raw_value: Any) -> list[str]:
    if raw_value is None:
        return []
    if isinstance(raw_value, (list, tuple, set)):
        items = raw_value
    else:
        items = re.split(r"[,\n;]+", str(raw_value))
    tokens: list[str] = []
    for item in items:
        text = str(item).strip()
        if text:
            tokens.append(text)
    return tokens


def normalize_region_list(raw_value: Any) -> list[str]:
    regions: list[str] = []
    seen: set[str] = set()
    for token in parse_token_list(raw_value):
        region = normalize_region_token(token)
        if not region or region in seen:
            continue
        seen.add(region)
        regions.append(region)
    return regions


def normalize_township_list(raw_value: Any) -> list[str]:
    townships: list[str] = []
    seen: set[str] = set()
    for token in parse_token_list(raw_value):
        canonical = canonical_sheet_name(token)
        if not canonical or canonical in seen:
            continue
        seen.add(canonical)
        townships.append(canonical)
    return townships


def default_access_control() -> dict[str, Any]:
    return {
        "users": {
            "owner": {
                "role": ROLE_OWNER,
                "display_name": "Owner",
            }
        },
        "rsm_regions": {},
        "user_to_rsm": {},
        "asm_townships": {},
    }


def load_access_control() -> dict[str, Any]:
    defaults = default_access_control()
    if not ACCESS_CONTROL_PATH.exists():
        return defaults

    try:
        raw = ACCESS_CONTROL_PATH.read_text(encoding="utf-8")
        payload = json.loads(raw)
    except (OSError, ValueError):
        return defaults
    if not isinstance(payload, dict):
        return defaults

    users_raw = payload.get("users")
    users: dict[str, dict[str, str]] = {}
    if isinstance(users_raw, dict):
        for username_raw, meta in users_raw.items():
            username = normalize_username(username_raw)
            if not username:
                continue
            role = ROLE_USER
            display_name = username
            if isinstance(meta, dict):
                role = normalize_user_role(meta.get("role"), default=ROLE_USER)
                display_name = normalize_display_name(meta.get("display_name"), username)
            users[username] = {
                "role": role,
                "display_name": display_name,
            }

    if "owner" not in users:
        users["owner"] = {"role": ROLE_OWNER, "display_name": "Owner"}
    else:
        users["owner"]["role"] = ROLE_OWNER

    rsm_regions: dict[str, list[str]] = {}
    rsm_regions_raw = payload.get("rsm_regions")
    if isinstance(rsm_regions_raw, dict):
        for username_raw, regions_raw in rsm_regions_raw.items():
            username = normalize_username(username_raw)
            if not username:
                continue
            regions = normalize_region_list(regions_raw)
            rsm_regions[username] = regions
            if username not in users:
                users[username] = {"role": ROLE_RSM, "display_name": username}

    user_to_rsm: dict[str, str] = {}
    user_to_rsm_raw = payload.get("user_to_rsm")
    if isinstance(user_to_rsm_raw, dict):
        for username_raw, rsm_raw in user_to_rsm_raw.items():
            username = normalize_username(username_raw)
            manager_rsm = normalize_username(rsm_raw)
            if not username or not manager_rsm:
                continue
            user_to_rsm[username] = manager_rsm
            if username not in users:
                users[username] = {"role": ROLE_USER, "display_name": username}
            if manager_rsm not in users:
                users[manager_rsm] = {"role": ROLE_RSM, "display_name": manager_rsm}

    asm_townships: dict[str, dict[str, list[str]]] = {}
    asm_townships_raw = payload.get("asm_townships")
    if isinstance(asm_townships_raw, dict):
        for asm_raw, region_map_raw in asm_townships_raw.items():
            asm_username = normalize_username(asm_raw)
            if not asm_username or not isinstance(region_map_raw, dict):
                continue
            region_map: dict[str, list[str]] = {}
            for region_raw, towns_raw in region_map_raw.items():
                region = normalize_region_token(region_raw)
                townships = normalize_township_list(towns_raw)
                if townships:
                    region_map[region] = townships
            if region_map:
                asm_townships[asm_username] = region_map
                if asm_username not in users:
                    users[asm_username] = {"role": ROLE_ASM, "display_name": asm_username}

    for username, regions in rsm_regions.items():
        if regions and users[username]["role"] != ROLE_OWNER:
            users[username]["role"] = ROLE_RSM
    for username, regions in asm_townships.items():
        if regions and users[username]["role"] not in {ROLE_OWNER, ROLE_RSM}:
            users[username]["role"] = ROLE_ASM

    return {
        "users": users,
        "rsm_regions": rsm_regions,
        "user_to_rsm": user_to_rsm,
        "asm_townships": asm_townships,
    }


def save_access_control(access: dict[str, Any]) -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    payload = {
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "users": access.get("users", {}),
        "rsm_regions": access.get("rsm_regions", {}),
        "user_to_rsm": access.get("user_to_rsm", {}),
        "asm_townships": access.get("asm_townships", {}),
    }
    ACCESS_CONTROL_PATH.write_text(
        json.dumps(payload, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def ensure_access_user(
    access: dict[str, Any],
    username: str,
    role: str = ROLE_USER,
    display_name: str | None = None,
) -> dict[str, str]:
    users = access.setdefault("users", {})
    record = users.get(username)
    if not isinstance(record, dict):
        record = {}
        users[username] = record
    if "role" not in record:
        record["role"] = normalize_user_role(role, default=ROLE_USER)
    if display_name:
        record["display_name"] = normalize_display_name(display_name, username)
    if "display_name" not in record:
        record["display_name"] = username
    if username == "owner":
        record["role"] = ROLE_OWNER
    return record


def all_known_regions(access: dict[str, Any], entries: list[dict[str, Any]]) -> list[str]:
    regions = {str(entry["region"]) for entry in entries}
    for values in access.get("rsm_regions", {}).values():
        for region in values:
            regions.add(normalize_region_token(region))
    for region_map in access.get("asm_townships", {}).values():
        if not isinstance(region_map, dict):
            continue
        for region in region_map:
            regions.add(normalize_region_token(region))
    return sorted(regions)


def derive_allowed_regions(
    access: dict[str, Any],
    username: str,
    role: str,
    regions_universe: list[str],
) -> list[str]:
    region_set = set(regions_universe)
    if role == ROLE_OWNER:
        return sorted(region_set)
    if role == ROLE_RSM:
        assigned = [
            normalize_region_token(region)
            for region in access.get("rsm_regions", {}).get(username, [])
        ]
        return [region for region in assigned if region in region_set]
    if role == ROLE_ASM:
        asm_regions = []
        region_map = access.get("asm_townships", {}).get(username, {})
        if isinstance(region_map, dict):
            asm_regions = [normalize_region_token(region) for region in region_map.keys()]
        return [region for region in asm_regions if region in region_set]

    manager_rsm = access.get("user_to_rsm", {}).get(username)
    if manager_rsm:
        assigned = [
            normalize_region_token(region)
            for region in access.get("rsm_regions", {}).get(manager_rsm, [])
        ]
        return [region for region in assigned if region in region_set]
    return []


def resolve_selected_region(
    requested_region: str | None,
    allowed_regions: list[str],
    allow_all_regions: bool,
) -> str:
    requested = normalize_region_scope(requested_region)
    if allow_all_regions and requested == ALL_REGION_TOKEN:
        return ALL_REGION_TOKEN
    if requested in allowed_regions:
        return requested
    if allowed_regions:
        return allowed_regions[0]
    return ALL_REGION_TOKEN


def current_request_user_and_region() -> tuple[str, str | None]:
    user_raw = request.args.get("user")
    if not user_raw:
        user_raw = request.form.get("user")
    if not user_raw and request.is_json:
        payload = request.get_json(silent=True) or {}
        if isinstance(payload, dict):
            user_raw = payload.get("user")
    username = normalize_username(user_raw) or "owner"

    region_raw = request.args.get("region")
    if region_raw is None:
        region_raw = request.form.get("region")
    if region_raw is None and request.is_json:
        payload = request.get_json(silent=True) or {}
        if isinstance(payload, dict):
            region_raw = payload.get("region")
    return username, region_raw


def resolve_principal(
    requested_user: str | None = None,
    requested_region: str | None = None,
) -> dict[str, Any]:
    access = load_access_control()
    requested_username = normalize_username(requested_user) or "owner"
    users = access.get("users", {})
    if requested_username not in users:
        requested_username = "owner"
    user_record = users.get(requested_username, {})
    role = normalize_user_role(user_record.get("role"), default=ROLE_USER)

    entries = discover_workbook_entries()
    regions_universe = all_known_regions(access, entries)
    allowed_regions = derive_allowed_regions(access, requested_username, role, regions_universe)
    allow_all_regions = role == ROLE_OWNER
    selected_region = resolve_selected_region(
        requested_region,
        allowed_regions,
        allow_all_regions=allow_all_regions,
    )

    asm_townships = access.get("asm_townships", {}).get(requested_username, {})
    allowed_townships = None
    if role == ROLE_ASM and isinstance(asm_townships, dict):
        region_for_townships = selected_region
        if region_for_townships == ALL_REGION_TOKEN and allowed_regions:
            region_for_townships = allowed_regions[0]
        allowed_townships = set(
            normalize_township_list(asm_townships.get(region_for_townships, []))
        )

    return {
        "username": requested_username,
        "role": role,
        "display_name": normalize_display_name(
            user_record.get("display_name"),
            requested_username,
        ),
        "access": access,
        "entries": entries,
        "regions_universe": regions_universe,
        "allowed_regions": allowed_regions,
        "selected_region": selected_region,
        "allow_all_regions": allow_all_regions,
        "allowed_townships": allowed_townships,
        "rsm_regions": normalize_region_list(
            access.get("rsm_regions", {}).get(requested_username, [])
        ),
    }


def principal_from_request() -> dict[str, Any]:
    requested_user, requested_region = current_request_user_and_region()
    return resolve_principal(
        requested_user=requested_user,
        requested_region=requested_region,
    )


def principal_can_upload_to_region(principal: dict[str, Any], region: str) -> bool:
    role = principal["role"]
    if role == ROLE_OWNER:
        return True
    if role == ROLE_RSM:
        return region in set(principal.get("rsm_regions", []))
    return False


def principal_can_manage_rsm(principal: dict[str, Any]) -> bool:
    return principal["role"] == ROLE_OWNER


def principal_can_manage_asm(principal: dict[str, Any]) -> bool:
    return principal["role"] in {ROLE_OWNER, ROLE_RSM}


def principal_can_manage_region(principal: dict[str, Any], region: str) -> bool:
    if principal["role"] == ROLE_OWNER:
        return True
    if principal["role"] == ROLE_RSM:
        return region in set(principal.get("rsm_regions", []))
    return False


def filter_data_by_allowed_townships(
    data: dict[str, Any],
    allowed_townships: set[str] | None,
) -> dict[str, Any]:
    if allowed_townships is None:
        return data

    filtered_main = {
        canonical: sheet
        for canonical, sheet in data["main"].items()
        if canonical in allowed_townships
    }
    filtered_reference = {
        canonical: sheet
        for canonical, sheet in data["reference"].items()
        if canonical in allowed_townships
    }

    allowed_main_names = {sheet["sheet_name"] for sheet in filtered_main.values()}
    allowed_reference_names = {sheet["sheet_name"] for sheet in filtered_reference.values()}

    filtered_main_tabs = [
        tab
        for tab in data["main_sheet_tabs"]
        if tab.get("canonical") in allowed_townships and tab.get("sheet_name") in allowed_main_names
    ]
    filtered_reference_tabs = [
        tab
        for tab in data["reference_sheet_tabs"]
        if tab.get("canonical") in allowed_townships
        and tab.get("sheet_name") in allowed_reference_names
    ]
    filtered_sheet_index = [
        row
        for row in data["sheet_index"]
        if row.get("canonical") in allowed_townships
    ]

    return {
        **data,
        "main": filtered_main,
        "reference": filtered_reference,
        "sheet_index": filtered_sheet_index,
        "main_sheet_tabs": filtered_main_tabs,
        "reference_sheet_tabs": filtered_reference_tabs,
        "main_sheet_names": normalize_sheet_name_list(
            [tab["sheet_name"] for tab in filtered_main_tabs]
        ),
        "reference_sheet_names": normalize_sheet_name_list(
            [tab["sheet_name"] for tab in filtered_reference_tabs]
        ),
    }

def discover_workbook_entries() -> list[dict[str, Any]]:
    registry = load_workbook_registry()
    entries: list[dict[str, Any]] = []
    seen_names: set[str] = set()
    search_dirs = [ROOT_DIR]
    if UPLOAD_DIR.exists():
        search_dirs.append(UPLOAD_DIR)

    for folder in search_dirs:
        for path in folder.iterdir():
            if not path.is_file():
                continue
            if path.name.startswith("~$"):
                continue
            if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
                continue
            if path.name in seen_names:
                # Keep first entry if duplicate filenames exist in different folders.
                continue

            meta = registry.get(path.name)
            region_token = (
                normalize_region_token(meta.get("region"))
                if meta and isinstance(meta, dict)
                else infer_region_from_filename(path)
            )
            seen_names.add(path.name)
            entries.append({"name": path.name, "path": path, "region": region_token})

    entries.sort(key=lambda item: str(item["name"]).lower())
    return entries


def workbook_name_score(path: Path, role: str) -> int:
    token = normalize_token(path.stem)
    if role == "main":
        keywords = ("overview", "main", "master", "mhl", "summary", "analysis")
    else:
        keywords = ("detail", "reference", "township", "breakdown", "summary", "mtl")

    score = 0
    for idx, keyword in enumerate(keywords):
        if keyword in token:
            score += max(1, 12 - (idx * 2))
    return score


def default_workbook(workbooks: list[Path], role: str, avoid_name: str | None = None) -> Path:
    candidates = [path for path in workbooks if path.name != avoid_name] or workbooks
    ranked = sorted(
        candidates,
        key=lambda path: (
            workbook_name_score(path, role),
            path.stat().st_mtime_ns,
            path.name.lower(),
        ),
        reverse=True,
    )
    return ranked[0]


def resolve_workbook_pair(
    main_name: str | None,
    reference_name: str | None,
    viewer_role: str | None = None,
    region_scope: str | None = None,
    allowed_regions: list[str] | None = None,
    allow_all_regions: bool | None = None,
) -> dict[str, Any]:
    entries = discover_workbook_entries()
    if not entries:
        raise FileNotFoundError("No supported Excel files found in this folder.")

    resolved_role = normalize_viewer_role(viewer_role)
    if allow_all_regions is None:
        allow_all_regions = resolved_role == ROLE_OWNER

    requested_region = normalize_region_scope(region_scope)

    scoped_entries = entries
    if allowed_regions is not None:
        allowed_region_set = {
            normalize_region_token(region)
            for region in allowed_regions
            if region is not None
        }
        scoped_entries = [
            entry for entry in entries if entry["region"] in allowed_region_set
        ]

    available_regions = sorted({str(entry["region"]) for entry in scoped_entries})
    selected_region = resolve_selected_region(
        requested_region,
        available_regions,
        allow_all_regions=allow_all_regions,
    )

    if selected_region == ALL_REGION_TOKEN and allow_all_regions:
        visible_entries = scoped_entries
    else:
        visible_entries = [
            entry for entry in scoped_entries if entry["region"] == selected_region
        ]

    if not visible_entries:
        if selected_region == ALL_REGION_TOKEN:
            raise FileNotFoundError("No supported Excel files found in this folder.")
        raise FileNotFoundError(
            f"No supported Excel files found for region: {selected_region}."
        )

    visible_paths = [entry["path"] for entry in visible_entries]
    by_name = {str(entry["name"]): entry["path"] for entry in visible_entries}
    default_main = default_workbook(visible_paths, "main")
    default_reference = default_workbook(
        visible_paths,
        "reference",
        avoid_name=default_main.name if len(visible_paths) > 1 else None,
    )

    if main_name and main_name not in by_name:
        raise ValueError(f"Unknown main workbook for current scope: {main_name}")
    if reference_name and reference_name not in by_name:
        raise ValueError(f"Unknown reference workbook for current scope: {reference_name}")

    main_path = by_name.get(main_name) if main_name else default_main
    reference_path = by_name.get(reference_name) if reference_name else default_reference

    return {
        "available": [str(entry["name"]) for entry in visible_entries],
        "regions": available_regions,
        "viewer_role": resolved_role,
        "selected_region": selected_region,
        "default_main": default_main.name,
        "default_reference": default_reference.name,
        "main_path": main_path,
        "reference_path": reference_path,
    }


def load_viewer_data(main_path: Path, reference_path: Path) -> dict[str, Any]:
    main_stat = main_path.stat()
    ref_stat = reference_path.stat()
    main_data = parse_workbook_cached(str(main_path), main_stat.st_mtime_ns)
    reference_data = parse_workbook_cached(str(reference_path), ref_stat.st_mtime_ns)
    version = f"{main_stat.st_mtime_ns}:{ref_stat.st_mtime_ns}"
    try:
        main_sheet_names = normalize_sheet_name_list(
            workbook_sheetnames_cached(str(main_path), main_stat.st_mtime_ns)
        )
    except Exception:
        main_sheet_names = normalize_sheet_name_list(
            [sheet["sheet_name"] for sheet in main_data.values()]
        )
    try:
        reference_sheet_names = normalize_sheet_name_list(
            workbook_sheetnames_cached(str(reference_path), ref_stat.st_mtime_ns)
        )
    except Exception:
        reference_sheet_names = normalize_sheet_name_list(
            [sheet["sheet_name"] for sheet in reference_data.values()]
        )

    sheet_index = []
    fixed_by_sheet_name: dict[str, dict[str, Any]] = {}
    for canonical, main_sheet in main_data.items():
        reference_sheet = reference_data.get(canonical)
        fixed_by_sheet_name[main_sheet["sheet_name"]] = {
            "canonical": canonical,
            "has_reference": bool(reference_sheet),
        }
        sheet_index.append(
            {
                "canonical": canonical,
                "main_sheet_name": main_sheet["sheet_name"],
                "reference_sheet_name": (
                    reference_sheet["sheet_name"] if reference_sheet else None
                ),
                "main_month_count": len(main_sheet["months"]),
                "reference_month_count": (
                    len(reference_sheet["months"]) if reference_sheet else 0
                ),
                "main_row_count": len(main_sheet["rows"]),
            }
        )

    main_sheet_tabs = []
    for sheet_name in main_sheet_names:
        fixed_info = fixed_by_sheet_name.get(sheet_name)
        main_sheet_tabs.append(
            {
                "sheet_name": sheet_name,
                "canonical": fixed_info["canonical"] if fixed_info else None,
                "filterable": bool(fixed_info),
                "has_reference": fixed_info["has_reference"] if fixed_info else False,
            }
        )

    fixed_reference_by_sheet_name: dict[str, dict[str, Any]] = {}
    for canonical, reference_sheet in reference_data.items():
        fixed_reference_by_sheet_name[reference_sheet["sheet_name"]] = {
            "canonical": canonical,
            "has_main": canonical in main_data,
        }

    reference_sheet_tabs = []
    for sheet_name in reference_sheet_names:
        fixed_info = fixed_reference_by_sheet_name.get(sheet_name)
        canonical = fixed_info["canonical"] if fixed_info else None
        reference_sheet_tabs.append(
            {
                "sheet_name": sheet_name,
                "canonical": canonical,
                "filterable": bool(fixed_info),
                "has_main": fixed_info["has_main"] if fixed_info else False,
            }
        )

    return {
        "main_workbook": main_path.name,
        "reference_workbook": reference_path.name,
        "version": version,
        "main_sheet_names": main_sheet_names,
        "reference_sheet_names": reference_sheet_names,
        "main": main_data,
        "reference": reference_data,
        "sheet_index": sheet_index,
        "main_sheet_tabs": main_sheet_tabs,
        "reference_sheet_tabs": reference_sheet_tabs,
    }



def clear_data_caches() -> None:
    parse_workbook_cached.cache_clear()
    load_workbook_cached.cache_clear()
    workbook_sheetnames_cached.cache_clear()

def safe_uploaded_filename(original_name: str) -> str:
    path = Path(original_name)
    stem = re.sub(r"[^A-Za-z0-9._ -]+", "_", path.stem).strip(" ._")
    if not stem:
        stem = "workbook"
    suffix = path.suffix.lower()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    return f"{stem}__{timestamp}{suffix}"


def upload_region_for_file(upload_region: str | None, original_name: str) -> str:
    scoped_region = normalize_region_scope(upload_region)
    if scoped_region != ALL_REGION_TOKEN:
        return scoped_region
    return infer_region_from_filename(Path(original_name))


def request_json_payload() -> dict[str, Any]:
    payload = request.get_json(silent=True)
    if isinstance(payload, dict):
        return payload
    return {}


def request_value(key: str) -> Any:
    if key in request.args:
        return request.args.get(key)
    if key in request.form:
        return request.form.get(key)
    payload = request_json_payload()
    return payload.get(key)


def workbook_selection_for_principal(
    principal: dict[str, Any],
    main_name: str | None,
    reference_name: str | None,
) -> dict[str, Any]:
    return resolve_workbook_pair(
        main_name,
        reference_name,
        viewer_role=principal["role"],
        region_scope=principal["selected_region"],
        allowed_regions=principal["allowed_regions"],
        allow_all_regions=principal["allow_all_regions"],
    )


def collect_region_townships(entries: list[dict[str, Any]]) -> dict[str, list[str]]:
    region_map: dict[str, set[str]] = {}
    for entry in entries:
        path = entry["path"]
        region = str(entry["region"])
        try:
            stat = path.stat()
            parsed = parse_workbook_cached(str(path), stat.st_mtime_ns)
        except Exception:
            continue
        canonical_names = set(parsed.keys())
        if not canonical_names:
            continue
        if region not in region_map:
            region_map[region] = set()
        region_map[region].update(canonical_names)

    return {
        region: sorted(names)
        for region, names in sorted(region_map.items())
    }


def user_summary(access: dict[str, Any], username: str) -> dict[str, Any]:
    users = access.get("users", {})
    user_record = users.get(username, {})
    role = normalize_user_role(user_record.get("role"), default=ROLE_USER)
    user_to_rsm = access.get("user_to_rsm", {})
    rsm_regions = normalize_region_list(access.get("rsm_regions", {}).get(username, []))
    asm_regions = sorted(
        normalize_region_token(region)
        for region in access.get("asm_townships", {}).get(username, {}).keys()
    )
    return {
        "username": username,
        "display_name": normalize_display_name(user_record.get("display_name"), username),
        "role": role,
        "assigned_rsm": user_to_rsm.get(username),
        "rsm_regions": rsm_regions,
        "asm_regions": asm_regions,
    }


def users_visible_to_principal(principal: dict[str, Any]) -> list[dict[str, Any]]:
    access = principal["access"]
    user_names = sorted(access.get("users", {}).keys())
    role = principal["role"]
    username = principal["username"]

    if role == ROLE_OWNER:
        return [user_summary(access, user_name) for user_name in user_names]

    if role == ROLE_RSM:
        visible = {username}
        for candidate, assigned_rsm in access.get("user_to_rsm", {}).items():
            if assigned_rsm == username:
                visible.add(candidate)
        return [user_summary(access, user_name) for user_name in sorted(visible)]

    return [user_summary(access, username)]


def access_context_payload(principal: dict[str, Any]) -> dict[str, Any]:
    access = principal["access"]
    users = users_visible_to_principal(principal)
    visible_usernames = {user["username"] for user in users}
    entries = principal["entries"]

    if principal["allow_all_regions"]:
        region_options = principal["regions_universe"]
    else:
        region_options = principal["allowed_regions"]

    if principal["role"] == ROLE_OWNER:
        selected_region = principal["selected_region"]
    else:
        selected_region = resolve_selected_region(
            principal["selected_region"],
            region_options,
            allow_all_regions=False,
        )

    region_townships_all = collect_region_townships(entries)
    visible_region_townships = {
        region: region_townships_all.get(region, [])
        for region in region_options
    }

    rsm_regions_source = access.get("rsm_regions", {})
    rsm_regions_visible: dict[str, list[str]] = {}
    if principal["role"] == ROLE_OWNER:
        for username, regions in rsm_regions_source.items():
            normalized_username = normalize_username(username)
            if not normalized_username:
                continue
            rsm_regions_visible[normalized_username] = normalize_region_list(regions)
    elif principal["role"] == ROLE_RSM:
        rsm_regions_visible[principal["username"]] = normalize_region_list(
            rsm_regions_source.get(principal["username"], [])
        )

    user_to_rsm_source = access.get("user_to_rsm", {})
    user_to_rsm_visible: dict[str, str] = {}
    for username, manager in user_to_rsm_source.items():
        normalized_username = normalize_username(username)
        normalized_manager = normalize_username(manager)
        if not normalized_username or not normalized_manager:
            continue
        if (
            principal["role"] == ROLE_OWNER
            or normalized_username in visible_usernames
            or normalized_manager == principal["username"]
        ):
            user_to_rsm_visible[normalized_username] = normalized_manager

    asm_townships_source = access.get("asm_townships", {})
    asm_townships_visible: dict[str, dict[str, list[str]]] = {}
    for asm_username, region_map in asm_townships_source.items():
        normalized_asm = normalize_username(asm_username)
        if not normalized_asm or not isinstance(region_map, dict):
            continue
        if principal["role"] != ROLE_OWNER and normalized_asm not in visible_usernames:
            continue
        normalized_region_map: dict[str, list[str]] = {}
        for region, townships in region_map.items():
            normalized_region = normalize_region_token(region)
            if (
                principal["role"] != ROLE_OWNER
                and normalized_region not in set(region_options)
            ):
                continue
            normalized_region_map[normalized_region] = normalize_township_list(townships)
        if normalized_region_map:
            asm_townships_visible[normalized_asm] = normalized_region_map

    return {
        "current_user": user_summary(access, principal["username"]),
        "users": users,
        "regions": region_options,
        "all_regions": principal["regions_universe"],
        "selected_region": selected_region,
        "can_view_all_regions": principal["allow_all_regions"],
        "permissions": {
            "can_upload": principal["role"] in {ROLE_OWNER, ROLE_RSM},
            "can_manage_rsm": principal_can_manage_rsm(principal),
            "can_manage_asm": principal_can_manage_asm(principal),
            "can_manage_files": principal["role"] in {ROLE_OWNER, ROLE_RSM},
        },
        "assignments": {
            "rsm_regions": rsm_regions_visible,
            "user_to_rsm": user_to_rsm_visible,
            "asm_townships": asm_townships_visible,
        },
        "region_townships": visible_region_townships,
    }


def workbook_entry_by_name(entries: list[dict[str, Any]], filename: str) -> dict[str, Any] | None:
    for entry in entries:
        if str(entry["name"]) == filename:
            return entry
    return None


def workbook_entry_by_path(entries: list[dict[str, Any]], path: Path) -> dict[str, Any] | None:
    try:
        target = path.resolve()
    except Exception:
        target = path
    for entry in entries:
        try:
            candidate = Path(entry["path"]).resolve()
        except Exception:
            continue
        if candidate == target:
            return entry
    return None


def is_uploaded_workbook_entry(entry: dict[str, Any]) -> bool:
    try:
        path = Path(entry["path"]).resolve()
    except Exception:
        return False
    try:
        uploads_root = UPLOAD_DIR.resolve()
    except Exception:
        return False
    return uploads_root in path.parents


def validate_sheet_title_input(raw_title: Any) -> tuple[str | None, str | None]:
    title = str(raw_title or "").strip()
    if not title:
        return None, "Sheet name cannot be empty."
    if len(title) > EXCEL_SHEET_TITLE_MAX_LENGTH:
        return None, f"Sheet name must be {EXCEL_SHEET_TITLE_MAX_LENGTH} characters or fewer."
    invalid_found = sorted({ch for ch in title if ch in EXCEL_SHEET_TITLE_INVALID_CHARS})
    if invalid_found:
        invalid_display = " ".join(invalid_found)
        return None, f"Sheet name contains invalid characters: {invalid_display}"
    return title, None


def rename_sheet_in_workbook(path: Path, old_sheet_name: str, new_sheet_name: str) -> tuple[str, str]:
    normalized_old, old_error = validate_sheet_title_input(old_sheet_name)
    if old_error:
        raise ValueError(old_error)

    normalized_new, new_error = validate_sheet_title_input(new_sheet_name)
    if new_error:
        raise ValueError(new_error)

    if normalized_old == normalized_new:
        return normalized_old, normalized_new

    workbook = load_workbook(
        path,
        read_only=False,
        data_only=False,
        keep_links=True,
        keep_vba=path.suffix.lower() in MACRO_ENABLED_EXTENSIONS,
    )
    try:
        if normalized_old not in workbook.sheetnames:
            raise ValueError(f"Unknown sheet name: {normalized_old}")
        if normalized_new in workbook.sheetnames:
            raise ValueError(f"Sheet name already exists: {normalized_new}")

        worksheet = workbook[normalized_old]
        worksheet.title = normalized_new
        workbook.save(path)
    finally:
        workbook.close()

    return normalized_old, normalized_new
