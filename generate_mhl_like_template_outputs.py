#!/usr/bin/env python3
"""Generate MLM outputs with the same sheet set as MHL and 7-MTL templates.

This script focuses on exact sheet names/order compatibility while filling values
from a single DailySales source workbook.
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook

import generate_daily_sales_reports as core

MHL_SHEET_ORDER = [
    "7-MTL",
    "Business Summary",
    "Ws Semi Ws ",
    "Final MTL SKU Wise",
    "MTL Individual",
    "MTL SKU Analysis",
    "Township wise Analysis",
    "Final Town SKU wise Analysis",
    "Outlet Summary",
    "Outlet List ",
    "Way Plan",
    "3-Van Wise SKU",
    "Competition Information",
    "Top 3 or 4 brands in Township",
    "Meiktila",
    "Tharzi",
    "Pyawbwe",
    "Wantwin",
    "Mahlaing",
    "Yamethin",
    "Kyaukpandaung",
    "Taungthar",
    "Myingan",
    "Bagan",
    "Pakokku",
]

TOWNS_MHL = [
    "Meiktila",
    "Tharzi",
    "Pyawbwe",
    "Wantwin",
    "Mahlaing",
    "Yamethin",
    "Kyaukpandaung",
    "Taungthar",
    "Myingan",
    "Bagan",
    "Pakokku",
]

TOWNS_7MTL = [
    "Meiktila",
    "Tharzi",
    "Pyaw Bwe",
    "Want Twin",
    "Mahaling",
    "Yamethin",
    "Kyaukpadaung",
    "Taungthar",
    "Myingyan",
    "Bagan",
    "Pakokku",
]


def month_keys_full_years(rows: list[core.SalesRow]) -> list[tuple[int, int]]:
    years = sorted({r.year for r in rows})
    keys: list[tuple[int, int]] = []
    for y in years:
        for m in range(1, 13):
            keys.append((y, m))
    return keys


def month_label(y: int, m: int) -> str:
    return f"{y}-{m:02d}"


def township_totals(rows: list[core.SalesRow]) -> dict[str, dict[str, float]]:
    totals = defaultdict(lambda: {"bot": 0.0, "lit": 0.0, "amt": 0.0})
    for r in rows:
        totals[r.township]["bot"] += r.sales_bot
        totals[r.township]["lit"] += r.liter
        totals[r.township]["amt"] += r.amount
    return totals


def map_placeholder_towns(rows: list[core.SalesRow], placeholders: list[str]) -> dict[str, str]:
    totals = township_totals(rows)
    actual_sorted = sorted(totals.keys(), key=lambda t: (-totals[t]["lit"], t))
    mapping: dict[str, str] = {}
    for idx, name in enumerate(placeholders):
        mapping[name] = actual_sorted[idx] if idx < len(actual_sorted) else ""
    return mapping


def group_sku_month(rows: list[core.SalesRow]) -> tuple[dict[tuple[Any, ...], dict[str, float]], dict[tuple[Any, ...], dict[str, float]]]:
    sku_total = defaultdict(lambda: {"bot": 0.0, "lit": 0.0, "amt": 0.0, "rows": 0.0})
    sku_month = defaultdict(lambda: {"bot": 0.0, "lit": 0.0, "amt": 0.0})
    for r in rows:
        key = (r.stock_id, r.stock_name, r.ml, r.packing)
        sku_total[key]["bot"] += r.sales_bot
        sku_total[key]["lit"] += r.liter
        sku_total[key]["amt"] += r.amount
        sku_total[key]["rows"] += 1
        mkey = (r.year, r.month)
        sku_month[(key, mkey)]["bot"] += r.sales_bot
        sku_month[(key, mkey)]["lit"] += r.liter
        sku_month[(key, mkey)]["amt"] += r.amount
    return sku_total, sku_month


def group_town_month(rows: list[core.SalesRow]) -> dict[tuple[str, tuple[int, int]], dict[str, float]]:
    town_month = defaultdict(lambda: {"bot": 0.0, "lit": 0.0, "amt": 0.0, "rows": 0.0})
    for r in rows:
        key = (r.township, (r.year, r.month))
        town_month[key]["bot"] += r.sales_bot
        town_month[key]["lit"] += r.liter
        town_month[key]["amt"] += r.amount
        town_month[key]["rows"] += 1
    return town_month


def write_simple_sheet(wb: Workbook, title: str, headers: list[str], data_rows: list[list[Any]], formats: dict[int, str] | None = None) -> None:
    core.write_table_sheet(wb, title, headers, data_rows, number_formats=formats)


def build_mhl_like(
    output_path: Path,
    source_path: Path,
    rows: list[core.SalesRow],
    customer_map: dict[str, dict[str, Any]],
    profile: dict[str, int],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    months = month_keys_full_years(rows)
    sku_total, sku_month = group_sku_month(rows)
    town_month = group_town_month(rows)
    town_map = map_placeholder_towns(rows, TOWNS_MHL)
    town_totals_actual = township_totals(rows)

    sorted_skus = sorted(sku_total.keys(), key=lambda k: (-sku_total[k]["lit"], k[1], k[0]))

    headers_7 = ["Sr.", "StockID", "Product Name", "ML", "Packing"]
    for y, m in months:
        lbl = month_label(y, m)
        headers_7.extend([f"{lbl} Bot", f"{lbl} Lit"])
    headers_7.extend(["Total Bot", "Total Lit", "Total Amount"])
    rows_7: list[list[Any]] = []
    for idx, sku in enumerate(sorted_skus, start=1):
        row = [idx, sku[0], sku[1], sku[2], sku[3]]
        for mk in months:
            cell = sku_month.get((sku, mk), {"bot": 0.0, "lit": 0.0})
            row.extend([cell["bot"], cell["lit"]])
        row.extend([sku_total[sku]["bot"], sku_total[sku]["lit"], sku_total[sku]["amt"]])
        rows_7.append(row)
    write_simple_sheet(wb, "7-MTL", headers_7, rows_7, formats={4: "0.00"})

    bs_headers = ["Year", "Month", "MonthName", "TxnRows", "Outlets", "SKUs", "Townships", "SalesBot", "Liter", "Amount", "WS_Liter", "SemiWS_Liter", "Van_Liter"]
    month_rows: list[list[Any]] = []
    by_month = defaultdict(lambda: {"rows": 0, "customers": set(), "skus": set(), "towns": set(), "bot": 0.0, "lit": 0.0, "amt": 0.0, "ws": 0.0, "semi": 0.0, "van": 0.0})
    for r in rows:
        key = (r.year, r.month)
        b = by_month[key]
        b["rows"] += 1
        b["customers"].add(r.customer_id)
        b["skus"].add((r.stock_id, r.stock_name, r.ml, r.packing))
        b["towns"].add(r.township)
        b["bot"] += r.sales_bot
        b["lit"] += r.liter
        b["amt"] += r.amount
        team_lower = (r.team or "").lower()
        if "semi" in team_lower:
            b["semi"] += r.liter
        elif "van" in team_lower:
            b["van"] += r.liter
        else:
            b["ws"] += r.liter
    for (y, m) in sorted(by_month.keys()):
        b = by_month[(y, m)]
        month_rows.append([y, m, core.month_name(m), b["rows"], len(b["customers"]), len(b["skus"]), len(b["towns"]), b["bot"], b["lit"], b["amt"], b["ws"], b["semi"], b["van"]])
    write_simple_sheet(wb, "Business Summary", bs_headers, month_rows, formats={8: "0.00", 9: "0.00", 10: "#,##0.00"})

    semi_headers = ["CustomerID", "CustomerName", "Township", "Team"] + [f"{month_label(y,m)} Lit" for y,m in months] + ["Total Lit", "Total Amount"]
    semi_rows = defaultdict(lambda: {"vals": defaultdict(float), "lit": 0.0, "amt": 0.0, "name": "", "town": "", "team": ""})
    for r in rows:
        if "semi" not in (r.team or "").lower():
            continue
        b = semi_rows[r.customer_id]
        b["name"] = r.customer_name
        b["town"] = r.township
        b["team"] = r.team
        b["vals"][(r.year, r.month)] += r.liter
        b["lit"] += r.liter
        b["amt"] += r.amount
    ws_rows = []
    for cid, b in sorted(semi_rows.items(), key=lambda item: (-item[1]["lit"], item[0])):
        row = [cid, b["name"], b["town"], b["team"]]
        for mk in months:
            row.append(b["vals"].get(mk, 0.0))
        row.extend([b["lit"], b["amt"]])
        ws_rows.append(row)
    write_simple_sheet(wb, "Ws Semi Ws ", semi_headers, ws_rows, formats={len(semi_headers)-1: "#,##0.00"})

    final_sku_headers = ["StockID", "StockName", "ML", "Packing", "TxnRows", "SalesBot", "Liter", "Amount", "LiterContribution", "AmountContribution"]
    grand_lit = sum(v["lit"] for v in sku_total.values())
    grand_amt = sum(v["amt"] for v in sku_total.values())
    final_sku_rows = []
    for sku in sorted_skus:
        v = sku_total[sku]
        final_sku_rows.append([sku[0], sku[1], sku[2], sku[3], int(v["rows"]), v["bot"], v["lit"], v["amt"], (v["lit"]/grand_lit if grand_lit else 0.0), (v["amt"]/grand_amt if grand_amt else 0.0)])
    write_simple_sheet(wb, "Final MTL SKU Wise", final_sku_headers, final_sku_rows, formats={3: "0.00", 9: "0.00%", 10: "0.00%"})

    indiv_headers = ["CustomerID", "CustomerName", "Township", "Team", "TxnRows", "SalesBot", "Liter", "Amount"]
    indiv = defaultdict(lambda: {"name": "", "town": "", "team": "", "rows": 0, "bot": 0.0, "lit": 0.0, "amt": 0.0})
    for r in rows:
        b = indiv[r.customer_id]
        b["name"] = r.customer_name
        b["town"] = r.township
        b["team"] = r.team
        b["rows"] += 1
        b["bot"] += r.sales_bot
        b["lit"] += r.liter
        b["amt"] += r.amount
    indiv_rows = []
    for cid, b in sorted(indiv.items(), key=lambda item: (-item[1]["lit"], item[0])):
        indiv_rows.append([cid, b["name"], b["town"], b["team"], b["rows"], b["bot"], b["lit"], b["amt"]])
    write_simple_sheet(wb, "MTL Individual", indiv_headers, indiv_rows, formats={7: "0.00", 8: "#,##0.00"})

    mtl_sku_headers = ["Year", "Month", "MonthName", "StockID", "StockName", "ML", "Packing", "SalesBot", "Liter", "Amount"]
    mtl_sku_rows = []
    for (sku, mk), v in sorted(sku_month.items(), key=lambda item: (item[0][1][0], item[0][1][1], -item[1]["lit"], item[0][0][1])):
        y, m = mk
        stock_id, stock_name, ml, pack = sku
        mtl_sku_rows.append([y, m, core.month_name(m), stock_id, stock_name, ml, pack, v["bot"], v["lit"], v["amt"]])
    write_simple_sheet(wb, "MTL SKU Analysis", mtl_sku_headers, mtl_sku_rows, formats={6: "0.00", 9: "0.00", 10: "#,##0.00"})

    tw_headers = ["No", "TownShip", "ActualTownship"]
    for y, m in months:
        tw_headers.extend([f"{month_label(y,m)} Bot", f"{month_label(y,m)} Lit"])
    tw_headers.extend(["Total Bot", "Total Lit", "Total Amount"])
    tw_rows = []
    for idx, placeholder in enumerate(TOWNS_MHL, start=1):
        actual = town_map.get(placeholder, "")
        row = [idx, placeholder, actual]
        for mk in months:
            v = town_month.get((actual, mk), {"bot": 0.0, "lit": 0.0}) if actual else {"bot": 0.0, "lit": 0.0}
            row.extend([v["bot"], v["lit"]])
        totals = town_totals_actual.get(actual, {"bot": 0.0, "lit": 0.0, "amt": 0.0}) if actual else {"bot": 0.0, "lit": 0.0, "amt": 0.0}
        row.extend([totals["bot"], totals["lit"], totals["amt"]])
        tw_rows.append(row)
    write_simple_sheet(wb, "Township wise Analysis", tw_headers, tw_rows, formats={len(tw_headers)-2: "0.00", len(tw_headers)-1: "#,##0.00"})

    town_sku = defaultdict(lambda: {"bot": 0.0, "lit": 0.0, "amt": 0.0, "rows": 0})
    for r in rows:
        town_sku[(r.township, r.stock_id, r.stock_name, r.ml, r.packing)]["bot"] += r.sales_bot
        town_sku[(r.township, r.stock_id, r.stock_name, r.ml, r.packing)]["lit"] += r.liter
        town_sku[(r.township, r.stock_id, r.stock_name, r.ml, r.packing)]["amt"] += r.amount
        town_sku[(r.township, r.stock_id, r.stock_name, r.ml, r.packing)]["rows"] += 1

    ft_headers = ["PlaceholderTownship", "ActualTownship", "StockID", "StockName", "ML", "Packing", "TxnRows", "SalesBot", "Liter", "Amount", "TownshipLiterContribution"]
    ft_rows = []
    for placeholder in TOWNS_MHL:
        actual = town_map.get(placeholder, "")
        denom = town_totals_actual.get(actual, {"lit": 0.0})["lit"] if actual else 0.0
        actual_keys = [k for k in town_sku.keys() if k[0] == actual]
        actual_keys.sort(key=lambda k: (-town_sku[k]["lit"], k[2]))
        for key in actual_keys:
            _, sid, sname, ml, pack = key
            v = town_sku[key]
            ft_rows.append([placeholder, actual, sid, sname, ml, pack, v["rows"], v["bot"], v["lit"], v["amt"], (v["lit"]/denom if denom else 0.0)])
    write_simple_sheet(wb, "Final Town SKU wise Analysis", ft_headers, ft_rows, formats={5: "0.00", 10: "#,##0.00", 11: "0.00%"})

    outlet_headers = ["CustomerID", "CustomerName", "Township", "Team", "TxnRows", "Liter", "Amount"]
    outlet_rows = []
    for cid, b in sorted(indiv.items(), key=lambda item: (-item[1]["lit"], item[0])):
        outlet_rows.append([cid, b["name"], b["town"], b["team"], b["rows"], b["lit"], b["amt"]])
    write_simple_sheet(wb, "Outlet Summary", outlet_headers, outlet_rows, formats={6: "0.00", 7: "#,##0.00"})

    outlet_list_headers = ["CustomerID", "CustomerName", "Address", "Township", "Team", "ActiveInPeriod"]
    outlet_list_rows = []
    for cid, ref in sorted(customer_map.items(), key=lambda item: (item[1].get("township", ""), item[1].get("customer_name", ""), item[0])):
        outlet_list_rows.append([cid, ref.get("customer_name", ""), ref.get("address", ""), ref.get("township", ""), ref.get("team", ""), "Y" if cid in indiv else "N"])
    write_simple_sheet(wb, "Outlet List ", outlet_list_headers, outlet_list_rows)

    way_plan = defaultdict(lambda: {"rows": 0, "customers": set(), "lit": 0.0})
    for r in rows:
        way_plan[(r.car_no or "(blank)", r.township)]["rows"] += 1
        way_plan[(r.car_no or "(blank)", r.township)]["customers"].add(r.customer_id)
        way_plan[(r.car_no or "(blank)", r.township)]["lit"] += r.liter
    way_headers = ["CarNo", "Township", "TxnRows", "Outlets", "Liter"]
    way_rows = []
    for key, v in sorted(way_plan.items(), key=lambda item: (-item[1]["lit"], item[0][0], item[0][1])):
        way_rows.append([key[0], key[1], v["rows"], len(v["customers"]), v["lit"]])
    write_simple_sheet(wb, "Way Plan", way_headers, way_rows, formats={5: "0.00"})

    van_sku = defaultdict(lambda: {"bot": 0.0, "lit": 0.0, "amt": 0.0})
    for r in rows:
        van_sku[(r.car_no or "(blank)", r.stock_id, r.stock_name, r.ml, r.packing)]["bot"] += r.sales_bot
        van_sku[(r.car_no or "(blank)", r.stock_id, r.stock_name, r.ml, r.packing)]["lit"] += r.liter
        van_sku[(r.car_no or "(blank)", r.stock_id, r.stock_name, r.ml, r.packing)]["amt"] += r.amount
    van_headers = ["CarNo", "StockID", "StockName", "ML", "Packing", "SalesBot", "Liter", "Amount"]
    van_rows = []
    for key, v in sorted(van_sku.items(), key=lambda item: (-item[1]["lit"], item[0][0], item[0][2])):
        van_rows.append([key[0], key[1], key[2], key[3], key[4], v["bot"], v["lit"], v["amt"]])
    write_simple_sheet(wb, "3-Van Wise SKU", van_headers, van_rows, formats={4: "0.00", 7: "0.00", 8: "#,##0.00"})

    comp_headers = ["Field", "Value"]
    comp_rows = [
        ["Data Source", source_path.name],
        ["Generated", datetime.now().isoformat(timespec="seconds")],
        ["Rows Used", profile.get("kept_rows", 0)],
        ["Scope", "DailySales derived only"],
        ["Note", "Competition metrics are not available in source workbook."],
        ["Township Mapping", "placeholder -> actual"],
    ]
    for placeholder in TOWNS_MHL:
        comp_rows.append([placeholder, town_map.get(placeholder, "")])
    write_simple_sheet(wb, "Competition Information", comp_headers, comp_rows)

    top_headers = ["PlaceholderTownship", "ActualTownship", "Rank", "StockID", "StockName", "Liter", "Amount"]
    top_rows = []
    for placeholder in TOWNS_MHL:
        actual = town_map.get(placeholder, "")
        keys = [k for k in town_sku if k[0] == actual]
        keys.sort(key=lambda k: (-town_sku[k]["lit"], k[2]))
        for rank, key in enumerate(keys[:4], start=1):
            _, sid, sname, _, _ = key
            v = town_sku[key]
            top_rows.append([placeholder, actual, rank, sid, sname, v["lit"], v["amt"]])
    write_simple_sheet(wb, "Top 3 or 4 brands in Township", top_headers, top_rows, formats={6: "0.00", 7: "#,##0.00"})

    for sheet_name in TOWNS_MHL:
        actual = town_map.get(sheet_name, "")
        headers = ["Sr", "ActualTownship", "StockID", "Product Name", "ML", "Packing"]
        for y, m in months:
            lbl = month_label(y, m)
            headers.extend([f"{lbl} Bot", f"{lbl} Lit"])
        headers.extend(["Total Bot", "Total Lit", "Total Amount"])

        keys = [k for k in sorted_skus]
        rows_out = []
        for idx, sku in enumerate(keys, start=1):
            sid, sname, ml, pack = sku
            row = [idx, actual, sid, sname, ml, pack]
            total_bot = total_lit = total_amt = 0.0
            for mk in months:
                if actual:
                    v = town_month.get((actual, mk), None)
                    sku_v = 0.0
                    sku_l = 0.0
                    # derive via direct scan for precision
                    for r in rows:
                        if r.township == actual and r.stock_id == sid and (r.year, r.month) == mk:
                            sku_v += r.sales_bot
                            sku_l += r.liter
                    row.extend([sku_v, sku_l])
                    total_bot += sku_v
                    total_lit += sku_l
                else:
                    row.extend([0.0, 0.0])
            if actual:
                for r in rows:
                    if r.township == actual and r.stock_id == sid:
                        total_amt += r.amount
            row.extend([total_bot, total_lit, total_amt])
            if total_lit > 0:
                rows_out.append(row)
        write_simple_sheet(wb, sheet_name, headers, rows_out, formats={5: "0.00", len(headers)-2: "0.00", len(headers)-1: "#,##0.00"})

    for name in MHL_SHEET_ORDER:
        if name not in wb.sheetnames:
            wb.create_sheet(name)

    wb._sheets = [wb[name] for name in MHL_SHEET_ORDER]
    wb.save(output_path)


def build_7mtl_like(
    output_path: Path,
    source_path: Path,
    rows: list[core.SalesRow],
    profile: dict[str, int],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    months = month_keys_full_years(rows)
    town_map = map_placeholder_towns(rows, TOWNS_7MTL)
    town_month = group_town_month(rows)

    h = ["No", "TownShip", "ActualTownship"]
    for y, m in months:
        h.extend([f"{month_label(y,m)} Bot", f"{month_label(y,m)} Lit"])
    h.extend(["Total Bot", "Total Lit", "Total Amount"])
    sum_rows = []
    town_tot = township_totals(rows)
    for idx, placeholder in enumerate(TOWNS_7MTL, start=1):
        actual = town_map.get(placeholder, "")
        row = [idx, placeholder, actual]
        for mk in months:
            v = town_month.get((actual, mk), {"bot": 0.0, "lit": 0.0}) if actual else {"bot": 0.0, "lit": 0.0}
            row.extend([v["bot"], v["lit"]])
        t = town_tot.get(actual, {"bot": 0.0, "lit": 0.0, "amt": 0.0}) if actual else {"bot": 0.0, "lit": 0.0, "amt": 0.0}
        row.extend([t["bot"], t["lit"], t["amt"]])
        sum_rows.append(row)
    write_simple_sheet(wb, "7-MTL", h, sum_rows, formats={len(h)-2: "0.00", len(h)-1: "#,##0.00"})

    sku_cache = defaultdict(
        lambda: defaultdict(lambda: defaultdict(lambda: {"bot": 0.0, "lit": 0.0, "amt": 0.0}))
    )
    for r in rows:
        sku_key = (r.stock_id, r.stock_name, r.ml, r.packing)
        sku_cache[r.township][sku_key][(r.year, r.month)]["bot"] += r.sales_bot
        sku_cache[r.township][sku_key][(r.year, r.month)]["lit"] += r.liter
        sku_cache[r.township][sku_key][(r.year, r.month)]["amt"] += r.amount

    for placeholder in TOWNS_7MTL:
        actual = town_map.get(placeholder, "")
        headers = ["Sr", "ActualTownship", "StockID", "Product Name", "ML", "Packing"]
        for y, m in months:
            headers.extend([f"{month_label(y,m)} Bot", f"{month_label(y,m)} Lit"])
        headers.extend(["Total Bot", "Total Lit", "Total Amount"])

        rows_out = []
        if actual and actual in sku_cache:
            sku_keys = sorted(sku_cache[actual].keys(), key=lambda k: (-sum(v["lit"] for v in sku_cache[actual][k].values()), k[1], k[0]))
            for idx, sku in enumerate(sku_keys, start=1):
                sid, sname, ml, pack = sku
                row = [idx, actual, sid, sname, ml, pack]
                tb = tl = ta = 0.0
                for mk in months:
                    v = sku_cache[actual][sku].get(mk, {"bot": 0.0, "lit": 0.0, "amt": 0.0})
                    row.extend([v["bot"], v["lit"]])
                    tb += v["bot"]
                    tl += v["lit"]
                    ta += v["amt"]
                row.extend([tb, tl, ta])
                rows_out.append(row)
        write_simple_sheet(wb, placeholder, headers, rows_out, formats={5: "0.00", len(headers)-2: "0.00", len(headers)-1: "#,##0.00"})

    order = ["7-MTL"] + TOWNS_7MTL
    for name in order:
        if name not in wb.sheetnames:
            wb.create_sheet(name)
    wb._sheets = [wb[name] for name in order]
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate outputs with the same sheet set as MHL/7-MTL templates.")
    parser.add_argument("--source", required=True, help="Source workbook path")
    parser.add_argument("--output-dir", default=".", help="Output directory")
    parser.add_argument("--mhl-name", default="MLM 2026 Feb.xlsx", help="Output filename for MHL-like workbook")
    parser.add_argument("--town-name", default="7-MLM for Township Summary_4.xlsx", help="Output filename for 7-MTL-like workbook")
    args = parser.parse_args()

    source_path = Path(args.source).expanduser().resolve()
    out_dir = Path(args.output_dir).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    stock_map, customer_map = core.load_masters(source_path)
    rows, profile = core.extract_sales_rows(source_path, stock_map, customer_map)

    mhl_path = out_dir / args.mhl_name
    town_path = out_dir / args.town_name

    build_mhl_like(mhl_path, source_path, rows, customer_map, profile)
    build_7mtl_like(town_path, source_path, rows, profile)

    print(f"Generated: {mhl_path}")
    print(f"Generated: {town_path}")


if __name__ == "__main__":
    main()
